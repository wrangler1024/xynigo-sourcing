# -*- coding: utf-8 -*-
"""Release contract tests for Xynigo Sourcing v0.13.10 candidate."""

from pathlib import Path
import hashlib
import json
import subprocess
import sys
import tempfile
import unittest

from purchase_tool import __version__


class ReleaseV01310Tests(unittest.TestCase):
    def test_version_and_packaging_are_aligned(self):
        root = Path(__file__).resolve().parents[1]
        self.assertEqual(__version__, '0.13.10')
        pyproject = (root / 'pyproject.toml').read_text(encoding='utf-8')
        self.assertIn('version = "0.13.10"', pyproject)
        cloud_project = (root / 'cloud' / 'auth-service' /
                         'pyproject.toml').read_text(encoding='utf-8')
        self.assertIn('version = "0.13.10"', cloud_project)
        cloud_init = (root / 'cloud' / 'auth-service' / 'src' /
                      'xynigo_auth' / '__init__.py').read_text(
                          encoding='utf-8')
        self.assertIn('__version__ = "0.13.10"', cloud_init)
        cloud_main = (root / 'cloud' / 'auth-service' / 'src' /
                      'xynigo_auth' / 'main.py').read_text(encoding='utf-8')
        self.assertIn('version="0.13.10"', cloud_main)
        self.assertTrue((root / 'release' / 'v0.13.10.zh-CN.json').is_file())
        self.assertTrue((root / 'release' / 'v0.13.10.zh-CN.md').is_file())
        script = (root / '组装Windows绿色包.sh').read_text(encoding='utf-8')
        self.assertIn('v${VERSION}${BUILD_SUFFIX}.zip', script)
        self.assertIn('XYNIGO_BUILD_LABEL', script)
        self.assertIn('XYNIGO_RELEASE_CHANNEL:-stable', script)
        self.assertIn('--channel "$CHANNEL"', script)
        self.assertIn("'channel': channel", script)
        self.assertIn('Xynigo Sourcing v%s 启动中', script)
        self.assertIn("'启动-本地执行器.bat'", script)
        self.assertIn('run.py --local-ui', script)
        self.assertIn("os.environ.setdefault('XYNIGO_INSTALL_DIR', ROOT)", script)
        self.assertNotIn(
            'from purchase_tool.updater import check_for_updates_at_startup',
            script)
        self.assertTrue((root / 'packaging' / 'windows' /
                         'update-helper.ps1').is_file())
        mac_script = (root / '组装macOS绿色包.sh').read_text(
            encoding='utf-8')
        self.assertIn('Xynigo_Sourcing_macOS_', mac_script)
        self.assertIn('XYNIGO_RELEASE_CHANNEL:-stable', mac_script)
        self.assertIn('--channel "$CHANNEL"', mac_script)
        self.assertIn("'channel': channel", mac_script)
        self.assertIn("'启动-本地执行器-Mac.command'", mac_script)
        self.assertIn('xynigo-sourcing --local-ui', mac_script)
        mac_entry = (root / 'packaging' / 'macos' / 'entry.py').read_text(
            encoding='utf-8')
        self.assertIn("os.environ.setdefault('XYNIGO_INSTALL_DIR'", mac_entry)
        self.assertNotIn('check_for_updates_at_startup', mac_entry)
        self.assertTrue((root / 'packaging' / 'macos' /
                         'update-helper.sh').is_file())
        self.assertTrue((root / 'src' / 'purchase_tool' /
                         'updater.py').is_file())

    def test_release_asset_builder_supports_test_channel(self):
        root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            asset = temp / 'Xynigo_Sourcing_test.zip'
            asset.write_bytes(b'xynigo-v0.13.10-test-package')
            manifest = temp / 'latest.json'
            sha_file = temp / 'SHA256SUMS.txt'
            subprocess.run(
                [
                    sys.executable,
                    str(root / 'scripts' / 'update_release_assets.py'),
                    '--version', '0.13.10',
                    '--channel', 'test',
                    '--platform', 'macos-arm64',
                    '--asset', str(asset),
                    '--notes', str(root / 'release' /
                                   'v0.13.10.zh-CN.json'),
                    '--manifest', str(manifest),
                    '--sha-file', str(sha_file),
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            payload = json.loads(manifest.read_text(encoding='utf-8'))
            digest = hashlib.sha256(asset.read_bytes()).hexdigest()
            self.assertEqual(payload['schemaVersion'], 2)
            self.assertEqual(payload['channel'], 'test')
            self.assertEqual(payload['version'], '0.13.10')
            self.assertEqual(payload['platforms']['macos-arm64']['sha256'],
                             digest)
            self.assertEqual(
                sha_file.read_text(encoding='utf-8'),
                '%s  %s\n' % (digest, asset.name),
            )

    def test_rebuilding_sole_platform_refreshes_legacy_manifest_hash(self):
        root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            asset = temp / 'Xynigo_Sourcing_macOS_test.zip'
            manifest = temp / 'latest.json'
            sha_file = temp / 'SHA256SUMS.txt'
            command = [
                sys.executable,
                str(root / 'scripts' / 'update_release_assets.py'),
                '--version', '0.13.10',
                '--channel', 'test',
                '--platform', 'macos-arm64',
                '--asset', str(asset),
                '--notes', str(root / 'release' / 'v0.13.10.zh-CN.json'),
                '--manifest', str(manifest),
                '--sha-file', str(sha_file),
            ]
            asset.write_bytes(b'first-macos-build')
            subprocess.run(command, check=True, capture_output=True, text=True)
            asset.write_bytes(b'second-macos-build')
            subprocess.run(command, check=True, capture_output=True, text=True)

            payload = json.loads(manifest.read_text(encoding='utf-8'))
            current = payload['platforms']['macos-arm64']
            self.assertEqual(payload['assetName'], current['assetName'])
            self.assertEqual(payload['sha256'], current['sha256'])
            self.assertEqual(payload['size'], current['size'])

    def test_web_bundle_contains_module_three_and_template(self):
        root = Path(__file__).resolve().parents[1]
        html = (root / 'src' / 'purchase_tool' / 'web' / 'index.html').read_text(
            encoding='utf-8')
        self.assertIn('订单物流查询', html)
        self.assertNotIn('物流订单查询', html)
        self.assertIn('买家号建环境', html)
        self.assertIn('/api/envbatch/start', html)
        self.assertIn('/api/tasks', html)
        self.assertIn(
            'async function pollTaskStatus() {\n  if (!authReady) return;',
            html,
        )
        self.assertIn('id="cfgSafeParallel"', html)
        self.assertIn('开启安全并行', html)
        self.assertIn('/api/envbatch/preflight', html)
        self.assertIn('id="envSiteGroup"', html)
        self.assertNotIn('id="envSiteGroup" disabled', html)
        self.assertIn('async function loadEnvGroups(options={})', html)
        self.assertIn('id="envSite"', html)
        self.assertIn('if (group) body.purchaseTags = {[site]:group};', html)
        self.assertIn("api('/api/envbatch/preferences'", html)
        self.assertIn('site:$(\'envSite\').value', html)
        self.assertIn('option.textContent = group', html)
        self.assertNotIn('option.textContent = `墨西哥站 · ${group}`', html)
        # 模块三：模式切换 + 采购员卡片 + 备用·测试环境（2026-08-19 设计定稿）
        self.assertIn('id="envModeBar"', html)
        self.assertIn('data-mode="bound"', html)
        self.assertIn('data-mode="backup"', html)
        self.assertIn('id="envCardParse"', html)
        self.assertIn('id="envAssignBound"', html)
        self.assertIn('id="envAssignBackup"', html)
        self.assertIn('id="envBuyerGrid"', html)
        self.assertIn('id="envBackupBuyerGrid"', html)
        self.assertIn('id="envAssignSum"', html)
        self.assertIn('id="envSpecOut"', html)
        self.assertIn('id="btnEnvEven"', html)
        self.assertIn('id="envBackupCount"', html)
        self.assertIn('name="envBackupType"', html)
        self.assertIn('id="envBackupPattern"', html)
        self.assertIn('id="envBoundExports"', html)
        self.assertIn('id="envBackupExports"', html)
        self.assertIn('id="btnEnvBackupResult"', html)
        self.assertNotIn('下载飞书直贴 TSV（无表头/含凭证）', html)
        self.assertNotIn('从第一空行的「站点」列粘贴', html)
        self.assertIn('旧表直贴已停用', html)
        self.assertIn('composeAssignmentSpec()', html)
        self.assertIn('/api/envbatch/backup/preview', html)
        self.assertIn('/api/envbatch/backup/start', html)
        self.assertIn('/api/envbatch/backup/progress', html)
        self.assertIn('/api/envbatch/backup/result', html)
        self.assertIn('function setEnvMode(next)', html)
        self.assertIn('function renderBackupBatch(snap)', html)
        # 分组默认选中站点标准分组 + 日期控件（2026-08-19 Jeff 指示）
        self.assertIn('type="date" id="envDate"', html)
        self.assertIn('SITE_DEFAULT_GROUPS', html)
        self.assertIn('function envGroupCompatibleWithSite(groupName, site)', html)
        self.assertIn("MX: '希音墨西哥采购'", html)
        self.assertIn("US: '美国采购分组'", html)
        self.assertIn('function envPurchaseDate()', html)
        # 采购分组下拉过滤店铺分组（2026-08-20 Jeff 指示）
        self.assertIn('BUYER_GROUP_PATTERN', html)
        self.assertIn('/采购|买家号|Registration/i', html)
        self.assertIn(".filter(name => BUYER_GROUP_PATTERN.test(name));", html)
        self.assertIn('— 暂无采购/买家号环境分组 —', html)
        self.assertNotIn('id="envAssign"', html)
        self.assertNotIn('assignmentTotal', html)
        # ④ 批量进度：限高内部滚动 + 吸顶表头（大批量不拉长页面）
        self.assertIn('id="envTableWrap"', html)
        self.assertIn('#envTableWrap { max-height: 480px; overflow-y: auto; }', html)
        self.assertIn('#envThead th { position: sticky; top: 0;', html)
        self.assertIn('id="querySite"', html)
        self.assertIn('<option value="US">美国站 · us.shein.com</option>', html)
        self.assertIn('const payload = {serials, site};', html)
        self.assertIn('body: JSON.stringify(payload)', html)
        self.assertIn("'Shipped': 'st-enviado'", html)
        self.assertIn("'Paid': 'st-procesando'", html)
        self.assertIn("'Risk verification': 'st-reembolsando'", html)
        self.assertIn('风险订单，待验证', html)
        self.assertIn('风险订单 <b id="cntRisk">0</b>', html)
        self.assertIn("if (r.riskOrder) cnt.risk++;", html)
        self.assertIn("else cnt.ok++;", html)
        self.assertIn("r.state === 'ok' && !r.riskOrder", html)
        self.assertIn("r.state === 'ok' && r.riskOrder", html)
        self.assertIn('id="cfgPurchaseTagMX"', html)
        self.assertIn('id="cfgPurchaseTagUS"', html)
        self.assertIn('id="cfgProxyLink"', html)
        self.assertIn('type="password" id="cfgProxyLink"', html)
        self.assertIn('id="cfgProxyClear"', html)
        self.assertIn('proxyLink, proxyClear', html)
        self.assertIn('可选：自定义链接；留空保持现状', html)
        self.assertIn("cfg.proxySource === 'custom'", html)
        self.assertIn('系统模板固定带表头', html)
        self.assertNotIn('https://proxy.example.test', html)
        self.assertIn('Xynigo Sourcing v0.13.10', html)
        self.assertIn('累计修复测试 v0.13.10', html)
        self.assertIn('测试环境 · 数据隔离', html)
        self.assertNotIn('本机数据不出站', html)
        self.assertIn('Xyni, GO!', html)
        self.assertIn('Xynigo 品牌字标', html)
        self.assertIn('小犀与 Xynigo 完整品牌一体图形', html)
        self.assertIn('src="xynigo-logo.png"', html)
        self.assertIn('src="xynigo-logo.png?v=6"', html)
        self.assertNotIn('src="xynigo-mascot-x.png"', html)
        self.assertIn('跨境采购协同系统', html)
        self.assertNotIn('跨境代采协同系统', html)
        self.assertIn('src="xynigo-x.ico?v=3"', html)
        self.assertIn('href="/xynigo-x.png?v=6"', html)
        self.assertIn('href="/favicon.ico?v=6"', html)
        self.assertNotIn('class="logo-x"', html)
        self.assertIn('小犀提示', html)
        self.assertIn('品牌表达', html)
        self.assertIn('持续迭代中', html)
        # 新系统 UI 第一阶段：一级产品层级 + 左侧可展开二级导航，现有功能入口保持不变。
        for primary in (
                'workbench', 'procurement', 'fulfillment', 'operations',
                'finance', 'resources', 'assistant', 'analytics', 'system'):
            self.assertIn('data-primary="%s"' % primary, html)
        for label in (
                '工作台', '采购中心', '履约追踪', '运营中心',
                '财务中心', '资源中心', '小犀助手', '数据分析',
                '系统管理'):
            self.assertIn('<b>%s</b>' % label, html)
        self.assertIn('id="secondaryTabs"', html)
        self.assertLess(html.index('id="secondaryTabs"'), html.index('</aside>'))
        self.assertNotIn('secondary-nav-shell', html)
        self.assertIn("primaryButton.insertAdjacentElement('afterend', $('secondaryTabs'))", html)
        self.assertIn('body.sidebar-collapsed .secondary-tabs { display: none; }', html)
        self.assertIn('data-parent="fulfillment" data-module="query"', html)
        self.assertIn(
            'data-parent="procurement" data-module="procurementorders"', html)
        self.assertIn(
            'data-parent="procurement" data-module="procurementexecution"', html)
        self.assertIn(
            '>采购任务<span class="secondary-status">云端</span>', html)
        self.assertIn('data-parent="resources" data-module="buyerlib"', html)
        self.assertIn('data-parent="resources" data-module="vendorimport"', html)
        self.assertIn('data-parent="resources" data-module="register"', html)
        self.assertIn('data-parent="resources" data-module="envbatch"', html)
        self.assertIn('data-parent="resources" data-module="storemanagement"', html)
        self.assertIn('data-parent="resources" data-module="proxymanagement"', html)
        for label in ('买家号库', '号商入库', '账号注册', '环境创建', '店铺管理', '代理IP'):
            self.assertIn('>%s<span class="secondary-status">' % label, html)
        self.assertIn('账号 · 店铺 · IP · 环境', html)
        for parent, label in (
                ('operations', '运营任务'), ('finance', '应付管理')):
            self.assertIn('data-parent="%s" data-planned="%s"' % (parent, label), html)
        self.assertIn(
            'data-parent="assistant" data-module="procurementimport"', html)
        self.assertIn("defaultModule: 'procurementimport'", html)
        self.assertIn('id="procurementImportPanel"', html)
        self.assertIn("$('secondaryTabs').hidden = visibleSecondaryCount === 0", html)
        self.assertIn('id="buyerLibraryPanel"', html)
        self.assertIn('id="vendorImportPanel"', html)
        self.assertIn('/api/buyer-library', html)
        self.assertIn('后台无头验证全部新建环境出口 IP', html)
        self.assertIn('不会打开可见的 HubStudio 环境窗口', html)
        self.assertIn('环境创建完成，正在检测出口 IP', html)
        self.assertIn("snap.phase === 'ip_checking'", html)
        self.assertIn('data-parent="system" data-module="localsettings"', html)
        self.assertIn(
            'data-parent="system" data-module="larkconnection" '
            'data-local-only="1" data-required-role="super_admin"', html)
        self.assertIn('data-module="localsettings" data-local-only="1"', html)
        self.assertIn(
            '>本机设置<span class="secondary-status">本机兼容</span>', html)
        self.assertIn('>飞书连接<span class="secondary-status restricted">本机超管</span>', html)
        self.assertIn(
            'data-parent="system" data-planned="云端服务配置"', html)
        self.assertIn(
            '>云端服务配置<span class="secondary-status restricted">超管</span>', html)
        for module, label in (
                ('organizationaccess', '组织与权限'),
                ('sessionmanagement', '登录会话')):
            self.assertIn(
                'data-parent="system" data-module="%s"' % module, html)
            self.assertIn(
                '>%s<span class="secondary-status">云端</span>' % label, html)
        self.assertNotIn('data-parent="system" data-module="membermanagement"', html)
        self.assertNotIn('data-parent="system" data-module="rolemanagement"', html)
        self.assertIn('id="planningPanel"', html)
        self.assertIn('功能规划中 · 不影响现有模块', html)
        self.assertIn('function setPrimaryModule(primary)', html)
        self.assertIn('function setFeaturePanel(module)', html)
        self.assertIn('id="workspaceTabs"', html)
        self.assertIn('aria-label="已打开业务页面"', html)
        self.assertIn("let openFeatureTabs = ['query']", html)
        self.assertIn('function openFeatureTab(module)', html)
        self.assertIn('function activateFeatureTab(module)', html)
        self.assertIn('function closeFeatureTab(module)', html)
        self.assertIn('MAX_OPEN_FEATURE_TABS = 10', html)
        self.assertIn('id="workspaceTabScrollLeft"', html)
        self.assertIn('id="workspaceTabScrollRight"', html)
        self.assertIn('id="workspaceTabListToggle"', html)
        self.assertIn('id="workspaceTabMenuList"', html)

        self.assertIn('function updateWorkspaceTabControls()', html)
        self.assertIn('function closeOtherFeatureTabs()', html)
        self.assertIn('function closeRightFeatureTabs()', html)
        self.assertIn("btn.onclick = () => openFeatureTab(btn.dataset.module)", html)
        self.assertNotIn('class="breadcrumb"', html)
        self.assertIn('id="sidebarToggle"', html)
        self.assertIn('id="btnRetryFail"', html)
        self.assertNotIn('id="updateNotice"', html)
        self.assertNotIn('id="updateCheck"', html)
        self.assertNotIn('/api/update/check', html)
        self.assertNotIn('/api/update/status', html)
        self.assertNotIn('/api/update/prompt', html)
        self.assertIn("location.protocol === 'file:'", html)
        self.assertIn('当前页面是本地文件预览', html)
        self.assertIn('无法连接 Xynigo 本地服务', html)
        self.assertIn('cfgHideEnvName', html)
        self.assertIn("activeModule !== 'query'", html)
        self.assertIn('data-module="localsettings"', html)
        self.assertIn('data-module="larkconnection"', html)
        self.assertIn('id="settingsPanel"', html)
        self.assertIn('data-settings-view="localsettings"', html)
        self.assertIn('data-settings-view="larkconnection"', html)
        self.assertIn('data-settings-view="organizationaccess"', html)
        self.assertIn('data-settings-view="sessionmanagement"', html)
        for view, label in (
                ('members', '成员管理'),
                ('roles', '角色管理'),
                ('policies', '权限策略')):
            self.assertIn('data-organization-view="%s"' % view, html)
            self.assertIn('>%s</button>' % label, html)
            self.assertIn('data-organization-panel="%s"' % view, html)
        self.assertIn(
            'data-organization-view="data-scope" data-planned="数据范围"', html)
        self.assertIn(
            'disabled>数据范围<span class="secondary-status planned">规划</span>', html)
        self.assertNotIn('data-organization-panel="data-scope"', html)
        self.assertIn(
            "requiredPermissionsAny: ['system.member.manage', 'system.role.manage']", html)
        self.assertIn('/api/admin/members', html)
        self.assertIn('/api/admin/members/invitations/resolve', html)
        self.assertIn('/api/admin/members/invitations', html)
        self.assertIn('/api/admin/roles', html)
        self.assertIn('/api/admin/sessions', html)
        self.assertIn('id="btnRoleCreate"', html)
        self.assertIn('id="btnMemberInvite"', html)
        self.assertIn('id="memberInviteEditor"', html)
        self.assertIn('id="memberInviteMobile"', html)
        self.assertIn('function resolveMemberInvitation()', html)
        self.assertIn('function createMemberInvitation()', html)
        self.assertIn("window.open('about:blank', 'xynigo-feishu-login')", html)
        self.assertIn('function closeAuthLoginWindow()', html)
        self.assertIn('closeAuthLoginWindow();\n  authIdentity = identity;', html)
        self.assertIn('if (state.loginPending)', html)
        self.assertIn('正在恢复登录状态', html)
        self.assertIn('if (authLoginStarting) return;', html)
        self.assertIn('started.resumed', html)
        self.assertIn('新增后状态固定为 pending', html)
        self.assertIn('不写入 Xynigo 数据库或审计日志', html)
        self.assertIn('id="roleEditor"', html)
        self.assertIn('function saveRoleDefinition()', html)
        self.assertIn('function deleteRoleDefinition(role)', html)
        self.assertIn('/rename`, {name}', html)
        self.assertIn('/delete`);', html)
        self.assertIn('已分配成员的角色必须先解除授权才能删除', html)
        self.assertIn('const PERMISSION_MENU_GROUPS = [', html)
        for primary in ('履约追踪', '资源中心', '系统管理'):
            self.assertIn("label: '%s'" % primary, html)
        for primary, permission in (
                ('工作台', 'workbench.access'),
                ('采购中心', 'procurement.access'),
                ('运营中心', 'operations.access'),
                ('财务中心', 'finance.access'),
                ('小犀助手', 'assistant.access'),
                ('数据分析', 'analytics.access')):
            self.assertIn("label: '%s'" % primary, html)
            self.assertIn("requiredPermission: '%s'" % permission, html)
            self.assertIn("codes: ['%s']" % permission, html)
        self.assertIn("label: '订单物流查询'", html)
        self.assertIn("label: '买家号库'", html)
        self.assertIn("label: '店铺管理'", html)
        self.assertIn("label: '代理IP'", html)
        self.assertIn("label: '运营采购单'", html)
        for permission in (
                'procurement.request.read', 'procurement.request.save',
                'procurement.request.submit',
                'procurement.execution.manage'):
            self.assertIn(permission, html)
        # 采购中心 P0：真实概览/筛选列表/分页/按需详情。
        self.assertIn("defaultModule: 'procurementorders'", html)
        self.assertIn("label: '采购任务', primary: 'procurement'", html)
        self.assertIn('id="procurementPanel"', html)
        self.assertIn('id="procurementExecutionPanel"', html)
        self.assertIn('id="executionFilterStatus"', html)
        self.assertIn('id="executionFilterSite"', html)
        self.assertIn('id="executionFilterBinding"', html)
        self.assertIn('id="executionFilterKeyword"', html)
        self.assertIn('销售订单号 / 采购单 / 店铺 / 运营', html)
        self.assertIn('id="executionQueueList"', html)
        self.assertIn('id="procurementOrderDetailPanel"', html)
        self.assertIn('id="procurementWorkspaceBody"', html)
        self.assertNotIn('const PROCUREMENT_EXECUTION_PAGE_SEED = [', html)
        self.assertIn('function renderProcurementExecutionPage()', html)
        self.assertIn('function renderProcurementWorkspace(data, focus=', html)
        self.assertIn('async function openProcurementOrderWorkspace(', html)
        self.assertIn("module === 'procurementexecution'", html)
        self.assertIn('认领成功的采购单直接进入本人工作台', html)
        self.assertIn("claimedByMe:'true'", html)
        self.assertIn("api('/api/procurement/orders?' + params.toString())", html)
        self.assertIn('id="executionPageSize"', html)
        self.assertIn('data-execution-action-toggle=', html)
        self.assertIn('data-execution-focus="tracking"', html)
        self.assertIn('function openExecutionActionMenu(toggle, menu)', html)
        self.assertIn('function placeExecutionActionMenu(toggle, menu)', html)
        self.assertIn('window.scrollBy({top: scrollAmount, behavior:', html)
        self.assertIn('upwardTop >= tableHeadBottom + 4', html)
        self.assertIn('function scheduleExecutionActionMenuReposition()', html)
        self.assertIn("menu.dataset.placement = openUp ? 'top' : 'bottom';", html)
        self.assertIn('.execution-action-menu { position: fixed; z-index: 120;', html)
        self.assertNotIn('bottom: calc(100% + 5px)', html)
        self.assertIn('操作始终保持两列：详情 / 快捷或处理', html)
        self.assertIn('快捷下单（绑定资源）', html)
        self.assertIn("'quick-checkout' : 'checkout'", html)
        self.assertIn('data-execution-return=', html)
        self.assertIn("'/return'", html)
        self.assertIn('物流单号：', html)
        self.assertIn('物流商：', html)
        self.assertIn('跟单摘要', html)
        self.assertIn('采购单状态', html)
        self.assertIn('.execution-page-filter-grid input:focus', html)
        self.assertIn('.procurement-action-stack { display: grid; grid-template-columns: repeat(2,minmax(72px,1fr))', html)
        self.assertIn('position: sticky; right: 0; z-index: 5;', html)
        self.assertNotIn('procurement-order-key', html)
        self.assertIn("const storeName = row.storeBaseName || row.storeName || '—';", html)
        self.assertIn('id="procurementCountUnclaimed"', html)
        self.assertIn('id="procurementFilterScope"', html)
        self.assertIn('id="procurementFilterSite"', html)
        self.assertIn('id="procurementFilterStore"', html)
        self.assertIn('id="procurementFilterOperator"', html)
        self.assertIn('id="procurementFilterKeyword"', html)
        self.assertIn('data-procurement-quick-scope="unclaimed"', html)
        for page_size in ('30', '50', '100', '300'):
            self.assertIn(
                '<option value="%s">%s 条</option>' % (page_size, page_size),
                html,
            )
        self.assertIn("params.set('pageSize', $('procurementPageSize').value || '30')", html)
        self.assertIn("$('procurementPageSize').value = '30'", html)
        self.assertIn('id="procurementTbody"', html)
        self.assertIn('id="procurementSelectAll"', html)
        self.assertIn('id="btnProcurementBatchClaim"', html)
        self.assertIn('data-procurement-select-order=', html)
        self.assertIn("params.set('store', store)", html)
        self.assertIn("params.set('operator', operator)", html)
        self.assertIn('预估利润', html)
        self.assertIn('利润率', html)
        self.assertIn('data-procurement-field="salesAmount"', html)
        self.assertIn('DEFAULT_PROCUREMENT_FIELD_VISIBILITY', html)
        self.assertIn('function applyProcurementFieldVisibility()', html)
        self.assertIn('id="btnProcurementPrev"', html)
        self.assertIn('id="btnProcurementNext"', html)
        self.assertIn('id="procurementDetailMask"', html)
        self.assertIn('id="procurementClaimMask"', html)
        self.assertIn('id="procurementExecutionMask"', html)
        self.assertIn('id="procurementCheckoutMask"', html)
        self.assertIn('id="procurementCheckoutResource"', html)
        self.assertIn('id="btnProcurementCheckoutConfirm"', html)
        self.assertIn('data-procurement-claim=', html)
        self.assertIn('data-procurement-claim-order=', html)
        self.assertIn('data-procurement-plan-open', html)
        self.assertIn('function openProcurementClaimPreview(button)', html)
        self.assertIn('function openProcurementBatchClaimPreview()', html)
        self.assertIn('async function confirmProcurementClaimPreview()', html)
        self.assertIn('function splitProcurementExecutionGroups()', html)
        self.assertIn('async function saveProcurementExecutionPreview()', html)
        self.assertIn('async function openProcurementCheckoutDialog(', html)
        self.assertIn('async function confirmProcurementCheckoutDialog()', html)
        self.assertIn("api('/api/cloud/buyer-accounts?' + params.toString())", html)
        self.assertIn("'/checkout-attempts'", html)
        self.assertIn("'/begin'", html)
        self.assertIn('idempotencyKey:state.idempotencyKey', html)
        self.assertIn("selectableOnly:'true'", html)
        self.assertIn('item.hubEnvironment.ref && item.hubEnvironment.name', html)
        self.assertNotIn('快捷下单入口已保留：下一步绑定环境与买家号', html)
        self.assertIn('function safeProcurementImageUrl(value)', html)
        self.assertIn('店铺 / 运营', html)
        self.assertIn('收件人 / 国家', html)
        self.assertIn('主规格（采购备注）', html)
        self.assertIn('次规格（采购备注）', html)
        self.assertIn('打开环境查单', html)
        self.assertIn("api('/api/procurement/claims'", html)
        self.assertIn("'/splits'", html)
        self.assertNotIn('.procurement-table-scroll { max-height:', html)
        self.assertIn('列表按每页数量向下展开', html)
        self.assertIn("api('/api/procurement/overview')", html)
        self.assertIn("return '/api/procurement/orders?'", html)
        self.assertIn("api('/api/procurement/orders/'", html)
        self.assertIn('function loadProcurementWorkspace(resetPage=false)', html)
        self.assertIn('function renderProcurementRows(data)', html)
        self.assertIn('function renderProcurementDetail(data)', html)
        self.assertIn("$('procurementDetailBody').replaceChildren();", html)
        self.assertIn('列表只展示收件人和国家', html)
        for permission in (
                'resource.store.read', 'resource.store.configure',
                'resource.store.credential.update', 'resource.store.clone',
                'resource.ip.read', 'resource.ip.test',
                'resource.ip.allocate', 'resource.ip.credential.manage'):
            self.assertIn(permission, html)
        self.assertIn("label: '组织与权限'", html)
        self.assertIn("label: '登录会话'", html)
        self.assertGreaterEqual(html.count("codes: ['resource.buyer.import']"), 2)
        self.assertGreaterEqual(html.count("codes: ['system.member.manage']"), 1)
        self.assertIn('data-permission-primary-toggle=', html)
        self.assertIn('data-permission-secondary-toggle=', html)
        self.assertIn('function changePermissionPolicySelection(event)', html)
        self.assertIn('function hasPrimaryAccess(primary)', html)
        self.assertIn('!hasPrimaryAccess(button.dataset.primary)', html)
        self.assertIn('Array.from(selectedPermissionCodesForRole(roleId)).sort()', html)
        self.assertIn('成员状态、成员角色、会话、数据范围和无权限码入口不在此修改', html)
        self.assertIn('class="permission-workspace"', html)
        self.assertIn('class="card permission-policy-card"', html)
        self.assertIn('id="permissionRoleSearch"', html)
        self.assertIn('id="permissionRoleList" role="listbox"', html)
        self.assertIn('id="permissionRoleCount"', html)
        self.assertIn('function renderPermissionRoleList()', html)
        self.assertIn('function selectPermissionPolicyRole(roleId)', html)
        self.assertIn('const permissionPolicyDirtyRoleIds = new Set()', html)
        self.assertIn('const expandedPermissionPrimaryIds = new Set()', html)
        self.assertIn('data-permission-primary-expand=', html)
        self.assertIn('data-permission-expand-all="true"', html)
        self.assertIn('data-permission-expand-all="false"', html)
        self.assertIn('.permission-secondary-grid[hidden] { display: none; }', html)
        self.assertIn('切换角色将丢弃这些调整', html)
        self.assertNotIn('roleGrid.innerHTML = adminRoles.map', html)
        self.assertIn("requiredRole: 'super_admin'", html)
        self.assertIn('仅超级管理员可查看和修改飞书连接', html)
        self.assertIn("roles.includes('admin') ? '管理员'", html)
        self.assertIn("label: '云端服务配置', badge: '仅超管'", html)
        self.assertIn('管理员默认拥有全部日常业务权限', html)
        self.assertIn('管理员不含云端服务配置和代理凭证管理权限', html)
        self.assertIn('id="storeManagementPanel"', html)
        self.assertIn('id="proxyManagementPanel"', html)
        self.assertIn('/api/resources/stores', html)
        self.assertIn('/api/resources/proxies/check/start', html)
        self.assertIn('普通列表跳过代理账号/密码列', html)
        self.assertIn('id="proxyTypeCatalog"', html)
        self.assertIn('id="proxyFilterType"', html)
        self.assertIn('id="proxyFilterScenario"', html)
        self.assertIn('711 动态住宅代理仅在采购环境创建时通过 API 提取', html)
        self.assertIn('动态住宅代理按采购任务实时提取，不生成固定库存行', html)
        self.assertIn('静态住宅 IP 当前暂无使用场景，尚未接入资产台账', html)
        self.assertIn('当前阶段只读；配置、克隆和凭证更新尚未开放', html)
        self.assertIn('async function loadLocalSettingsPage()', html)
        self.assertIn('async function loadLarkConnectionPage()', html)
        self.assertNotIn('id="btnSettings"', html)
        self.assertNotIn('id="settingsMask"', html)
        self.assertIn('id="cfgLarkAppId"', html)
        self.assertIn('id="cfgLarkAppSecret"', html)
        self.assertIn('id="cfgLarkLedgerUrl"', html)
        self.assertIn('id="cfgLarkAppIdState"', html)
        self.assertIn('id="cfgLarkAppSecretState"', html)
        self.assertIn('id="cfgLarkLedgerUrlState"', html)
        self.assertIn('id="cfgLarkConnectedTarget"', html)
        self.assertIn('id="larkSaveFeedback"', html)
        self.assertIn('function saveLarkConfig()', html)
        self.assertIn("button.textContent = '保存中…'", html)
        self.assertIn("setLarkSaveFeedback('bad', '⚠ 保存失败：'", html)
        self.assertIn('配置已保存，但连接名称获取失败', html)
        self.assertIn('重新验证连接与字段', html)
        self.assertIn('正在获取旧迁移源名称并只读检查字段', html)
        self.assertIn('await loadLarkConfigStatus()', html)
        self.assertIn('已安全保存 App Secret', html)
        self.assertIn('旧迁移源已配置；留空保持不变', html)
        self.assertIn('function renderLarkConnectedTarget(', html)
        self.assertIn('function createLarkTargetLink(text)', html)
        self.assertIn("link.href = '/api/lark/open-target'", html)
        self.assertIn("link.target = '_blank'", html)
        self.assertIn("link.rel = 'noopener noreferrer'", html)
        self.assertIn("link.className = 'config-target-link'", html)
        self.assertIn("event => event.stopPropagation()", html)
        self.assertIn('只读验证旧迁移源', html)
        self.assertIn('/api/lark/template', html)
        self.assertIn('下载旧台账兼容模板', html)
        self.assertIn('确认替换迁移源？', html)
        self.assertNotIn('id="cfgLarkBaseToken"', html)
        self.assertNotIn('id="cfgLarkTableId"', html)
        self.assertIn('包含 table=tbl...', html)
        self.assertIn('body:JSON.stringify({appId, appSecret, ledgerUrl, clearCredential, clearLedgerTarget, expectedRevision:larkConfigRevision})', html)
        self.assertIn('/api/lark/config', html)
        self.assertIn('/api/lark/preflight', html)
        self.assertNotIn('/api/lark/target-metadata', html)
        self.assertNotIn('refreshPending:true', html)
        self.assertNotIn('正在自动获取当前连接名称', html)
        self.assertIn('type="hidden" id="envWriteLedger"', html)
        self.assertIn('id="envExecutionState" role="status" aria-live="polite"', html)
        self.assertIn('id="envExecutionTitle"', html)
        self.assertIn('id="envExecutionDetail"', html)
        self.assertIn('let envSubmitting = false', html)
        self.assertIn('function beginEnvSubmission(title, detail)', html)
        self.assertIn('function paintEnvSubmissionState()', html)
        self.assertIn('function finishEnvSubmission()', html)
        self.assertIn('function renderEnvExecutionState(snap, mode)', html)
        self.assertIn("'⏳ 正在提交…'", html)
        self.assertIn('执行指令已确认，正在提交买家号建环境任务', html)
        self.assertNotIn('执行指令已确认，正在检查飞书台账并提交任务', html)
        self.assertIn('await paintEnvSubmissionState()', html)
        self.assertIn('建环境结果将先写数据库，再由云端自动同步独立 Base', html)
        self.assertIn('${taskName}部分失败：HubStudio', html)
        self.assertIn('任务状态暂时无法获取', html)
        self.assertIn('页面每 1.5 秒自动刷新，请勿重复提交', html)
        self.assertIn('id="envWriteLedgerLabel"', html)
        self.assertIn('id="regWriteLedgerLabel"', html)
        self.assertIn('function larkTargetDisplayName()', html)
        self.assertIn('function renderLarkWriteTargetText()', html)
        self.assertIn('测试版已停用旧统一台账直写', html)
        self.assertIn('id="authGate" role="dialog"', html)
        self.assertIn('id="authLoginButton"', html)
        self.assertIn('requiredPermission: \'resource.environment.create\'', html)
        self.assertIn('larkTargetBaseName = targetBaseName', html)
        self.assertNotIn('回写${larkTargetDisplayName()}', html)
        self.assertNotIn('回写飞书「买家号（统一）」', html)
        self.assertIn('id="btnEnvRetryLedger" disabled', html)
        self.assertIn('writeLarkLedger, confirmLarkWrite', html)
        self.assertIn('const writeLarkLedger = false', html)
        self.assertNotIn('补写本批次飞书台账', html)
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         '采购工具买家号入库模板.xlsx').is_file())
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         '买家号统一台账模板.xlsx').is_file())
        from openpyxl import load_workbook
        ledger_template = load_workbook(
            root / 'src' / 'purchase_tool' / 'web' /
            '买家号统一台账模板.xlsx', read_only=True)
        self.assertEqual(
            ledger_template.sheetnames,
            ['买家号统一台账', '字段配置说明'])
        self.assertEqual(
            [cell.value for cell in ledger_template['买家号统一台账'][1]],
            ['账号ID', '站点', '邮箱账号', '密码', '接码Key链接', 'Cookie',
             '号商购买单号', '购买日期', '账号状态', '绑定环境', '环境分组名',
             '环境序号', '采购员', '绑定时间', '首次登录日期', '最后使用日期',
             '创建时间', '备注', '累计下单数', '异常记录', '创建人',
             '迁移状态', '操作人'])
        guide_values = [
            cell.value for row in ledger_template['字段配置说明'].iter_rows()
            for cell in row if cell.value]
        self.assertIn('新刚、志恒、康德、宇航、熊、德、恒', guide_values)
        self.assertIn('正常、待复核（源表错列）', guide_values)
        self.assertIn('yyyy-mm-dd hh:mm', guide_values)
        self.assertIn('环境分组名', guide_values)
        guide = ledger_template['字段配置说明']
        self.assertIn('连续13列', guide['A2'].value)
        first_login = next(
            [cell.value for cell in row]
            for row in guide.iter_rows(
                min_row=1, max_row=guide.max_row, min_col=1, max_col=6)
            if row[0].value == '首次登录日期')
        self.assertEqual(first_login[4], '字段预检必需')
        self.assertIn('当前买家号建环境 API 不写入', first_login[5])
        ledger_template.close()
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         'xynigo-logo.png').is_file())
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         'xynigo-mascot-x.png').is_file())
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         'xynigo-x.ico').is_file())
        self.assertTrue((root / 'src' / 'purchase_tool' / 'web' /
                         'xynigo-x.png').is_file())
        ico = (root / 'src' / 'purchase_tool' / 'web' /
               'xynigo-x.ico').read_bytes()
        self.assertEqual(ico[:4], b'\x00\x00\x01\x00')
        self.assertGreaterEqual(int.from_bytes(ico[4:6], 'little'), 7)
        main_py = (root / 'src' / 'purchase_tool' / 'main.py').read_text(
            encoding='utf-8')
        self.assertIn("self._file(X_ICON_ICO, 'image/x-icon')", main_py)
        self.assertNotIn('FAVICON_ICO', main_py)
        export_py = (root / 'src' / 'purchase_tool' /
                     'excel_export.py').read_text(encoding='utf-8')
        self.assertIn('embed_cell_images', export_py)
        self.assertNotIn('TwoCellAnchor', export_py)
        self.assertNotIn('ws.add_image', export_py)
        cell_images_py = (root / 'src' / 'purchase_tool' /
                          'xlsx_cell_images.py').read_text(encoding='utf-8')
        self.assertIn("{'n': '_rvRel:LocalImageIdentifier', 't': 'i'}",
                      cell_images_py)
        self.assertIn("{'n': 'CalcOrigin', 't': 'i'}", cell_images_py)

    def test_device_channel_does_not_auto_install_owner_user_session(self):
        root = Path(__file__).resolve().parents[1]
        source = (root / 'src/purchase_tool/main.py').read_text(
            encoding='utf-8')
        self.assertIn('user_session_installer=None', source)
        self.assertNotIn(
            'user_session_installer=self.auth.install_executor_session',
            source)
        self.assertIn('interactive Feishu OAuth', source)


if __name__ == '__main__':
    unittest.main()
