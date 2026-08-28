# -*- coding: utf-8 -*-
"""店小秘 XYP2 → 单张采购共享协作表回归测试。"""
import base64
from io import BytesIO
import json
from pathlib import Path
import time
from xml.etree import ElementTree
import unittest
import zipfile

from openpyxl import Workbook, load_workbook

from purchase_tool.procurement_import import (
    LEGACY_OUTPUT_HEADERS_V1, ORDER_GROUP_COLORS, OUTPUT_HEADERS,
    ProcurementImportError,
    ProcurementImportService,
    _source_image_map, parse_xyp2_remark)
from purchase_tool.lark_sheet_sync import SheetTable


JPEG = b'\xff\xd8synthetic-xynigo-image\xff\xd9'
FEISHU_COLLABORATION_HEADERS = (
    '分单日期', '采购员', '销售订单号', '店铺', '运营', '包裹号',
    '采购状态', '优先级', '销售订单金额', '商品金额', '订单时间', '商品图片',
    '采购链接', '主规格', '次规格', '需求数量', '采购指导价',
    '收货人姓名', '收货人国家', '收货人州/省', '收货人城市',
    '地址1', '地址2', '邮编', '收货人电话', '采购备注',
    '下单批次', '买家号', '付款卡号', '采购订单号', '实际付款', '付款时间', '下单截图',
    '物流商', '物流单号', '物流截图', '跟单状态', '异常备注', '最近更新',
    '系统订单键', '导入操作人', '导入批次', '数据版本',
)
SHEET_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
OFFICE_REL_NS = (
    'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
CONTENT_NS = (
    'http://schemas.openxmlformats.org/package/2006/content-types')
XDR_NS = (
    'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing')
A_NS = 'http://schemas.openxmlformats.org/drawingml/2006/main'


class FakeSheetGateway:
    def __init__(self):
        self.headers = OUTPUT_HEADERS
        self.rows = ()
        self.images = {}
        self.backgrounds = {}
        self.links = {}
        self.writes = []
        self.append_calls = []
        self.presentation_calls = []
        self.header_presentation_calls = []
        self.hyperlink_calls = []

    def inspect(self, url):
        return {
            'url': 'https://tenant.feishu.cn/sheets/SheetToken123',
            'spreadsheetToken': 'SheetToken123',
            'revision': 8,
            'sheets': [{
                'sheetId': 'sheetA', 'sheetName': '采购分单协作区',
                'rowCount': 200, 'columnCount': len(self.headers),
                'hidden': False,
            }],
        }

    def read_table(self, url, sheet_id):
        return SheetTable(tuple(self.headers), tuple(self.rows), revision=8)

    def append_table_rows(self, url, sheet_name, columns, rows,
                          dtypes=None, formats=None):
        self.append_calls.append({
            'sheetName': sheet_name,
            'columns': tuple(columns),
            'rows': tuple(tuple(row) for row in rows),
            'dtypes': dict(dtypes or {}),
            'formats': dict(formats or {}),
        })
        next_row = max([number for number, _values in self.rows] or [1]) + 1
        current = list(self.rows)
        current.extend(
            (next_row + index, tuple(row)) for index, row in enumerate(rows))
        self.rows = tuple(current)
        return {'updated_rows_count': len(rows)}

    def normalize_collaboration_headers(self, url, sheet_id, sheet_name,
                                        headers, last_row=1, rows=()):
        old_headers = tuple(self.headers)
        aliases = {
            '分单标记': '分单日期', '分单时间': '分单日期',
            '采购单号': '包裹号', '销售金额': '销售订单金额',
            '平台订单号': '采购订单号',
            '收件人': '收货人姓名', '国家': '收货人国家',
            '收件地址': '地址1', '电话': '收货人电话',
        }
        canonical = [aliases.get(value, value) for value in old_headers]
        if '分单批次' in canonical:
            split_index = canonical.index('分单批次')
            import_index = canonical.index('导入批次')
            for row_number, raw in self.rows:
                split_value = raw[split_index] if split_index < len(raw) else ''
                import_value = raw[import_index] if import_index < len(raw) else ''
                if split_value and split_value != import_value:
                    raise AssertionError('unsafe split batch deletion at row %d' % row_number)
            canonical.pop(split_index)
            old_headers = tuple(
                value for index, value in enumerate(old_headers)
                if index != split_index)
            self.rows = tuple((row_number, tuple(
                value for index, value in enumerate(raw)
                if index != split_index)) for row_number, raw in self.rows)
        if '分单日期' not in canonical:
            canonical.insert(0, '分单日期')
        if '商品金额' not in canonical:
            canonical.insert(canonical.index('销售订单金额') + 1,
                             '商品金额')
        receiver_headers = [
            '收货人姓名', '收货人国家', '收货人州/省', '收货人城市',
            '地址1', '地址2', '邮编', '收货人电话',
        ]
        for value in receiver_headers:
            if value in canonical:
                canonical.remove(value)
        position = canonical.index('采购指导价') + 1
        canonical[position:position] = receiver_headers
        if '导入操作人' not in canonical:
            canonical.insert(canonical.index('系统订单键') + 1,
                             '导入操作人')
        remapped = []
        for row_number, raw in self.rows:
            values = {
                aliases.get(name, name): value
                for name, value in zip(old_headers, raw)
            }
            remapped.append((row_number, tuple(values.get(name) for name in canonical)))
        self.headers = tuple(canonical)
        self.rows = tuple(remapped)
        return {'operations': 1}

    def reorder_collaboration_headers(self, url, sheet_id, sheet_name,
                                      headers, desired_headers):
        old_headers = tuple(self.headers)
        ordered = list(old_headers)
        for target_index, name in enumerate(desired_headers):
            source_index = ordered.index(name)
            value = ordered.pop(source_index)
            ordered.insert(target_index, value)
        remapped = []
        for row_number, raw in self.rows:
            values = dict(zip(old_headers, raw))
            remapped.append((row_number, tuple(values.get(name) for name in ordered)))
        self.headers = tuple(ordered)
        self.rows = tuple(remapped)
        return {'operations': 1}

    def normalize_date_column(self, url, sheet_name, headers, rows,
                              header='分单日期', number_format='yyyy-mm-dd'):
        column_index = tuple(headers).index(header)
        normalized = []
        for row_number, raw in self.rows:
            values = list(raw)
            value = values[column_index] if column_index < len(values) else None
            if value:
                values[column_index] = str(value)[:10]
            normalized.append((row_number, tuple(values)))
        self.rows = tuple(normalized)
        return {'operations': 1, 'rows': len(normalized)}

    def apply_header_presentation(self, url, sheet_id, sheet_name, zones):
        self.header_presentation_calls.append(tuple(dict(item) for item in zones))
        return {'operations': 2}

    def image_presence(self, url, sheet_id, row_numbers, column='L'):
        return {int(row): bool(self.images.get(int(row))) for row in row_numbers}

    def row_backgrounds(self, url, sheet_id, row_numbers):
        return {int(row): self.backgrounds.get(int(row), '')
                for row in row_numbers}

    def apply_row_presentation(self, url, sheet_name, background_bands,
                               row_ranges, row_height=52, last_column='AO'):
        self.presentation_calls.append({
            'bands': tuple(dict(item) for item in background_bands),
            'ranges': tuple(tuple(item) for item in row_ranges),
            'height': row_height,
            'lastColumn': last_column,
        })
        for item in background_bands:
            for row in range(int(item['start']), int(item['end']) + 1):
                self.backgrounds[row] = item['color']
        return {'operations': 1}

    def hyperlink_presence(self, url, sheet_id, expected_links, column='M'):
        return {
            int(row): self.links.get(int(row)) == link
            for row, link in dict(expected_links).items()
        }

    def set_hyperlinks(self, url, sheet_id, links, column='M'):
        self.hyperlink_calls.append(tuple(links))
        for row, link in links:
            self.links[int(row)] = link
        return {'operations': 1}

    def set_image(self, url, sheet_id, row_number, image_bytes, mime,
                  column='L'):
        self.writes.append((sheet_id, int(row_number), bytes(image_bytes), mime))
        self.images[int(row_number)] = True
        return {'revision': 9}

    def verify_image(self, url, sheet_id, row_number, column='L'):
        return bool(self.images.get(int(row_number)))


def wait_for_sync(service, job_id):
    deadline = time.time() + 3
    while time.time() < deadline:
        result = service.image_sync_status(job_id)
        if result['state'] not in {
                'queued', 'validating', 'writing_rows', 'verifying_rows',
                'normalizing_headers', 'formatting_headers',
                'formatting_rows', 'writing_links', 'writing_images'}:
            return result
        time.sleep(0.01)
    raise AssertionError('image sync did not finish')


def qname(namespace, name):
    return '{%s}%s' % (namespace, name)


def xml_bytes(root):
    return ElementTree.tostring(
        root, encoding='utf-8', xml_declaration=True)


def add_test_drawings(xlsx_bytes, row_numbers):
    """Add standard DrawingML anchors without Pillow."""
    with zipfile.ZipFile(BytesIO(xlsx_bytes), 'r') as archive:
        infos = archive.infolist()
        entries = {info.filename: archive.read(info.filename)
                   for info in infos}

    sheet = ElementTree.fromstring(entries['xl/worksheets/sheet1.xml'])
    ElementTree.SubElement(
        sheet, qname(SHEET_NS, 'drawing'),
        {qname(OFFICE_REL_NS, 'id'): 'rId1'})
    entries['xl/worksheets/sheet1.xml'] = xml_bytes(sheet)

    relationships = ElementTree.Element(qname(REL_NS, 'Relationships'))
    ElementTree.SubElement(relationships, qname(REL_NS, 'Relationship'), {
        'Id': 'rId1', 'Type': OFFICE_REL_NS + '/drawing',
        'Target': '../drawings/drawing1.xml',
    })
    entries['xl/worksheets/_rels/sheet1.xml.rels'] = xml_bytes(relationships)

    drawing = ElementTree.Element(qname(XDR_NS, 'wsDr'))
    drawing_rels = ElementTree.Element(qname(REL_NS, 'Relationships'))
    for index, row_number in enumerate(row_numbers, start=1):
        anchor = ElementTree.SubElement(
            drawing, qname(XDR_NS, 'twoCellAnchor'), {'editAs': 'oneCell'})
        origin = ElementTree.SubElement(anchor, qname(XDR_NS, 'from'))
        for name, value in (
                ('col', 22), ('colOff', 0),
                ('row', row_number - 1), ('rowOff', 0)):
            ElementTree.SubElement(origin, qname(XDR_NS, name)).text = str(value)
        target = ElementTree.SubElement(anchor, qname(XDR_NS, 'to'))
        for name, value in (
                ('col', 23), ('colOff', 0),
                ('row', row_number), ('rowOff', 0)):
            ElementTree.SubElement(target, qname(XDR_NS, name)).text = str(value)
        picture = ElementTree.SubElement(anchor, qname(XDR_NS, 'pic'))
        properties = ElementTree.SubElement(
            picture, qname(XDR_NS, 'nvPicPr'))
        ElementTree.SubElement(
            properties, qname(XDR_NS, 'cNvPr'),
            {'id': str(index), 'name': 'Picture %d' % index})
        ElementTree.SubElement(properties, qname(XDR_NS, 'cNvPicPr'))
        fill = ElementTree.SubElement(picture, qname(XDR_NS, 'blipFill'))
        ElementTree.SubElement(
            fill, qname(A_NS, 'blip'),
            {qname(OFFICE_REL_NS, 'embed'): 'rId%d' % index})
        ElementTree.SubElement(picture, qname(XDR_NS, 'spPr'))
        ElementTree.SubElement(anchor, qname(XDR_NS, 'clientData'))
        ElementTree.SubElement(drawing_rels, qname(REL_NS, 'Relationship'), {
            'Id': 'rId%d' % index, 'Type': OFFICE_REL_NS + '/image',
            'Target': '../media/image%d.jpeg' % index,
        })
        entries['xl/media/image%d.jpeg' % index] = JPEG
    entries['xl/drawings/drawing1.xml'] = xml_bytes(drawing)
    entries['xl/drawings/_rels/drawing1.xml.rels'] = xml_bytes(drawing_rels)

    content_types = ElementTree.fromstring(entries['[Content_Types].xml'])
    ElementTree.SubElement(content_types, qname(CONTENT_NS, 'Default'), {
        'Extension': 'jpeg', 'ContentType': 'image/jpeg'})
    ElementTree.SubElement(content_types, qname(CONTENT_NS, 'Override'), {
        'PartName': '/xl/drawings/drawing1.xml',
        'ContentType': ('application/vnd.openxmlformats-officedocument.'
                        'drawing+xml')})
    entries['[Content_Types].xml'] = xml_bytes(content_types)

    output = BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as archive:
        existing = set()
        for info in infos:
            archive.writestr(info, entries[info.filename])
            existing.add(info.filename)
        for name, data in entries.items():
            if name not in existing:
                archive.writestr(name, data)
    return output.getvalue()


def xyp2_text():
    payload = {
        'd': 'mx', 'c': 'MXN',
        'i': [
            ['SOURCE-01', '422790137', 'I8mmn32aip2g7d', '27_447',
             'Multicolor', 'M', 110.09, 0.65, 38.53, 1],
            ['SOURCE-02', '422489591', 'I3mmn32amfnfut', '27_447',
             'Black', 'L', 142.93, 0.65, 50.03, 1],
        ],
        'r': 11.44,
    }
    return '[XYP2]%s[/XYP2]' % json.dumps(
        payload, ensure_ascii=False, separators=(',', ':'))


def source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'order_'
    headers = [
        '店铺账号', '订单号', '包裹号', '下单时间', '付款时间', '订单金额',
        '币种缩写', '收货人姓名', '收货人国家', '收货人州/省', '收货人城市',
        '地址1', '地址2', '邮编', '收货人电话', 'SKU', '产品名称', '产品规格',
        '产品售价', '单个产品数量', '产品图片网址', '客服备注', '产品图片',
    ]
    worksheet.append(headers)
    common = [
        '测试店铺-测试运营（二组）$', 'GSH-TEST-001', 'XMWU-TEST-001',
        '2026-08-26 12:00:00', '2026-08-26 11:59:00', 300, 'MXN',
        'Recipient Test', 'MEXICO', 'State', 'City', 'Address 1',
        'Address 2', '00123', '0012345678',
    ]
    worksheet.append(common + [
        'ERP-SKU-01', 'Product one', 'SOURCE-01:Dark Grey-M', 150, 1,
        'https://img.ltwebstatic.com/test/source-one.jpg', xyp2_text(), '',
    ])
    worksheet.append(common + [
        'ERP-SKU-02', 'Product two', 'SOURCE-02:Black-L', 150, 1,
        'https://img.ltwebstatic.com/test/source-two.jpg', xyp2_text(), '',
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return add_test_drawings(buffer.getvalue(), [2, 3])


def aggregated_source_workbook():
    payload = {
        'd': 'mx', 'c': 'MXN',
        'i': [
            ['SOURCE-S', '403511191', 'SKU-CODE-S', '27_447',
             'Black', 'S', 170.55, 0, 170.55, 4],
            ['SOURCE-M', '403511191', 'SKU-CODE-M', '27_447',
             'Black', 'M', 170.55, 0, 170.55, 4],
            ['SOURCE-L', '403511191', 'SKU-CODE-L', '27_447',
             'Black', 'L', 170.55, 0, 170.55, 4],
        ],
        'r': 0,
    }
    remark = '[XYP2]%s[/XYP2]' % json.dumps(
        payload, ensure_ascii=False, separators=(',', ':'))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'order_'
    headers = [
        '店铺账号', '订单号', '包裹号', '下单时间', '付款时间', '订单金额',
        '币种缩写', '收货人姓名', '收货人国家', '收货人州/省', '收货人城市',
        '地址1', '地址2', '邮编', '收货人电话', 'SKU', '产品名称', '产品规格',
        '产品售价', '单个产品数量', '产品图片网址', '客服备注', '产品图片',
    ]
    worksheet.append(headers)
    common = [
        '聚合测试店铺-测试运营（一组）$', 'GSH-AGGREGATED-001',
        'XMWU-AGGREGATED-001', '2026-08-26 12:00:00',
        '2026-08-26 11:59:00', 2046.60, 'MXN', 'Recipient Test',
        'MEXICO', 'State', 'City', 'Address 1', 'Address 2', '00123',
        '0012345678',
    ]
    for seller_sku, size in (
            ('SOURCE-S', 'S'), ('SOURCE-M', 'M'), ('SOURCE-L', 'L')):
        for _index in range(4):
            worksheet.append(common + [
                'ERP-' + seller_sku, 'Aggregated product',
                '%s:Black-%s' % (seller_sku, size), 170.55, 1,
                'https://img.ltwebstatic.com/test/%s.jpg' % seller_sku.lower(),
                remark, '',
            ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ProcurementImportTests(unittest.TestCase):
    def test_parses_compact_xyp2_and_rebuilds_precise_links(self):
        parsed = parse_xyp2_remark('普通备注 ' + xyp2_text())
        self.assertEqual(parsed.site, 'MX')
        self.assertEqual(parsed.currency, 'MXN')
        self.assertEqual(parsed.rounding_amount, 11.44)
        self.assertEqual(len(parsed.items), 2)
        self.assertEqual(parsed.items[0].seller_sku, 'SOURCE-01')
        self.assertIn('goods_id=422790137', parsed.items[0].purchase_link)
        self.assertIn('skucode=I8mmn32aip2g7d', parsed.items[0].purchase_link)
        self.assertIn('#xv=1&p=Multicolor&s=M', parsed.items[0].purchase_link)

    def test_rejects_truncated_legacy_remark_with_actionable_message(self):
        with self.assertRaisesRegex(
                ProcurementImportError, '最新版插件.*XYP2'):
            parse_xyp2_remark('[XYNIGO_PURCHASE_V1]{"truncated":')

    def test_reads_exported_drawing_images_without_pillow(self):
        images = _source_image_map(source_workbook(), 'order_')
        self.assertEqual(sorted(images), [2, 3])
        self.assertEqual(images[2], JPEG)

    def test_generates_single_collaboration_sheet_and_prevents_amount_duplication(self):
        source = source_workbook()
        service = ProcurementImportService()
        result = service.parse(
            'order_test.xlsx', base64.b64encode(source).decode('ascii'))
        self.assertEqual(result['sourceRows'], 2)
        self.assertEqual(result['orderCount'], 1)
        self.assertEqual(result['detailCount'], 2)
        self.assertEqual(result['orderImageCount'], 2)
        self.assertNotIn('purchaseImageCount', result)
        self.assertEqual(result['warningCount'], 0)
        self.assertEqual(
            [row['imageIndex'] for row in result['preview']], [0, 1])
        self.assertEqual(
            [row['orderGroupIndex'] for row in result['preview']], [0, 0])
        self.assertEqual(
            [row['itemSalesAmount'] for row in result['preview']], [150, 150])
        self.assertEqual(
            [row['salesCurrency'] for row in result['preview']], ['MXN', 'MXN'])
        self.assertNotIn('recipient', json.dumps(result).lower())

        preview_data, preview_mime = service.preview_image(
            result['planId'], '0')
        self.assertEqual(preview_data, JPEG)
        self.assertEqual(preview_mime, 'image/jpeg')
        with self.assertRaisesRegex(
                ProcurementImportError, '行号超出范围'):
            service.preview_image(result['planId'], '99')

        plan = service.pending[result['planId']]
        self.assertEqual(plan.rows[0].values['店铺'], '测试店铺')
        self.assertEqual(plan.rows[0].values['运营'], '测试运营')
        self.assertEqual(plan.rows[0].values['销售订单金额'], 300)
        self.assertIsNone(plan.rows[1].values['销售订单金额'])
        self.assertEqual(plan.rows[0].values['商品金额'], 150)
        self.assertEqual(plan.rows[1].values['商品金额'], 150)
        self.assertEqual(plan.rows[0].values['收货人姓名'], 'Recipient Test')
        self.assertEqual(plan.rows[0].values['收货人国家'], 'MX')
        self.assertEqual(plan.rows[0].values['收货人州/省'], 'State')
        self.assertEqual(plan.rows[0].values['收货人城市'], 'City')
        self.assertEqual(plan.rows[0].values['地址1'], 'Address 1')
        self.assertEqual(plan.rows[0].values['地址2'], 'Address 2')
        self.assertEqual(plan.rows[0].values['邮编'], '00123')
        self.assertEqual(plan.rows[0].values['收货人电话'], '0012345678')
        self.assertIn('凑单补差：11.44', plan.rows[0].values['采购备注'])
        self.assertNotIn('凑单补差', plan.rows[1].values['采购备注'])
        self.assertRegex(
            plan.rows[0].values['系统订单键'],
            r'^OK1-[0-9A-HJKMNP-TV-Z]{5}(?:-[0-9A-HJKMNP-TV-Z]{5}){3}$')
        self.assertRegex(
            plan.rows[0].values['分单日期'],
            r'^\d{4}-\d{2}-\d{2}$')
        self.assertEqual(
            plan.rows[1].values['分单日期'],
            plan.rows[0].values['分单日期'])
        self.assertNotIn('分单批次', plan.rows[0].values)
        self.assertEqual(plan.rows[0].values['导入批次'], result['importBatch'])
        self.assertRegex(result['importBatch'], r'^order_test-[0-9a-f]{12}$')
        self.assertEqual(plan.rows[0].values['数据版本'], 'XYP2')
        self.assertEqual(len(ORDER_GROUP_COLORS), 6)
        rgb = [tuple(int(color[index:index + 2], 16)
                     for index in (0, 2, 4))
               for color in ORDER_GROUP_COLORS]
        adjacent_distances = [
            sum(abs(left[channel] - right[channel]) for channel in range(3))
            for left, right in zip(rgb, rgb[1:] + rgb[:1])
        ]
        self.assertGreaterEqual(min(adjacent_distances), 45)

        repeated = service.parse(
            'order_test.xlsx', base64.b64encode(source).decode('ascii'))
        self.assertEqual(repeated['importBatch'], result['importBatch'])

        data, name, mime = service.export(result['planId'])
        self.assertTrue(name.endswith('.xlsx'))
        self.assertIn('spreadsheetml.sheet', mime)
        workbook = load_workbook(BytesIO(data), data_only=False)
        self.assertEqual(workbook.sheetnames, ['采购协作区'])
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(tuple(headers), OUTPUT_HEADERS)
        self.assertEqual(tuple(headers), FEISHU_COLLABORATION_HEADERS)
        self.assertEqual(len(headers), 43)
        self.assertEqual(headers[11], '商品图片')
        self.assertEqual(headers[28], '付款卡号')
        self.assertEqual(headers[29], '采购订单号')
        self.assertEqual(headers[32], '下单截图')
        self.assertEqual(headers[35], '物流截图')
        self.assertEqual(headers[40], '导入操作人')
        guide_index = headers.index('采购指导价')
        self.assertEqual(headers[guide_index + 1:guide_index + 9], [
            '收货人姓名', '收货人国家', '收货人州/省', '收货人城市',
            '地址1', '地址2', '邮编', '收货人电话'])
        self.assertNotIn('采购商品图', headers)
        amount_column = headers.index('销售订单金额') + 1
        self.assertEqual(worksheet.cell(2, amount_column).value, 300)
        self.assertIsNone(worksheet.cell(3, amount_column).value)
        self.assertEqual(
            worksheet.cell(2, 1).fill.fgColor.rgb[-6:],
            ORDER_GROUP_COLORS[0])
        self.assertEqual(
            worksheet.cell(3, 1).fill.fgColor.rgb[-6:],
            ORDER_GROUP_COLORS[0])
        self.assertEqual(worksheet.row_dimensions[2].height, 52.5)
        self.assertEqual(worksheet.column_dimensions['L'].width, 10.5)
        self.assertEqual(worksheet.freeze_panes, 'F2')
        self.assertEqual(worksheet['A2'].number_format, 'yyyy-mm-dd')
        self.assertEqual(worksheet['A2'].value.date().isoformat(),
                         plan.rows[0].values['分单日期'])
        with zipfile.ZipFile(BytesIO(data)) as archive:
            self.assertIn('xl/richData/rdrichvalue.xml', archive.namelist())
            self.assertEqual(len([
                name for name in archive.namelist()
                if name.startswith('xl/media/image')]), 2)

    def test_aggregates_repeated_source_rows_without_false_warnings(self):
        service = ProcurementImportService()
        result = service.parse(
            'order_aggregated.xlsx',
            base64.b64encode(aggregated_source_workbook()).decode('ascii'))

        self.assertEqual(result['sourceRows'], 12)
        self.assertEqual(result['orderCount'], 1)
        self.assertEqual(result['detailCount'], 3)
        self.assertEqual(result['warningCount'], 0)
        self.assertEqual(result['errorCount'], 0)
        self.assertEqual(result['issues'], [])
        self.assertEqual(
            [row['quantity'] for row in result['preview']], [4, 4, 4])
        self.assertEqual(
            [row['itemSalesAmount'] for row in result['preview']],
            [682.2, 682.2, 682.2])

        plan = service.pending[result['planId']]
        self.assertEqual(
            [row.values['商品金额'] for row in plan.rows],
            [682.2, 682.2, 682.2])
        self.assertEqual(
            [row.values['需求数量'] for row in plan.rows], [4, 4, 4])
        self.assertEqual(plan.rows[0].values['销售订单金额'], 2046.6)
        self.assertIsNone(plan.rows[1].values['销售订单金额'])
        self.assertIsNone(plan.rows[2].values['销售订单金额'])

    def test_dynamic_sheet_target_appends_rows_and_idempotently_writes_images(self):
        gateway = FakeSheetGateway()
        # 下单/物流截图是可选采购执行字段；现有 32 列表可直接导入。
        gateway.headers = LEGACY_OUTPUT_HEADERS_V1
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        inspected = service.inspect_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123')
        self.assertEqual(inspected['sheets'][0]['sheetId'], 'sheetA')
        validated = service.validate_target(
            result['planId'], inspected['spreadsheetUrl'], 'sheetA')
        self.assertTrue(validated['valid'])
        self.assertEqual(validated['headerCount'], 32)
        self.assertEqual(validated['missingRecommendedColumns'],
                         ['收货人州/省', '收货人城市', '地址1', '地址2',
                          '邮编', '收货人电话', '商品金额', '导入操作人'])
        with self.assertRaisesRegex(ProcurementImportError, '明确确认'):
            service.start_image_sync(result['planId'], confirm_write=False)

        started = service.start_image_sync(
            result['planId'], confirm_write=True,
            operator_name='测试导入员')
        completed = wait_for_sync(service, started['jobId'])
        self.assertEqual(completed['state'], 'completed')
        self.assertEqual(completed['rowsTotal'], 2)
        self.assertEqual(completed['rowsWritten'], 2)
        self.assertEqual(completed['rowsExisting'], 0)
        self.assertEqual(completed['rowsStyled'], 2)
        self.assertEqual(completed['linksWritten'], 2)
        self.assertEqual(completed['written'], 2)
        self.assertEqual(completed['skippedExisting'], 0)
        self.assertEqual(len(gateway.append_calls), 1)
        self.assertEqual(len(gateway.presentation_calls), 1)
        self.assertEqual(gateway.presentation_calls[0]['height'], 52)
        self.assertEqual(gateway.presentation_calls[0]['lastColumn'], 'AN')
        self.assertEqual(gateway.presentation_calls[0]['bands'], ({
            'start': 2, 'end': 3,
            'color': '#' + ORDER_GROUP_COLORS[0]},))
        self.assertEqual(len(gateway.hyperlink_calls), 1)
        self.assertEqual(
            gateway.append_calls[0]['columns'], gateway.headers)
        operator_index = gateway.headers.index('导入操作人')
        self.assertEqual(
            {row[operator_index] for row in gateway.append_calls[0]['rows']},
            {'测试导入员'})
        self.assertEqual(gateway.append_calls[0]['dtypes']['需求数量'], 'int64')
        self.assertEqual(gateway.append_calls[0]['dtypes']['销售订单金额'], 'Float64')
        self.assertEqual(
            gateway.append_calls[0]['dtypes']['分单日期'], 'datetime64[ns]')
        self.assertEqual(
            gateway.append_calls[0]['formats']['分单日期'], 'yyyy-mm-dd')
        self.assertEqual(
            [(item[0], item[1], item[3]) for item in gateway.writes],
            [('sheetA', 2, 'image/jpeg'), ('sheetA', 3, 'image/jpeg')])

        # 采购员后续可用另一种整行底色标任务；幂等重试不得覆盖。
        gateway.backgrounds[2] = '#ABCDEF'
        retried = service.start_image_sync(
            result['planId'], confirm_write=True,
            operator_name='测试导入员')
        idempotent = wait_for_sync(service, retried['jobId'])
        self.assertEqual(idempotent['state'], 'completed')
        self.assertEqual(idempotent['rowsWritten'], 0)
        self.assertEqual(idempotent['rowsExisting'], 2)
        self.assertEqual(idempotent['written'], 0)
        self.assertEqual(idempotent['skippedExisting'], 2)
        self.assertEqual(idempotent['rowsStyled'], 0)
        self.assertEqual(idempotent['linksWritten'], 0)
        self.assertEqual(gateway.backgrounds[2], '#ABCDEF')
        self.assertEqual(len(gateway.writes), 2)
        self.assertEqual(len(gateway.append_calls), 1)

    def test_partial_same_batch_is_resumed_without_duplicating_existing_rows(self):
        gateway = FakeSheetGateway()
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        plan = service.pending[result['planId']]
        gateway.rows = ((8, tuple(
            plan.rows[0].values[name] for name in OUTPUT_HEADERS)),)
        service.validate_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123', 'sheetA')

        started = service.start_sheet_sync(
            result['planId'], confirm_write=True)
        completed = wait_for_sync(service, started['jobId'])
        self.assertEqual(completed['state'], 'completed')
        self.assertEqual(completed['rowsWritten'], 1)
        self.assertEqual(completed['rowsExisting'], 1)
        self.assertEqual(len(gateway.append_calls), 1)
        self.assertEqual(len(gateway.append_calls[0]['rows']), 1)
        self.assertEqual(len(gateway.rows), 2)

    def test_cross_batch_mixed_orders_skip_history_and_append_only_new(self):
        gateway = FakeSheetGateway()
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        plan = service.pending[result['planId']]
        # Turn the second synthetic detail into a separate new order so the
        # target contains one complete historical order and one absent order.
        plan.rows[1].values['销售订单号'] = 'GSH-NEW-002'
        plan.rows[1].values['包裹号'] = 'XMWU-NEW-002'
        plan.rows[1].values['系统订单键'] = (
            '测试店铺|GSH-NEW-002|XMWU-NEW-002')
        historical_rows = []
        for copy_index in range(7):
            historical = dict(plan.rows[0].values)
            historical['导入批次'] = 'older-batch-%d' % copy_index
            row_number = 8 + copy_index
            historical_rows.append((row_number, tuple(
                historical[name] for name in OUTPUT_HEADERS)))
            gateway.images[row_number] = True
            gateway.backgrounds[row_number] = '#PRESERVE'
            gateway.links[row_number] = historical['采购链接']
        gateway.rows = tuple(historical_rows)
        service.validate_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123', 'sheetA')

        started = service.start_sheet_sync(
            result['planId'], confirm_write=True)
        completed = wait_for_sync(service, started['jobId'])
        self.assertEqual(completed['state'], 'completed')
        self.assertEqual(completed['rowsWritten'], 1)
        self.assertEqual(completed['rowsExisting'], 1)
        self.assertEqual(completed['rowsStyled'], 1)
        self.assertEqual(completed['written'], 1)
        self.assertEqual(completed['skippedExisting'], 0)
        self.assertEqual(len(gateway.append_calls), 1)
        self.assertEqual(len(gateway.append_calls[0]['rows']), 1)
        self.assertEqual(
            gateway.append_calls[0]['rows'][0][
                OUTPUT_HEADERS.index('销售订单号')],
            'GSH-NEW-002')
        self.assertEqual(len(gateway.rows), 8)
        self.assertEqual(gateway.presentation_calls[0]['ranges'], ((15, 15),))
        self.assertEqual([item[1] for item in gateway.writes], [15])
        self.assertTrue(all(
            gateway.backgrounds[row_number] == '#PRESERVE'
            for row_number in range(8, 15)))

    def test_cross_batch_changed_order_aborts_before_append(self):
        gateway = FakeSheetGateway()
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        plan = service.pending[result['planId']]
        historical_rows = []
        for index, row in enumerate(plan.rows):
            historical = dict(row.values)
            historical['导入批次'] = 'older-batch'
            if index == 0:
                historical['需求数量'] = int(historical['需求数量']) + 1
            historical_rows.append((8 + index, tuple(
                historical[name] for name in OUTPUT_HEADERS)))
        gateway.rows = tuple(historical_rows)
        service.validate_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123', 'sheetA')

        started = service.start_sheet_sync(
            result['planId'], confirm_write=True)
        failed = wait_for_sync(service, started['jobId'])
        self.assertEqual(failed['state'], 'failed')
        self.assertIn('不支持修改已导入订单', failed['error'])
        self.assertIn('认领前修改请走独立修订流程', failed['error'])
        self.assertEqual(gateway.append_calls, [])
        self.assertEqual(gateway.writes, [])

    def test_cross_batch_all_historical_orders_remain_read_only(self):
        gateway = FakeSheetGateway()
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        plan = service.pending[result['planId']]
        historical_rows = []
        row_number = 8
        for copy_index in range(7):
            for row in plan.rows:
                historical = dict(row.values)
                historical['导入批次'] = 'older-batch-%d' % copy_index
                # 旧共享表仍可用中文/竖线键参与新版本防重，无需人工迁移。
                historical['系统订单键'] = (
                    '测试店铺|GSH-TEST-001|XMWU-TEST-001')
                historical_rows.append((row_number, tuple(
                    historical[name] for name in OUTPUT_HEADERS)))
                gateway.images[row_number] = True
                gateway.backgrounds[row_number] = '#PRESERVE'
                gateway.links[row_number] = historical['采购链接']
                row_number += 1
        gateway.rows = tuple(historical_rows)
        service.validate_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123', 'sheetA')

        started = service.start_sheet_sync(
            result['planId'], confirm_write=True)
        completed = wait_for_sync(service, started['jobId'])
        self.assertEqual(completed['state'], 'completed')
        self.assertEqual(completed['rowsWritten'], 0)
        self.assertEqual(completed['rowsExisting'], 2)
        self.assertEqual(completed['rowsStyled'], 0)
        self.assertEqual(completed['linksWritten'], 0)
        self.assertEqual(completed['total'], 0)
        self.assertEqual(completed['written'], 0)
        self.assertEqual(completed['skippedExisting'], 0)
        self.assertEqual(gateway.append_calls, [])
        self.assertEqual(gateway.presentation_calls, [])
        self.assertEqual(gateway.hyperlink_calls, [])
        self.assertEqual(gateway.writes, [])
        self.assertTrue(all(
            gateway.backgrounds[number] == '#PRESERVE'
            for number in range(8, 22)))

    def test_batch_row_mismatch_aborts_before_any_image_write(self):
        gateway = FakeSheetGateway()
        service = ProcurementImportService(
            sheet_gateway=gateway, sleep_fn=lambda _seconds: None)
        result = service.parse(
            'order_test.xlsx',
            base64.b64encode(source_workbook()).decode('ascii'))
        plan = service.pending[result['planId']]
        valid_rows = [
            (20 + index, tuple(row.values[name] for name in OUTPUT_HEADERS))
            for index, row in enumerate(plan.rows)]
        gateway.rows = tuple(valid_rows)
        service.validate_target(
            result['planId'],
            'https://tenant.feishu.cn/sheets/SheetToken123', 'sheetA')
        gateway.rows = tuple(valid_rows + [(99, valid_rows[0][1])])

        started = service.start_image_sync(
            result['planId'], confirm_write=True)
        failed = wait_for_sync(service, started['jobId'])
        self.assertEqual(failed['state'], 'failed')
        self.assertIn('1 条与解析计划不一致或重复', failed['error'])
        self.assertEqual(gateway.writes, [])
        self.assertEqual(gateway.append_calls, [])

    def test_frontend_and_same_origin_routes_are_wired(self):
        root = Path(__file__).resolve().parents[1]
        html = (root / 'src/purchase_tool/web/index.html').read_text(
            encoding='utf-8')
        main = (root / 'src/purchase_tool/main.py').read_text(
            encoding='utf-8')
        self.assertIn('data-module="procurementimport"', html)
        self.assertIn('id="procurementImportPanel"', html)
        self.assertIn('/api/assistant/procurement-import/parse', html)
        self.assertIn('/api/assistant/procurement-import/image?planId=', html)
        self.assertIn('/api/assistant/procurement-import/export?planId=', html)
        self.assertIn('/api/assistant/procurement-import/target/inspect', html)
        self.assertIn('/api/assistant/procurement-import/target/validate', html)
        self.assertIn('/api/assistant/procurement-import/sheet-sync', html)
        self.assertIn('function procurementImportResourcePath(path)', html)
        self.assertIn("'/v1/assistant/procurement-import/'", html)
        self.assertIn('id="procurementImportTargetUrl"', html)
        self.assertIn('无需人工复制粘贴', html)
        self.assertIn('不会覆盖已有单元格', html)
        self.assertIn('<th>商品金额</th>', html)
        self.assertIn('销售订单 / 包裹号', html)
        self.assertIn('包裹号：${esc(row.packageNo)', html)
        self.assertIn('procurement-import-stack', html)
        self.assertIn('procurement-import-order-band-', html)
        self.assertIn('procurement-import-order-band-5', html)
        self.assertIn('background: #e0f1ff', html)
        self.assertIn('background: #fdebf5', html)
        self.assertIn('groupIndex % 6', html)
        self.assertIn('系统追踪区仅用于幂等防重', html)
        self.assertIn('已导入订单保持原样，不能在这里修改', html)
        self.assertIn('采购任务认领前修改属于独立修订流程', html)
        self.assertNotIn('商品图片缩小 25%', html)
        self.assertIn(
            '.procurement-import-target-grid select, '
            '.procurement-import-target-actions .btn', html)
        self.assertIn('height: 42px; min-height: 42px', html)
        self.assertIn('id="procurementImportSyncBar"', html)
        self.assertIn('aria-live="polite"', html)
        self.assertIn('@keyframes procurement-sync-shimmer', html)
        self.assertIn('procurementImportProgressPercent(status)', html)
        self.assertIn('正在重新校验工作表核心字段和跨批次重复状态', html)
        self.assertIn(
            '导入完成：新追加 ${status.rowsWritten} 行，已存在 '
            '${status.rowsExisting} 行；新设底色', html)
        self.assertIn("'/api/assistant/procurement-import/image'", main)
        self.assertIn(
            "'/api/assistant/procurement-import/parse': 'assistant.access'",
            main)
        self.assertIn(
            "'/api/assistant/procurement-import/export': 'assistant.access'",
            main)
        self.assertIn(
            "'/api/assistant/procurement-import/image-sync': 'assistant.access'",
            main)
        self.assertIn(
            "'/api/assistant/procurement-import/sheet-sync': 'assistant.access'",
            main)


if __name__ == '__main__':
    unittest.main()
