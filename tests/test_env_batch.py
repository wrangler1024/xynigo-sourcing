# -*- coding: utf-8 -*-
from io import BytesIO
import json
import os
from pathlib import Path
import random
import re
import tempfile
import threading
import unittest
from contextlib import nullcontext
from unittest.mock import patch
import zipfile

from openpyxl import Workbook, load_workbook

os.environ.setdefault(
    'XYNIGO_PROXY_LINK', 'https://proxy.example.test/{region}')

from purchase_tool.env_batch import (
    BUYER_ROSTER, BatchEnvOrchestrator, BatchPlanItem,
    BackupEnvOrchestrator, BuyerAccount,
    EnvironmentSnapshotIndex, EnvBatchError, RES_POOL, ResumeStateStore,
    VENDOR_TEMPLATE_HEADERS,
    backup_env_names, backup_result_tsv_bytes,
    batch_fingerprint, build_batch_plan, build_env_create_body,
    choose_resolution, envbatch_preflight, extract_vendor_order_no,
    format_remark, load_vendor_xlsx,
    mapping_workbook_bytes, normalize_buyer, normalize_env_site,
    parse_assignment, parse_vendor_workbook, probe_env_ip,
    deserialize_buyer_accounts, serialize_buyer_accounts,
    validate_assignment_template, validate_purchase_tag)
from purchase_tool.hub_api import HubApiError, HubStudioApi


TEST_TAG = 'MX-Purchase'
TEST_US_TAG = 'US-Purchase'
TEST_PROXY = 'https://proxy.example.test/{region}'


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def demo_rows():
    return [
        ['alpha@example.com', 'secret-pass-a',
         'https://codes.example.test/get?orderNo=a1b2c3',
         '[{"name":"sid","value":"secret-cookie-a"}]'],
        ['beta@example.com', 'secret-pass-b',
         'https://codes.example.test/get?orderNo=d4e5f6',
         '[{"name":"sid","value":"secret-cookie-b"}]'],
    ]


class FakeHub(object):
    def __init__(self, fail_cookie_once=False, fail_create_with_link=False):
        self.envs = []
        self.calls = []
        self.lock = threading.Lock()
        self.fail_cookie_once = fail_cookie_once
        self.fail_create_with_link = fail_create_with_link
        self.env_list_calls = 0
        self.env_lookup_calls = []

    def group_list(self):
        return [TEST_TAG, TEST_US_TAG]

    def env_list(self, tag=None):
        with self.lock:
            self.env_list_calls += 1
            return [dict(item) for item in self.envs
                    if not tag or item.get('tagName', TEST_TAG) == tag]

    def env_lookup(self, container_code=None, container_name=None,
                   tag_name=None):
        with self.lock:
            self.env_lookup_calls.append({
                'containerCode': str(container_code or ''),
                'containerName': str(container_name or ''),
                'tagName': str(tag_name or ''),
            })
            matched = []
            for item in self.envs:
                item_tag = item.get('tagName', TEST_TAG)
                if tag_name and item_tag != tag_name:
                    continue
                if (container_code and
                        str(item.get('containerCode') or '') ==
                        str(container_code)):
                    matched.append(item)
                elif (container_name and
                      str(item.get('containerName') or '') ==
                      str(container_name)):
                    matched.append(item)
            if len(matched) > 1:
                raise RuntimeError('synthetic ambiguous environment')
            return dict(matched[0]) if matched else None

    def env_create(self, body):
        with self.lock:
            self.calls.append(('create', dict(body)))
            if self.fail_create_with_link:
                raise RuntimeError('proxy rejected: ' + body['linkCode'])
            created = {
                'containerName': body['containerName'],
                'containerCode': str(9000 + len(self.envs)),
                'serialNumber': 1100 + len(self.envs),
                'tagName': body['tagName'],
                'remark': '',
            }
            self.envs.append(created)
        return {'containerCode': created['containerCode']}

    def env_import_cookie(self, code, cookie_text):
        self.calls.append(('cookie', str(code), cookie_text))
        if self.fail_cookie_once:
            self.fail_cookie_once = False
            raise RuntimeError('cookie import failed')
        return {}

    def container_add_account(self, code, email, password, site):
        self.calls.append(('account', str(code), email, password, site))
        return {}

    def env_update(self, code, name, remark):
        self.calls.append(('remark', str(code), name, remark))
        for env in self.envs:
            if str(env['containerCode']) == str(code):
                env['remark'] = remark
        return {}

    def env_delete(self, container_codes):
        wanted = {str(code) for code in container_codes}
        self.calls.append(('delete', tuple(sorted(wanted))))
        with self.lock:
            self.envs = [
                env for env in self.envs
                if str(env.get('containerCode') or '') not in wanted
            ]
        return True


class EnvBatchTests(unittest.TestCase):
    def test_repository_sample_matches_contract(self):
        repo = Path(__file__).resolve().parents[1]
        samples = (
            repo / 'examples' / 'buyer-import-template.xlsx',
            repo / 'src' / 'purchase_tool' / 'web' /
            '采购工具买家号入库模板.xlsx',
        )
        for sample in samples:
            workbook = load_workbook(sample, read_only=True, data_only=True)
            try:
                header = tuple(
                    cell.value for cell in workbook.worksheets[0][1][:4])
            finally:
                workbook.close()
            self.assertEqual(header, VENDOR_TEMPLATE_HEADERS)
            accounts = parse_vendor_workbook(str(sample))
            self.assertEqual(len(accounts), 5)
            self.assertTrue(all(account.cookie_text.startswith('[')
                                for account in accounts))
            self.assertTrue(all(account.order_no for account in accounts))

    def test_parser_accepts_recognized_header_or_no_header(self):
        for header in (
                list(VENDOR_TEMPLATE_HEADERS),
                ['Email', 'Password', 'Verification URL', 'Cookie JSON']):
            accounts = parse_vendor_workbook(BytesIO(workbook_bytes(
                [header] + demo_rows())))
            self.assertEqual(len(accounts), 2)
            self.assertEqual(accounts[0].row_number, 2)
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows())))
        self.assertEqual(len(accounts), 2)
        self.assertEqual(accounts[0].row_number, 1)

        rows = demo_rows()
        rows[0][1] = 'password'
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(rows)))
        self.assertEqual(accounts[0].password, 'password')

    def test_parser_rejects_mismatched_header_and_extra_business_column(self):
        with self.assertRaisesRegex(EnvBatchError, '表头.*列序'):
            parse_vendor_workbook(BytesIO(workbook_bytes([
                ['邮箱账号', 'Cookie', '接码Key链接', '密码']
            ] + demo_rows())))
        row = demo_rows()[0] + ['unexpected']
        with self.assertRaises(EnvBatchError):
            parse_vendor_workbook(BytesIO(workbook_bytes([row])))

    def test_parser_preserves_cookie_string_exactly(self):
        cookie = ' [ {"name": "sid", "value": "abc"} ] '
        row = ['exact@example.com', 'pass',
               'https://codes.example.test/?orderNo=abcdef', cookie]
        account = parse_vendor_workbook(
            BytesIO(workbook_bytes([row])))[0]
        self.assertEqual(account.cookie_text, cookie)

    def test_parser_accepts_new_vendor_email_key_formats(self):
        cookie = '[{"name":"sid","value":"synthetic"}]'
        rows = [
            ['first@example.com', 'pass-one',
             ('https://mail.example.test/api/boobar-graph?'
              'id=VendorKey_01&email=first%40example.com'), cookie],
            ['second@example.com', 'pass-two',
             ('https://mail.example.test/api/boobar-graph?'
              'email=second%40example.com'), cookie],
            ['third@example.com', 'pass-three',
             'https://codes.example.test/key?orderNo=abcdef12&type=mail',
             cookie],
        ]
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(rows)))
        self.assertEqual(len(accounts), 3)
        self.assertRegex(accounts[0].order_no, r'^[0-9a-f]{64}$')
        self.assertRegex(accounts[1].order_no, r'^[0-9a-f]{64}$')
        self.assertNotEqual(accounts[0].order_no, accounts[1].order_no)
        self.assertEqual(accounts[2].order_no, 'abcdef12')
        self.assertEqual(
            accounts[0].order_no,
            extract_vendor_order_no(rows[0][2], 'FIRST@example.com'))

    def test_vendor_email_path_format_is_normalized_for_deduplication(self):
        cookie = '[{"name":"sid","value":"synthetic"}]'
        query_url = (
            'https://mail.example.test/api/boobar-graph?'
            'email=first%40example.com')
        mail_url = (
            'https://mail.example.test/api?'
            'type=html&mail=first%40example.com')
        path_url = (
            'https://mail.example.test/api/mail-key/first%40example.com')
        account = parse_vendor_workbook(BytesIO(workbook_bytes([[
            'first@example.com', 'pass', path_url, cookie]])))[0]
        self.assertRegex(account.order_no, r'^[0-9a-f]{64}$')
        self.assertEqual(
            account.order_no,
            extract_vendor_order_no(query_url, 'FIRST@example.com'))
        self.assertEqual(
            account.order_no,
            extract_vendor_order_no(mail_url, 'FIRST@example.com'))

    def test_parser_accepts_mail_query_vendor_format_without_header(self):
        cookie = '[{"name":"sid","value":"synthetic"}]'
        rows = [
            ['first@example.com', 'pass-one',
             ('https://mail.example.test/api?'
              'type=html&mail=first%40example.com'), cookie],
            ['second@example.com', 'pass-two',
             ('https://mail.example.test/api?'
              'type=html&mail=second%40example.com'), cookie],
        ]
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(rows)))
        self.assertEqual(len(accounts), 2)
        self.assertEqual([account.row_number for account in accounts], [1, 2])
        self.assertTrue(all(
            re.fullmatch(r'[0-9a-f]{64}', account.order_no)
            for account in accounts))
        self.assertNotEqual(accounts[0].order_no, accounts[1].order_no)

    def test_vendor_email_path_format_rejects_unsafe_or_ambiguous_links(self):
        cases = (
            ('https://mail.example.test/api/mail-key/other%40example.com',
             '邮箱与账号邮箱不一致'),
            ('https://mail.example.test/api/mail-key/first%40example.com?x=1',
             '不含可识别'),
            ('https://user@mail.example.test/api/mail-key/'
             'first%40example.com', '不含可识别'),
            ('https://mail.example.test/first%40example.com', '不含可识别'),
            ('https://mail.example.test/api/mail-key/first%ZZexample.com',
             '路径编码无效'),
        )
        for key_url, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(EnvBatchError, message):
                    extract_vendor_order_no(key_url, 'first@example.com')

        for key_url in (
                'https://user@mail.example.test/key?orderNo=abcdef12',
                'https://mail.example.test/key#?orderNo=abcdef12'):
            with self.assertRaisesRegex(EnvBatchError, '不含可识别'):
                extract_vendor_order_no(key_url, 'first@example.com')

    def test_new_vendor_key_contract_rejects_mismatch_and_guessing(self):
        cookie = '[{"name":"sid","value":"synthetic"}]'
        cases = (
            ('https://mail.example.test/api/boobar-graph?'
             'email=other%40example.com', '邮箱与账号邮箱不一致'),
            ('https://mail.example.test/api/boobar-graph?'
             'email=first%40example.com&extra=value', '参数不符合'),
            ('https://mail.example.test/api/boobar-graph?'
             'email=first%40example.com&email=first%40example.com',
             '参数重复'),
            ('https://mail.example.test/api?'
             'type=html&mail=other%40example.com', '邮箱与账号邮箱不一致'),
            ('https://mail.example.test/api?'
             'type=text&mail=first%40example.com', '参数不符合'),
            ('https://mail.example.test/api?'
             'type=html&mail=first%40example.com&extra=value', '参数不符合'),
            ('https://mail.example.test/api?'
             'type=html&mail=first%40example.com&mail=first%40example.com',
             '参数重复'),
            ('https://mail.example.test/api?'
             'mail=first%40example.com', '参数不符合'),
            ('https://mail.example.test/arbitrary?email=first%40example.com',
             '不含可识别'),
        )
        for key_url, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(EnvBatchError, message):
                    parse_vendor_workbook(BytesIO(workbook_bytes([[
                        'first@example.com', 'pass', key_url, cookie]])))

    def test_new_vendor_id_remains_a_strict_duplicate_key(self):
        cookie = '[{"name":"sid","value":"synthetic"}]'
        rows = [
            ['first@example.com', 'pass-one',
             ('https://mail.example.test/api/boobar-graph?'
              'id=SameKey_001&email=first%40example.com'), cookie],
            ['second@example.com', 'pass-two',
             ('https://mail.example.test/api/boobar-graph?'
              'id=SameKey_001&email=second%40example.com'), cookie],
        ]
        with self.assertRaisesRegex(EnvBatchError, '号商单号重复'):
            parse_vendor_workbook(BytesIO(workbook_bytes(rows)))

    def test_assignment_and_daily_continuation_are_per_buyer(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        existing = [
            {'containerName': 'XG-MX-0819-003'},
            {'containerName': 'ZH-MX-0819-009'},
        ]
        plan = build_batch_plan(
            accounts, '1:新刚,1:志恒', existing, purchase_date='20260819')
        self.assertEqual(plan[0].env_name, 'XG-MX-0819-004')
        self.assertEqual(plan[1].env_name, 'ZH-MX-0819-010')

        # 旧「采购-」格式不参与新代号续排
        us_plan = build_batch_plan(
            accounts, '1:新刚,1:志恒',
            [{'containerName': 'XG-US-0819-006'},
             {'containerName': '采购-熊-US-0819-099'}],
            site='US', purchase_date='20260819')
        self.assertEqual(us_plan[0].env_name, 'XG-US-0819-007')
        self.assertEqual(us_plan[1].env_name, 'ZH-US-0819-001')

    def test_cloud_reserved_names_override_local_suffix_scan(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        planned = {
            accounts[0].account_id: 'XG-MX-0819-101',
            accounts[1].account_id: 'ZH-MX-0819-205',
        }
        plan = build_batch_plan(
            accounts, '1:新刚,1:志恒',
            existing_envs=[{'containerName': 'XG-MX-0819-999'}],
            purchase_date='20260819', planned_env_names=planned)
        self.assertEqual(
            [row.env_name for row in plan],
            ['XG-MX-0819-101', 'ZH-MX-0819-205'])
        with self.assertRaisesRegex(EnvBatchError, '不一致'):
            build_batch_plan(
                accounts, '1:新刚,1:志恒', purchase_date='20260819',
                planned_env_names={
                    accounts[0].account_id: 'ZH-MX-0819-101',
                    accounts[1].account_id: 'ZH-MX-0819-205',
                })

    def test_cloud_reserved_plan_skips_growing_global_environment_scan(self):
        hub = FakeHub()
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None)
        rows = runner.prepare(
            accounts, '1:新刚',
            planned_env_names={
                accounts[0].account_id: 'XG-MX-0819-321',
            })
        self.assertEqual(rows[0].env_name, 'XG-MX-0819-321')
        self.assertEqual(hub.env_list_calls, 1)

    def test_ip_verification_sample_honors_requested_count_and_all_mode(self):
        buyers = ['新刚', '志恒', '康德', '宇航']
        rows = []
        for index in range(200):
            account = BuyerAccount(
                row_number=index + 1,
                email='verify%d@example.com' % index,
                password='password',
                key_url='https://codes.example.test/get?orderNo=%08x' % (
                    index + 1),
                cookie_text='[{"name":"sid","value":"cookie"}]',
                order_no='%08x' % (index + 1),
                buyer=buyers[index % len(buyers)],
            )
            rows.append(BatchPlanItem(
                account=account,
                env_name='XG-MX-0903-%03d' % (index + 1),
                container_code=str(10000 + index),
                state='done',
            ))

        sampled = BatchEnvOrchestrator._verification_sample(rows, 17)
        self.assertEqual(len(sampled), 17)
        self.assertEqual(
            {row.account.buyer for row in sampled}, set(buyers))
        self.assertEqual(len({id(row) for row in sampled}), 17)
        self.assertEqual(
            BatchEnvOrchestrator._verification_sample(rows, 200), rows)

    def test_assignment_roster_validation_and_code_normalization(self):
        self.assertEqual(parse_assignment('2:XG,1:志恒', 3),
                         [(2, '新刚'), (1, '志恒')])
        self.assertEqual(parse_assignment('1:xg', 1), [(1, '新刚')])
        self.assertEqual(normalize_buyer('kd'), '康德')
        with self.assertRaisesRegex(EnvBatchError, '不在名单内'):
            parse_assignment('1:甲', 1)
        with self.assertRaisesRegex(EnvBatchError, '重复'):
            parse_assignment('1:XG,1:新刚', 2)
        self.assertEqual(validate_assignment_template('2:XG,1:志恒'),
                         '2:XG,1:志恒')
        with self.assertRaises(EnvBatchError):
            validate_assignment_template('1:Operator-A')
        roster_codes = {code for _name, code in BUYER_ROSTER}
        self.assertEqual(roster_codes, {'XG', 'ZH', 'KD', 'YH'})

    def test_existing_order_remark_makes_completed_rerun_idempotent(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows()[:1])))
        existing = [{
            'containerName': 'XG-MX-0819-007',
            'containerCode': '7007', 'serialNumber': 1006,
            'remark': '邮箱接码:https://redacted | 单号:a1b2c3 | 采购员:新刚 | 购买:20260819',
        }]
        plan = build_batch_plan(
            accounts, '1:新刚', existing, purchase_date='20260819')
        self.assertEqual(plan[0].state, 'done')
        self.assertEqual(plan[0].env_name, 'XG-MX-0819-007')
        self.assertEqual(plan[0].completed_steps,
                         {'env_created', 'cookie_imported', 'account_bound',
                          'remarked', 'done'})
        with self.assertRaisesRegex(EnvBatchError, '另一站点'):
            build_batch_plan(
                accounts, '1:新刚', existing, site='US',
                purchase_date='20260819')

    def test_global_guard_keeps_same_group_idempotent_recovery(self):
        hub = FakeHub()
        hub.envs.append({
            'containerName': 'XG-MX-0820-007',
            'containerCode': 'same-1', 'serialNumber': 2007,
            'tagName': TEST_TAG,
            'remark': ('邮箱接码:https://redacted | 单号:a1b2c3 | '
                       '采购员:新刚 | 购买:20260820'),
        })
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260820', sleep_fn=lambda _seconds: None)
        rows = runner.prepare(accounts, '1:新刚')
        self.assertTrue(rows[0].recovered_existing)
        self.assertEqual(rows[0].state, 'done')
        runner.run()
        self.assertFalse(any(call[0] in {
            'create', 'cookie', 'account', 'remark'} for call in hub.calls))

    def test_prepare_rejects_order_already_present_in_other_group(self):
        hub = FakeHub()
        hub.envs.append({
            'containerName': 'XG-MX-0820-001',
            'containerCode': 'other-1', 'serialNumber': 2001,
            'tagName': 'Other-Purchase',
            'remark': ('邮箱接码:https://redacted | 单号:a1b2c3 | '
                       '采购员:新刚 | 购买:20260820'),
        })
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260820', sleep_fn=lambda _seconds: None)
        with self.assertRaisesRegex(EnvBatchError, '其他分组'):
            runner.prepare(accounts, '1:新刚')
        self.assertFalse(any(call[0] in {
            'create', 'cookie', 'account', 'remark'} for call in hub.calls))

    def test_resolution_pool_tracks_weights_and_caps_size(self):
        rng = random.Random(20260819)
        counts = {pair[:2]: 0 for pair in RES_POOL}
        for _index in range(10000):
            resolution = choose_resolution(rng)
            counts[resolution] += 1
            self.assertLessEqual(resolution[0], 2560)
            self.assertLessEqual(resolution[1], 1440)
        for width, height, weight in RES_POOL:
            observed = counts[(width, height)] / 10000.0
            self.assertLess(abs(observed - weight), 0.025)
        body = build_env_create_body(
            'XG-MX-0819-001', rng=rng,
            proxy_link=TEST_PROXY, purchase_tag=TEST_TAG)
        self.assertNotIn('coreVersion', body)
        self.assertEqual(body['advancedBo']['languageType'], 0)
        self.assertEqual(body['tagName'], TEST_TAG)
        self.assertEqual(body['linkCode'],
                         'https://proxy.example.test/MX')
        us_body = build_env_create_body(
            'XG-US-0819-001', site='US', rng=rng,
            proxy_link=TEST_PROXY, purchase_tag=TEST_TAG)
        self.assertEqual(us_body['linkCode'],
                         'https://proxy.example.test/US')
        self.assertEqual(normalize_env_site('us'), 'US')
        with self.assertRaisesRegex(EnvBatchError, 'MX.*US'):
            normalize_env_site('CA')

    def test_us_orchestrator_captures_site_for_proxy_name_and_binding(self):
        hub = FakeHub()
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            site='US', purchase_date='20260819',
            sleep_fn=lambda _seconds: None)
        runner.prepare(accounts, '1:新刚')
        runner.run()
        self.assertEqual(runner.rows[0].env_name, 'XG-US-0819-001')
        create = next(call[1] for call in hub.calls if call[0] == 'create')
        account = next(call for call in hub.calls if call[0] == 'account')
        self.assertEqual(create['linkCode'], 'https://proxy.example.test/US')
        self.assertEqual(account[-1], 'US')

    def test_preflight_requires_exact_group_and_never_returns_proxy(self):
        hub = FakeHub()
        ready = envbatch_preflight(hub, TEST_TAG, TEST_PROXY)
        self.assertTrue(ready['ready'])
        self.assertTrue(ready['groupFound'])
        self.assertNotIn(TEST_PROXY, json.dumps(ready))
        us_ready = envbatch_preflight(
            hub, TEST_US_TAG, TEST_PROXY, site='US')
        self.assertTrue(us_ready['ready'])
        self.assertEqual(us_ready['site'], 'US')

        wrong_us_group = envbatch_preflight(
            hub, TEST_TAG, TEST_PROXY, site='US')
        self.assertFalse(wrong_us_group['ready'])
        self.assertIn('美国站不能使用墨西哥', wrong_us_group['message'])
        wrong_mx_group = envbatch_preflight(
            hub, TEST_US_TAG, TEST_PROXY, site='MX')
        self.assertFalse(wrong_mx_group['ready'])
        self.assertIn('墨西哥站不能使用美国', wrong_mx_group['message'])

        no_proxy = envbatch_preflight(hub, TEST_TAG, '')
        self.assertFalse(no_proxy['ready'])
        self.assertTrue(no_proxy['hubConnected'])
        self.assertTrue(no_proxy['groupFound'])
        self.assertFalse(no_proxy['proxyConfigured'])

        missing = envbatch_preflight(hub, 'MX-Purchas', TEST_PROXY)
        self.assertFalse(missing['ready'])
        self.assertFalse(missing['groupFound'])
        self.assertNotIn(TEST_PROXY, json.dumps(missing))

        invalid = envbatch_preflight(
            hub, TEST_TAG, 'https://proxy.example.test/{secret}')
        self.assertFalse(invalid['ready'])
        self.assertNotIn('proxy.example.test', json.dumps(invalid))
        with self.assertRaisesRegex(EnvBatchError, '换行'):
            validate_purchase_tag(TEST_TAG + '\n')

    def test_create_error_never_exposes_proxy_link(self):
        secret_proxy = 'https://secret-user:secret-pass@proxy.example.test/{region}'
        hub = FakeHub(fail_create_with_link=True)
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes([demo_rows()[0]])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=secret_proxy,
            purchase_date='20260819', sleep_fn=lambda _seconds: None)
        rows = runner.prepare(accounts, '1:新刚')
        runner.run()
        rendered = json.dumps(rows[0].public_dict(), ensure_ascii=False)
        self.assertEqual(rows[0].state, 'failed')
        self.assertNotIn(secret_proxy, rendered)
        self.assertNotIn('secret-user', rendered)
        self.assertNotIn('secret-pass', rendered)

    def test_hub_cookie_import_requires_and_preserves_string(self):
        api = HubStudioApi()
        calls = []
        api._post = lambda path, body: calls.append((path, body)) or {}
        cookie = '[{"name":"sid","value":"unchanged"}]'
        api.env_import_cookie(123, cookie)
        self.assertEqual(calls[0], (
            '/env/import-cookie',
            {'containerCode': '123', 'cookie': cookie}))
        with self.assertRaises(TypeError):
            api.env_import_cookie(123, [])
        api.container_add_account(
            123, 'us@example.com', 'secret-pass', site='US')
        us_body = calls[-1][1]
        self.assertEqual(us_body['siteAlias'], '希音美国站')
        self.assertEqual(us_body['domainName'], 'https://us.shein.com')
        with self.assertRaisesRegex(ValueError, 'MX.*US'):
            api.container_add_account(
                123, 'ca@example.com', 'secret-pass', site='CA')

    def test_hub_headless_start_uses_official_read_only_parameters(self):
        api = HubStudioApi()
        calls = []
        api._post = lambda path, body: calls.append((path, body)) or {
            'ip': '203.0.113.10'}
        result = api.browser_start(123, headless=True)
        self.assertEqual(result['ip'], '203.0.113.10')
        self.assertEqual(calls, [('/browser/start', {
            'containerCode': '123',
            'isHeadless': True,
            'isWebDriverReadOnlyMode': True,
            'args': ['--headless=new'],
        })])

    def test_hub_rate_limit_uses_shared_cooldown_then_retries(self):
        class FakeResponse(object):
            def __init__(self, payload):
                self.payload = payload

            def __enter__(self):
                return self

            def __exit__(self, _type, _value, _traceback):
                return False

            def read(self):
                return json.dumps(self.payload).encode('utf-8')

        class FakeOpener(object):
            def __init__(self):
                self.responses = [
                    {'code': 'E010205', 'msg': '请求太频繁，请稍后再试'},
                    {'code': 0, 'msg': 'ok', 'data': {'list': [], 'total': 0}},
                ]

            def open(self, _request, timeout=None):
                self.timeout = timeout
                return FakeResponse(self.responses.pop(0))

        class FakeGate(object):
            def __init__(self):
                self.cooldowns = []

            def request(self):
                return nullcontext()

            def defer_requests(self, seconds):
                self.cooldowns.append(seconds)

        gate = FakeGate()
        api = HubStudioApi(runtime_gate=gate, retries=3)
        opener = FakeOpener()
        with patch('purchase_tool.hub_api.OPENER', opener), \
                patch('purchase_tool.hub_api.time.sleep') as fallback_sleep:
            rows = api.env_list()

        self.assertEqual(rows, [])
        self.assertEqual(gate.cooldowns, [2.0])
        fallback_sleep.assert_not_called()

    def test_ip_probe_never_uses_visible_browser_mode(self):
        class ProbeHub(object):
            def __init__(self):
                self.calls = []

            def browser_start(self, code, headless=False):
                self.calls.append(('start', str(code), headless))
                return {'ip': '203.0.113.11'}

            def browser_stop(self, code):
                self.calls.append(('stop', str(code)))

        hub = ProbeHub()
        result = probe_env_ip(
            hub, 'US', 'XG-US-0824-001', '123',
            lambda _ip: {'countryCode': 'US', 'country': 'United States'})
        self.assertTrue(result['ok'])
        self.assertEqual(hub.calls, [
            ('start', '123', True), ('stop', '123')])

    def test_ip_probe_translates_missing_browser_core_to_actionable_error(self):
        class MissingCoreHub(object):
            def browser_start(self, _code, headless=False):
                raise HubApiError(
                    'HubStudio Local API code=-10007',
                    api_code=-10007)

            def browser_stop(self, _code):
                raise AssertionError('未启动成功的环境不能执行关闭')

        result = probe_env_ip(
            MissingCoreHub(), 'US', 'XG-US-0901-011', '123',
            lambda _ip: {})

        self.assertFalse(result['ok'])
        self.assertEqual(result['errorCode'], 'hub_browser_core_missing')
        self.assertIn('浏览器内核不存在', result['error'])
        self.assertIn('安装可用内核', result['error'])
        self.assertNotIn('-10007', result['error'])

    def test_ip_probe_explains_missing_ip_instead_of_unknown(self):
        class NoIpHub(object):
            def browser_start(self, _code, headless=False):
                return {}

            def browser_stop(self, _code):
                return {}

        result = probe_env_ip(
            NoIpHub(), 'US', 'KD-US-0901-021', '123', lambda _ip: {})

        self.assertFalse(result['ok'])
        self.assertEqual(result['errorCode'], 'hub_ip_missing')
        self.assertIn('未返回出口 IP', result['error'])

    def test_ip_probe_preserves_geo_lookup_failure_reason(self):
        class ProbeHub(object):
            def browser_start(self, _code, headless=False):
                return {'ip': '203.0.113.11'}

            def browser_stop(self, _code):
                return {}

        result = probe_env_ip(
            ProbeHub(), 'US', 'KD-US-0901-021', '123',
            lambda _ip: (_ for _ in ()).throw(
                EnvBatchError('IP 归属地查询失败')))

        self.assertFalse(result['ok'])
        self.assertEqual(result['errorCode'], 'ip_geo_lookup_failed')
        self.assertEqual(result['ip'], '203.0.113.11')
        self.assertIn('归属地查询失败', result['error'])

    def test_hub_unfiltered_env_list_paginates_across_all_groups(self):
        api = HubStudioApi()
        calls = []

        def fake_post(path, body):
            calls.append((path, dict(body)))
            current = body['current']
            return {
                'total': 201,
                'list': ([{'containerCode': str(index)}
                          for index in range(200)]
                         if current == 1 else [{'containerCode': '200'}]),
            }

        api._post = fake_post
        rows = api.env_list()
        self.assertEqual(len(rows), 201)
        self.assertEqual([body['current'] for _path, body in calls], [1, 2])
        self.assertTrue(all('tagNames' not in body
                            for _path, body in calls))

    def test_failed_step_resumes_without_repeating_completed_steps(self):
        source = workbook_bytes(demo_rows()[:1])
        accounts = parse_vendor_workbook(BytesIO(source))
        hub = FakeHub(fail_cookie_once=True)
        with tempfile.TemporaryDirectory() as tmp:
            store = ResumeStateStore(
                batch_fingerprint(source, '1:新刚', 'MX', '20260819'), tmp)
            first = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', state_store=store,
                sleep_fn=lambda _seconds: None)
            first.prepare(accounts, '1:新刚')
            first.run()
            self.assertEqual(first.rows[0].state, 'failed')
            self.assertEqual(first.rows[0].error_step, 'cookie_imported')
            self.assertTrue(store.path.exists())
            state_text = store.path.read_text(encoding='utf-8')
            for secret in ('secret-pass-a', 'secret-cookie-a',
                           'codes.example.test', 'alpha@example.com'):
                self.assertNotIn(secret, state_text)
            if os.name != 'nt':
                self.assertEqual(store.path.stat().st_mode & 0o777, 0o600)

            second = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', state_store=store,
                sleep_fn=lambda _seconds: None)
            second.prepare(accounts, '1:新刚')
            second.run()
            self.assertEqual(second.rows[0].state, 'done')
            self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 1)
            create_body = next(call[1] for call in hub.calls
                               if call[0] == 'create')
            self.assertEqual(create_body['tagName'], TEST_TAG)
            self.assertEqual(create_body['linkCode'],
                             'https://proxy.example.test/MX')
            self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 2)
            self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 1)
            self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 1)
            self.assertFalse(store.path.exists())

    def test_retry_failed_retries_all_failed_rows_without_repeating_steps(self):
        class FailCookieTwiceHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.cookie_failures = 2

            def env_import_cookie(self, code, cookie_text):
                self.calls.append(('cookie', str(code), cookie_text))
                if self.cookie_failures:
                    self.cookie_failures -= 1
                    raise RuntimeError('synthetic cookie failure')
                return {}

        source = workbook_bytes(demo_rows())
        accounts = parse_vendor_workbook(BytesIO(source))
        hub = FailCookieTwiceHub()
        with tempfile.TemporaryDirectory() as tmp:
            store = ResumeStateStore(
                batch_fingerprint(source, '2:新刚', 'MX', '20260819'), tmp)
            runner = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', state_store=store,
                sleep_fn=lambda _seconds: None, max_workers=2)
            runner.prepare(accounts, '2:新刚')
            runner.run()
            self.assertEqual(
                [row.state for row in runner.rows], ['failed', 'failed'])
            self.assertTrue(all(
                row.error_step == 'cookie_imported' for row in runner.rows))

            retried = runner.retry_failed()

            self.assertEqual(len(retried), 2)
            self.assertEqual(
                [row.state for row in runner.rows], ['done', 'done'])
            self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 2)
            self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 4)
            self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 2)
            self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 2)
            self.assertFalse(store.path.exists())
            with self.assertRaisesRegex(EnvBatchError, '没有失败行'):
                runner.retry_failed()

    def test_retry_failed_honors_safe_stop_for_queued_failed_rows(self):
        class BlockingCookieHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.block_cookie = False
                self.cookie_started = threading.Event()
                self.release_cookie = threading.Event()

            def env_import_cookie(self, code, cookie_text):
                if self.block_cookie and not self.cookie_started.is_set():
                    self.cookie_started.set()
                    if not self.release_cookie.wait(2):
                        raise RuntimeError('test timed out waiting for cookie')
                return super().env_import_cookie(code, cookie_text)

        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        hub = BlockingCookieHub()
        stop_event = threading.Event()
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1, stop_event=stop_event)
        runner.prepare(accounts, '2:新刚')
        for row in runner.rows:
            row.completed_steps.add('env_created')
            row.container_code = 'existing-' + row.account.account_id[:8]
            row.serial_number = row.account.row_number
            row.state = 'failed'
            row.error_step = 'cookie_imported'
        hub.block_cookie = True

        worker = threading.Thread(target=runner.retry_failed)
        worker.start()
        self.assertTrue(hub.cookie_started.wait(1))
        stop_event.set()
        hub.release_cookie.set()
        worker.join(3)

        self.assertFalse(worker.is_alive())
        self.assertEqual(
            [row.state for row in runner.rows], ['done', 'stopped'])
        self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 1)

    def test_safe_stop_finishes_active_row_and_resumes_unstarted_rows(self):
        class BlockingFirstCreateHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.first_started = threading.Event()
                self.release_first = threading.Event()

            def env_create(self, body):
                if not self.first_started.is_set():
                    self.first_started.set()
                    if not self.release_first.wait(2):
                        raise RuntimeError('test timed out waiting to release create')
                return super().env_create(body)

        source = workbook_bytes(demo_rows())
        accounts = parse_vendor_workbook(BytesIO(source))
        hub = BlockingFirstCreateHub()
        stop_event = threading.Event()
        with tempfile.TemporaryDirectory() as tmp:
            store = ResumeStateStore(
                batch_fingerprint(source, '2:新刚', 'MX', '20260819'), tmp)
            first = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', state_store=store,
                sleep_fn=lambda _seconds: None, max_workers=1,
                stop_event=stop_event)
            first.prepare(accounts, '2:新刚')
            worker = threading.Thread(target=first.run)
            worker.start()
            self.assertTrue(hub.first_started.wait(1))
            stop_event.set()
            hub.release_first.set()
            worker.join(3)
            self.assertFalse(worker.is_alive())

            self.assertEqual(
                [row.state for row in first.rows], ['done', 'stopped'])
            self.assertEqual(
                first.rows[0].completed_steps,
                {'env_created', 'cookie_imported', 'account_bound',
                 'remarked', 'done'})
            self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 1)
            self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 1)
            self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 1)
            self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 1)
            self.assertTrue(store.path.exists())

            resumed = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', state_store=store,
                sleep_fn=lambda _seconds: None, max_workers=1)
            resumed.prepare(accounts, '2:新刚')
            resumed.run()
            self.assertEqual(
                [row.state for row in resumed.rows], ['done', 'done'])
            self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 2)
            self.assertFalse(store.path.exists())

    def test_stop_compensation_deletes_only_environments_created_by_run(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        hub = FakeHub()
        stop_event = threading.Event()
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1, stop_event=stop_event)
        runner.prepare(accounts, '2:新刚')

        runner._run_one(runner.rows[0])
        stop_event.set()
        runner._run_one(runner.rows[1])
        self.assertTrue(runner.rows[0].created_in_run)
        self.assertEqual(runner.rows[1].state, 'stopped')

        cleaned = runner.rollback_created_environments()

        self.assertEqual(cleaned, [runner.rows[0]])
        self.assertEqual(runner.rows[0].state, 'rolled_back')
        self.assertEqual(runner.rows[0].cleanup_status, 'deleted')
        self.assertEqual(runner.rows[1].cleanup_status, 'not_required')
        self.assertEqual(hub.envs, [])

    def test_stop_compensation_never_deletes_recovered_environment(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        hub = FakeHub()
        hub.envs.append({
            'containerName': 'XG-MX-0819-001',
            'containerCode': '8999', 'serialNumber': 1099,
            'tagName': TEST_TAG,
            'remark': format_remark(accounts[0], '20260819'),
        })
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1)
        runner.prepare(accounts, '2:新刚')
        recovered = next(row for row in runner.rows if row.recovered_existing)

        self.assertEqual(runner.rollback_created_environments(), [])
        self.assertIsNotNone(hub.env_lookup(
            container_code=recovered.container_code))

        blocked = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1,
            reject_existing_account_refs={recovered.account.account_id})
        with self.assertRaisesRegex(
                EnvBatchError, '上次销毁失败.*任何写入前'):
            blocked.prepare(accounts, '2:新刚')
        self.assertEqual(
            sum(call[0] == 'create' for call in hub.calls), 0)

    def test_stop_compensation_reconciles_delete_timeout_without_duplicate(self):
        class TimeoutAfterDeleteHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.delete_calls = 0

            def env_delete(self, container_codes):
                self.delete_calls += 1
                super().env_delete(container_codes)
                raise HubApiError(
                    'HubStudio Local API 请求超时',
                    'hubstudio_local_api_timeout')

        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        hub = TimeoutAfterDeleteHub()
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1)
        runner.prepare(accounts[:1], '1:新刚')
        runner.run()

        runner.rollback_created_environments()

        self.assertEqual(hub.delete_calls, 1)
        self.assertEqual(runner.rows[0].cleanup_status, 'deleted')
        self.assertEqual(runner.rows[0].cleanup_attempts, 1)

    def test_backup_safe_stop_does_not_start_queued_rows(self):
        class BlockingFirstCreateHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.first_started = threading.Event()
                self.release_first = threading.Event()

            def env_create(self, body):
                if not self.first_started.is_set():
                    self.first_started.set()
                    if not self.release_first.wait(2):
                        raise RuntimeError('test timed out waiting to release create')
                return super().env_create(body)

        hub = BlockingFirstCreateHub()
        stop_event = threading.Event()
        runner = BackupEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            sleep_fn=lambda _seconds: None, max_workers=1,
            stop_event=stop_event)
        runner.prepare('新刚', 2, '备用', '20260819')
        worker = threading.Thread(target=runner.run)
        worker.start()
        self.assertTrue(hub.first_started.wait(1))
        stop_event.set()
        hub.release_first.set()
        worker.join(3)
        self.assertFalse(worker.is_alive())
        self.assertEqual([row.state for row in runner.rows], ['done', 'stopped'])
        self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 1)
        result = backup_result_tsv_bytes(runner.rows, 'MX').decode('utf-8-sig')
        self.assertIn('已停止', result)

    def test_remark_and_mapping_export_contract(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows()[:1])))
        plan = build_batch_plan(
            accounts, '1:新刚', purchase_date='20260819')
        row = plan[0]
        row.serial_number = 1004
        row.binding_time = '2026-08-19 10:00:00'
        row.state = 'done'
        remark = format_remark(row.account, '20260819')
        self.assertIn('邮箱接码:', remark)
        self.assertIn('单号:a1b2c3', remark)
        self.assertIn('采购员:新刚', remark)
        self.assertIn('购买:20260819', remark)

        output = mapping_workbook_bytes(plan)
        with zipfile.ZipFile(BytesIO(output)) as archive:
            text = ''.join(
                archive.read(name).decode('utf-8', errors='ignore')
                for name in archive.namelist() if name.endswith('.xml'))
        self.assertIn('绑定映射清单', text)
        self.assertIn('alpha@example.com', text)
        for secret in ('secret-pass-a', 'secret-cookie-a',
                       'codes.example.test'):
            self.assertNotIn(secret, text)

    def test_cookie_domain_validates_account_site(self):
        def rows_with_cookie(cookie):
            row = ['site@example.com', 'pass',
                   'https://codes.example.test/?orderNo=abc123', cookie]
            return workbook_bytes([row])
        mx_cookie = '[{"name":"sid","domain":".shein.com.mx"}]'
        us_cookie = '[{"name":"sid","domain":".us.shein.com"}]'
        from purchase_tool.env_batch import (count_mixed_site_accounts,
                                              detect_cookie_site,
                                              validate_accounts_site)
        self.assertEqual(detect_cookie_site(mx_cookie), 'MX')
        self.assertEqual(detect_cookie_site(us_cookie), 'US')
        self.assertIsNone(detect_cookie_site('[{"name":"sid"}]'))
        self.assertEqual(
            detect_cookie_site(mx_cookie + us_cookie), 'CONFLICT')

        # 站点一致放行；纯错站拒收（无 Cookie 标记不拦截）。
        accounts = parse_vendor_workbook(
            BytesIO(rows_with_cookie(mx_cookie)))
        plan = build_batch_plan(accounts, '1:新刚', purchase_date='20260819')
        self.assertTrue(plan[0].env_name.startswith('XG-MX-'))
        with self.assertRaisesRegex(EnvBatchError, 'MX.*US.*不一致'):
            build_batch_plan(accounts, '1:新刚', site='US',
                             purchase_date='20260819')
        us_accounts = parse_vendor_workbook(
            BytesIO(rows_with_cookie(us_cookie)))
        build_batch_plan(us_accounts, '1:新刚', site='US',
                         purchase_date='20260819')
        plain = parse_vendor_workbook(
            BytesIO(rows_with_cookie('[{"name":"sid"}]')))
        build_batch_plan(plain, '1:新刚', site='US', purchase_date='20260819')
        both = parse_vendor_workbook(BytesIO(rows_with_cookie(
            '[{"name":"a","domain":".shein.com.mx"},'
            '{"name":"b","domain":".us.shein.com"}]')))
        mixed_plan = build_batch_plan(
            both, '1:新刚', purchase_date='20260819')
        self.assertEqual(len(mixed_plan), 1)
        self.assertEqual(mixed_plan[0].account.cookie_text,
                         both[0].cookie_text)

        mixed_cookie = (
            '[{"name":"a","domain":".shein.com.mx"},'
            '{"name":"b","domain":".us.shein.com"}]')
        aggregate_rows = [
            ['mix1@example.com', 'pass',
             'https://codes.example.test/?orderNo=abc001', mixed_cookie],
            ['ok@example.com', 'pass',
             'https://codes.example.test/?orderNo=abc002', mx_cookie],
            ['mix3@example.com', 'pass',
             'https://codes.example.test/?orderNo=abc003', mixed_cookie],
            ['mix4@example.com', 'pass',
             'https://codes.example.test/?orderNo=abc004', mixed_cookie],
        ]
        aggregate_accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(aggregate_rows)))
        self.assertEqual(count_mixed_site_accounts(aggregate_accounts), 3)
        # 通用台账导入仍可使用严格模式；仅环境创建显式兼容混合登录态。
        with self.assertRaisesRegex(
                EnvBatchError, '共 3 行数据异常.*第1行、第3-4行'):
            validate_accounts_site(aggregate_accounts, 'MX')
        aggregate_plan = build_batch_plan(
            aggregate_accounts, '4:新刚', purchase_date='20260819')
        self.assertEqual(len(aggregate_plan), 4)

    def test_encrypted_cloud_plan_round_trip_is_strict_and_site_bound(self):
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(demo_rows())))
        payload = serialize_buyer_accounts(accounts)
        restored = deserialize_buyer_accounts(payload, site='MX')
        self.assertEqual(
            [(item.email, item.password, item.cookie_text) for item in restored],
            [(item.email, item.password, item.cookie_text) for item in accounts])

        altered = [dict(item) for item in payload]
        altered[0]['orderNo'] = 'deadbeef'
        with self.assertRaisesRegex(EnvBatchError, '号商单号校验失败'):
            deserialize_buyer_accounts(altered, site='MX')

        us_payload = [dict(payload[0])]
        us_payload[0]['cookie'] = '[{"domain":".us.shein.com"}]'
        with self.assertRaisesRegex(EnvBatchError, 'Cookie 站点校验失败'):
            deserialize_buyer_accounts(us_payload, site='MX')

    def test_bound_parallel_run_creates_all_rows(self):
        rows = []
        for i in range(5):
            rows.append([
                'par%d@example.com' % i, 'secret-pass',
                'https://codes.example.test/get?orderNo=%s' % ('a' * 13 + '%02x' % i),
                '[{"name":"sid"}]'])
        accounts = parse_vendor_workbook(BytesIO(workbook_bytes(rows)))
        hub = FakeHub()
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _s: None,
            max_workers=3)
        runner.prepare(accounts, '5:新刚')
        runner.run()
        self.assertEqual({row.state for row in runner.rows}, {'done'})
        self.assertEqual(len(hub.envs), 5)
        self.assertEqual(len({env['containerCode'] for env in hub.envs}), 5)
        self.assertEqual(
            [row.env_name for row in runner.rows],
            ['XG-MX-0819-%03d' % n for n in range(1, 6)])

    def test_bound_run_reuses_snapshot_and_only_uses_targeted_lookups(self):
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows())))
        hub = FakeHub()
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=2)

        runner.prepare(accounts, '2:新刚')
        runner.run()

        # 准备阶段固定为正式分组 + 全局各一次；执行阶段不得再做全量翻页。
        self.assertEqual(hub.env_list_calls, 2)
        self.assertEqual(len(hub.env_lookup_calls), 2)
        self.assertEqual(
            sum(bool(call['containerName'])
                for call in hub.env_lookup_calls), 0)
        self.assertEqual(
            sum(bool(call['containerCode'])
                for call in hub.env_lookup_calls), 2)
        self.assertEqual({row.state for row in runner.rows}, {'done'})

    def test_environment_snapshot_indexes_name_and_vendor_order(self):
        selected = [{
            'containerName': 'XG-MX-0819-001',
            'containerCode': 'safe-1',
            'serialNumber': 1001,
            'remark': '单号:a1b2c3 | 采购员:新刚',
        }]
        index = EnvironmentSnapshotIndex(selected, list(selected))

        self.assertEqual(
            index.find_by_name('XG-MX-0819-001')['containerCode'],
            'safe-1')
        self.assertEqual(
            index.by_order['a1b2c3'][0]['containerName'],
            'XG-MX-0819-001')

    def test_backup_names_share_daily_serial_and_skip_existing(self):
        existing = [
            {'containerName': 'XG-MX-0819-003'},
            {'containerName': 'XG-MX-0819-004'},
            {'containerName': '采购-熊-MX-0819-099'},   # 旧格式不参与续排
        ]
        names = backup_env_names(
            existing, '新刚', 2, '备用', 'MX', '20260819')
        self.assertEqual(names, ['XG-MX-0819-005', 'XG-MX-0819-006'])
        test_names = backup_env_names(
            [{'containerName': 'ZH-US-测试-02'}], '志恒', 2, '测试', 'US',
            '20260819')
        self.assertEqual(test_names, ['ZH-US-测试-03', 'ZH-US-测试-04'])
        with self.assertRaisesRegex(EnvBatchError, '1-25'):
            backup_env_names([], '新刚', 26, '备用', 'MX', '20260819')
        with self.assertRaisesRegex(EnvBatchError, '备用 或 测试'):
            backup_env_names([], '新刚', 1, '临时', 'MX', '20260819')
        with self.assertRaisesRegex(EnvBatchError, '不在名单内'):
            backup_env_names([], '甲', 1, '备用', 'MX', '20260819')
        with self.assertRaisesRegex(EnvBatchError, 'YYYYMMDD'):
            backup_env_names([], '新刚', 1, '备用', 'MX', '0819')

    def test_backup_orchestrator_creates_and_remarks_without_binding(self):
        hub = FakeHub()
        runner = BackupEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY, site='MX',
            sleep_fn=lambda _seconds: None)
        runner.prepare('新刚', 2, '备用', '20260819')
        rows = runner.run()
        self.assertEqual([row.env_name for row in rows],
                         ['XG-MX-0819-001', 'XG-MX-0819-002'])
        self.assertEqual({row.state for row in rows}, {'done'})
        create_calls = [call for call in hub.calls if call[0] == 'create']
        self.assertEqual(len(create_calls), 2)
        self.assertEqual(create_calls[0][1]['linkCode'],
                         'https://proxy.example.test/MX')
        self.assertNotIn('coreVersion', create_calls[0][1])
        remark_calls = [call for call in hub.calls if call[0] == 'remark']
        self.assertEqual(len(remark_calls), 2)
        self.assertEqual({call[3] for call in remark_calls}, {'备用环境'})
        # 备用模式零 Cookie、零绑号
        self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 0)
        self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 0)
        # 重跑按名幂等：已存在的环境不再重复创建
        runner2 = BackupEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY, site='MX',
            sleep_fn=lambda _seconds: None)
        runner2.prepare('新刚', 2, '备用', '20260819')
        self.assertEqual([row.env_name for row in runner2.rows],
                         ['XG-MX-0819-003', 'XG-MX-0819-004'])

    def test_bound_adoption_refuses_nonempty_remark_env(self):
        # 模拟计划后、执行前另一机器/批次占用同名环境（同步延迟或并发撞名）
        for occupied_remark in ('备用环境', '测试环境',
                                '邮箱接码:https://x | 单号:abc | 采购员:新刚'):
            hub = FakeHub()
            accounts = parse_vendor_workbook(
                BytesIO(workbook_bytes(demo_rows()[:1])))
            runner = BatchEnvOrchestrator(
                hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
                purchase_date='20260819', sleep_fn=lambda _s: None)
            runner.prepare(accounts, '1:新刚')   # 计划 XG-MX-0819-001
            hub.envs.append({
                'containerName': 'XG-MX-0819-001',
                'containerCode': '8001', 'serialNumber': 2001,
                'remark': occupied_remark})
            runner.rows[0].state = 'failed'
            runner.run()
            self.assertEqual(runner.rows[0].state, 'failed')
            self.assertIn('拒绝收养', runner.rows[0].error)
            # 拒绝收养 = 零写入（不建、不导 Cookie、不绑号、不改备注）
            self.assertEqual(sum(call[0] in {
                'create', 'cookie', 'account', 'remark'}
                for call in hub.calls), 0)

    def test_bound_adoption_allows_empty_remark_env(self):
        # 合法收养：本批次建环境后中断（备注为空、未绑号），重跑同名收养续做
        hub = FakeHub()
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _s: None)
        runner.prepare(accounts, '1:新刚')   # 计划 XG-MX-0819-001
        hub.envs.append({
            'containerName': 'XG-MX-0819-001',
            'containerCode': '8001', 'serialNumber': 2001, 'remark': ''})
        runner.rows[0].state = 'failed'
        runner.run()
        self.assertEqual(runner.rows[0].state, 'done')
        self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 0)
        self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 1)

    def test_rate_limit_after_create_retries_by_adopting_half_finished_env(self):
        class RateLimitedAfterCreateHub(FakeHub):
            def __init__(self):
                super().__init__()
                self.fail_next_lookup = False

            def env_create(self, body):
                result = super().env_create(body)
                self.fail_next_lookup = True
                return result

            def env_lookup(self, container_code=None, container_name=None,
                           tag_name=None):
                if self.fail_next_lookup:
                    self.fail_next_lookup = False
                    raise RuntimeError(
                        '/env/list 返回 code=E010205: 请求太频繁，请稍后再试')
                return super().env_lookup(
                    container_code=container_code,
                    container_name=container_name,
                    tag_name=tag_name)

        hub = RateLimitedAfterCreateHub()
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1)
        runner.prepare(accounts, '1:新刚')

        runner.run()
        self.assertEqual(runner.rows[0].state, 'failed')
        self.assertEqual(runner.rows[0].error_step, 'env_created')
        self.assertEqual(len(hub.envs), 1)

        runner.retry_one(runner.rows[0].account.account_id)
        self.assertEqual(runner.rows[0].state, 'done')
        self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 1)

    def test_uncertain_create_result_is_adopted_without_duplicate_create(self):
        class UncertainCreateHub(FakeHub):
            def env_create(self, body):
                super().env_create(body)
                raise RuntimeError('synthetic response lost after create')

        hub = UncertainCreateHub()
        accounts = parse_vendor_workbook(
            BytesIO(workbook_bytes(demo_rows()[:1])))
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY,
            purchase_date='20260819', sleep_fn=lambda _seconds: None,
            max_workers=1)
        runner.prepare(accounts, '1:新刚')

        runner.run()

        self.assertEqual(runner.rows[0].state, 'done')
        self.assertEqual(len(hub.envs), 1)
        self.assertEqual(sum(call[0] == 'create' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'cookie' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'account' for call in hub.calls), 1)
        self.assertEqual(sum(call[0] == 'remark' for call in hub.calls), 1)

    def test_backup_result_tsv_contains_no_credentials(self):
        hub = FakeHub()
        runner = BackupEnvOrchestrator(
            hub, purchase_tag=TEST_TAG, proxy_link=TEST_PROXY, site='US',
            sleep_fn=lambda _seconds: None)
        runner.prepare('志恒', 1, '测试', '20260819')
        rows = runner.run()
        text = backup_result_tsv_bytes(rows, 'US').decode('utf-8-sig')
        self.assertIn('环境名', text)
        self.assertIn('ZH-US-测试-01', text)
        self.assertIn('完成', text)
        self.assertNotIn(TEST_PROXY, text)
        self.assertNotIn('proxy.example.test', text)
        self.assertNotIn('secret', text.lower())


if __name__ == '__main__':
    unittest.main()
