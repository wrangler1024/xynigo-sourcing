from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from test_auth_flow import build_test_app
from test_executor_channel import login
from xynigo_auth.models import (
    LogisticsQueryResult,
    LogisticsQueryRun,
    Permission,
    Role,
    RolePermission,
    Tenant,
    User,
    UserRole,
)


def _run(
    *,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    created_at: datetime,
    serials: list[str],
    site: str = "MX",
    status: str = "completed",
    parent: LogisticsQueryRun | None = None,
) -> LogisticsQueryRun:
    run_id = uuid.uuid4()
    root_id = (parent.root_run_id or parent.id) if parent is not None else run_id
    return LogisticsQueryRun(
        id=run_id,
        tenant_id=tenant_id,
        actor_user_id=user_id,
        source_run_key=f"history-{run_id}",
        payload_hash="a" * 64,
        result_payload_hash="b" * 64,
        executor_id=None,
        executor_task_id=None,
        parent_run_id=parent.id if parent is not None else None,
        root_run_id=root_id,
        query_mode="single_retry" if parent is not None else "initial",
        site=site,
        status=status,
        phase="completed" if status in {"completed", "partial_failure", "failed"} else status,
        attempt=0,
        progress_completed=len(serials),
        progress_total=len(serials),
        stop_requested=False,
        total_count=len(serials),
        success_count=len(serials) if status == "completed" else 0,
        failed_count=0 if status == "completed" else len(serials),
        request_summary={
            "environmentSerials": serials,
            "parentRunId": str(parent.id) if parent is not None else None,
        },
        started_at=created_at,
        completed_at=created_at + timedelta(minutes=1),
        last_heartbeat_at=created_at + timedelta(minutes=1),
        source="cloud_web",
        client_version="0.13.20",
        created_at=created_at,
        updated_at=created_at + timedelta(minutes=1),
    )


def _result(
    run: LogisticsQueryRun,
    serial: str,
    *,
    status: str,
    order_no: str = "",
) -> LogisticsQueryResult:
    return LogisticsQueryResult(
        id=uuid.uuid4(),
        run_id=run.id,
        tenant_id=run.tenant_id,
        environment_serial=serial,
        environment_name=f"ENV-{serial}",
        status=status,
        current_step="done",
        completed_steps=["query_completed"],
        platform_order_no=order_no or None,
        tracking_numbers=[],
        package_numbers=[],
        cancelled=False,
        risk_order=False,
        feishu_sync_status="pending",
        created_at=run.created_at,
        updated_at=run.updated_at,
    )


def test_logistics_history_admin_can_filter_users_and_merges_retry_in_input_order(
    tmp_path,
) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    now = datetime.now(UTC)

    with TestClient(app) as client:
        login(client)
        with database.session_factory() as session:
            tenant = session.scalar(
                select(Tenant).where(Tenant.feishu_tenant_key == "tenant_allowed")
            )
            assert tenant is not None
            user = session.scalar(
                select(User).where(
                    User.tenant_id == tenant.id,
                    User.feishu_open_id == "ou_admin",
                )
            )
            assert user is not None
            other_user = User(
                tenant_id=tenant.id,
                feishu_open_id="ou_history_other",
                display_name="History other",
                status="active",
            )
            session.add(other_user)
            session.flush()

            root = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(hours=2),
                serials=["20", "10"],
                status="partial_failure",
            )
            session.add(root)
            session.flush()
            session.add_all(
                [
                    _result(root, "20", status="ok", order_no="ORDER-20"),
                    _result(root, "10", status="fail"),
                ]
            )
            child = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(hours=1),
                serials=["10"],
                parent=root,
            )
            session.add(child)
            session.flush()
            session.add(_result(child, "10", status="ok", order_no="ORDER-10"))
            newest_retry = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(minutes=30),
                serials=["20"],
                parent=root,
            )
            session.add(newest_retry)
            session.flush()
            session.add(
                _result(newest_retry, "20", status="ok", order_no="ORDER-20-NEW")
            )

            other_root = _run(
                tenant_id=tenant.id,
                user_id=other_user.id,
                created_at=now,
                serials=["99"],
            )
            session.add(other_root)
            session.flush()
            session.add(_result(other_root, "99", status="ok", order_no="SECRET"))
            session.commit()
            root_id = str(root.id)
            newest_retry_id = str(newest_retry.id)
            other_root_id = str(other_root.id)
            user_id = str(user.id)
            tenant_id = tenant.id
            user_name = user.display_name
            other_user_name = other_user.display_name
            other_user_id = str(other_user.id)

        history = client.get("/v1/operation-runs/logistics-query/history")
        assert history.status_code == 200, history.text
        history_data = history.json()["data"]
        items = history_data["items"]
        assert [item["rootRunId"] for item in items] == [other_root_id, root_id]
        own_item = next(item for item in items if item["rootRunId"] == root_id)
        assert own_item["latestRunId"] == newest_retry_id
        assert own_item["retryCount"] == 2
        assert own_item["originalEnvironmentSerials"] == ["20", "10"]
        assert own_item["totalCount"] == 2
        assert own_item["successCount"] == 2
        assert own_item["durationSec"] == 180
        assert own_item["actorDisplayName"] == user_name
        assert {actor["displayName"] for actor in history_data["actors"]} == {
            user_name,
            other_user_name,
        }

        own_history = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"userId": user_id},
        )
        assert own_history.status_code == 200, own_history.text
        assert [
            item["rootRunId"]
            for item in own_history.json()["data"]["items"]
        ] == [root_id]

        latest_status = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"status": "completed"},
        )
        assert latest_status.status_code == 200, latest_status.text
        assert set(
            item["rootRunId"] for item in latest_status.json()["data"]["items"]
        ) == {root_id, other_root_id}
        stale_root_status = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"status": "partial_failure"},
        )
        assert stale_root_status.status_code == 200, stale_root_status.text
        assert stale_root_status.json()["data"]["items"] == []

        detail = client.get(
            f"/v1/operation-runs/logistics-query/history/{root_id}"
        )
        assert detail.status_code == 200, detail.text
        payload = detail.json()["data"]
        assert payload["rootRunId"] == root_id
        assert payload["latestRunId"] == newest_retry_id
        assert payload["durationSec"] == 180
        assert payload["attemptDurationSec"] == 60
        assert [row["environmentSerial"] for row in payload["rows"]] == ["20", "10"]
        assert [row["platformOrderNo"] for row in payload["rows"]] == [
            "ORDER-20-NEW",
            "ORDER-10",
        ]

        other_detail = client.get(
            f"/v1/operation-runs/logistics-query/history/{other_root_id}"
        )
        assert other_detail.status_code == 200
        assert other_detail.json()["data"]["actorDisplayName"] == "History other"

        quick_export = client.get(
            f"/v1/operation-runs/logistics-query/{root_id}/export",
            params={"includeScreenshots": "false"},
        )
        assert quick_export.status_code == 200, quick_export.text
        assert quick_export.headers["x-xynigo-screenshot-included"] == "0"
        assert quick_export.headers["x-xynigo-screenshot-missing"] == "0"
        workbook = load_workbook(BytesIO(quick_export.content), read_only=True)
        assert [
            workbook.active.cell(row=index, column=1).value
            for index in (2, 3)
        ] == ["20", "10"]
        assert [
            workbook.active.cell(row=index, column=3).value
            for index in (2, 3)
        ] == ["ORDER-20-NEW", "ORDER-10"]
        workbook.close()

        other_export = client.get(
            f"/v1/operation-runs/logistics-query/{other_root_id}/export",
            params={"includeScreenshots": "false"},
        )
        assert other_export.status_code == 200
        missing_screenshot = client.get(
            f"/v1/operation-runs/logistics-query/{other_root_id}/screenshots/99"
        )
        assert missing_screenshot.status_code == 404

        with database.session_factory() as session:
            logistics_reader = Role(
                tenant_id=tenant_id,
                code="logistics_reader",
                name="物流查询成员",
                is_system=False,
            )
            session.add(logistics_reader)
            session.flush()
            read_permission = session.scalar(
                select(Permission).where(
                    Permission.code == "fulfillment.order.read"
                )
            )
            assert read_permission is not None
            for assignment in session.scalars(
                select(UserRole).where(UserRole.user_id == uuid.UUID(user_id))
            ):
                session.delete(assignment)
            session.add(UserRole(
                user_id=uuid.UUID(user_id), role_id=logistics_reader.id
            ))
            session.add(RolePermission(
                role_id=logistics_reader.id,
                permission_id=read_permission.id,
            ))
            session.commit()

        member_history = client.get("/v1/operation-runs/logistics-query/history")
        assert member_history.status_code == 200, member_history.text
        assert [
            item["rootRunId"]
            for item in member_history.json()["data"]["items"]
        ] == [root_id]
        assert member_history.json()["data"]["actors"] == []
        forbidden_filter = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"userId": other_user_id},
        )
        assert forbidden_filter.status_code == 403
        hidden_detail = client.get(
            f"/v1/operation-runs/logistics-query/history/{other_root_id}"
        )
        assert hidden_detail.status_code == 404


def test_active_retry_reports_logical_batch_progress(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    now = datetime.now(UTC)

    with TestClient(app) as client:
        login(client)
        with database.session_factory() as session:
            tenant = session.scalar(
                select(Tenant).where(Tenant.feishu_tenant_key == "tenant_allowed")
            )
            user = session.scalar(
                select(User).where(
                    User.tenant_id == tenant.id,
                    User.feishu_open_id == "ou_admin",
                )
            )
            root = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(minutes=10),
                serials=["10", "20", "30"],
                status="partial_failure",
            )
            session.add(root)
            session.flush()
            session.add_all(
                [
                    _result(root, "10", status="ok", order_no="ORDER-10"),
                    _result(root, "20", status="fail"),
                    _result(root, "30", status="ok", order_no="ORDER-30"),
                ]
            )
            retry = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now,
                serials=["20"],
                status="running",
                parent=root,
            )
            retry.progress_completed = 0
            retry.progress_total = 1
            retry.total_count = 1
            retry.completed_at = None
            session.add(retry)
            session.commit()
            retry_id = str(retry.id)

        response = client.get(
            f"/v1/operation-runs/logistics-query/{retry_id}"
        )
        assert response.status_code == 200, response.text
        payload = response.json()["data"]
        assert payload["progressCompleted"] == 2
        assert payload["progressTotal"] == 3
        assert payload["retryProgressCompleted"] == 0
        assert payload["retryProgressTotal"] == 1
        assert [row["environmentSerial"] for row in payload["rows"]] == [
            "10",
            "20",
            "30",
        ]
        assert [row["status"] for row in payload["rows"]] == [
            "ok",
            "pending",
            "ok",
        ]


def test_logistics_history_paginates_newest_first_and_filters_latest_status(
    tmp_path,
) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    now = datetime.now(UTC)

    with TestClient(app) as client:
        login(client)
        with database.session_factory() as session:
            tenant = session.scalar(
                select(Tenant).where(Tenant.feishu_tenant_key == "tenant_allowed")
            )
            user = session.scalar(
                select(User).where(
                    User.tenant_id == tenant.id,
                    User.feishu_open_id == "ou_admin",
                )
            )
            assert tenant is not None and user is not None
            oldest = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(days=2),
                serials=["1"],
                site="MX",
            )
            middle = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now - timedelta(days=1),
                serials=["2"],
                site="US",
                status="failed",
            )
            newest = _run(
                tenant_id=tenant.id,
                user_id=user.id,
                created_at=now,
                serials=["3"],
                site="MX",
            )
            session.add_all([oldest, middle, newest])
            session.flush()
            session.add_all(
                [
                    _result(oldest, "1", status="ok"),
                    _result(middle, "2", status="fail"),
                    _result(newest, "3", status="ok"),
                ]
            )
            session.commit()
            middle_id = str(middle.id)
            newest_id = str(newest.id)

        first = client.get(
            "/v1/operation-runs/logistics-query/history", params={"limit": 1}
        )
        assert first.status_code == 200, first.text
        first_page = first.json()["data"]
        assert [item["rootRunId"] for item in first_page["items"]] == [
            newest_id
        ]
        assert first_page["hasMore"] is True
        assert first_page["nextCursor"] == newest_id

        second = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"limit": 1, "cursor": first_page["nextCursor"]},
        )
        assert second.status_code == 200, second.text
        second_page = second.json()["data"]
        assert [item["rootRunId"] for item in second_page["items"]] == [
            middle_id
        ]

        filtered = client.get(
            "/v1/operation-runs/logistics-query/history",
            params={"site": "US", "status": "failed"},
        )
        assert filtered.status_code == 200, filtered.text
        assert [item["rootRunId"] for item in filtered.json()["data"]["items"]] == [
            middle_id
        ]
