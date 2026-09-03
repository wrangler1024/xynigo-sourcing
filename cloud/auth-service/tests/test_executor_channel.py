from __future__ import annotations

from copy import deepcopy
from datetime import UTC, datetime, timedelta
import base64
import hashlib
from io import BytesIO
import json
import uuid

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlalchemy import func, select

from test_auth_flow import build_test_app, start_login
from xynigo_auth.models import (
    EnvironmentAccountPlan,
    EnvironmentAccountRunGuard,
    EnvironmentCreationRun,
    EnvironmentNameSequence,
    ExecutorPairingCode,
    ExecutorTask,
    HubEnvironmentInventory,
    LogisticsQueryRun,
    LocalExecutor,
    Tenant,
    User,
)
from xynigo_auth.security import hash_token


CSRF = {"X-Xynigo-Web-CSRF": "same-origin"}
REVISION_A = "a" * 64
REVISION_B = "b" * 64


def environment_workbook(count: int = 2, *, site: str = "MX") -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["邮箱账号", "密码", "接码Key链接", "Cookie"])
    domain = ".shein.com.mx" if site == "MX" else ".us.shein.com"
    for index in range(1, count + 1):
        email = f"buyer{index}@example.test"
        sheet.append([
            email,
            "synthetic-password",
            f"https://vendor.example/api?orderNo={index:08x}",
            json.dumps([{"name": "session", "value": "test", "domain": domain}]),
        ])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return base64.b64encode(output.getvalue()).decode("ascii")


def login(client: TestClient) -> None:
    state, _challenge = start_login(client)
    response = client.get(
        "/v1/auth/feishu/callback",
        params={"code": "authorization-code", "state": state},
        follow_redirects=False,
    )
    assert response.status_code == 303


def create_pairing_code_payload(client: TestClient) -> dict[str, object]:
    response = client.post(
        "/v1/executors/pairing-codes",
        json={"displayNameHint": "采购电脑"},
        headers=CSRF,
    )
    assert response.status_code == 201, response.text
    payload = response.json()
    assert payload["expiresIn"] == 300
    assert payload["pairingRequestId"]
    return payload


def create_pairing_code(client: TestClient) -> str:
    return str(create_pairing_code_payload(client)["pairingCode"])


def pair(
    device_client: TestClient,
    pairing_code: str,
    *,
    capabilities: list[str] | None = None,
) -> dict[str, object]:
    response = device_client.post(
        "/v1/executor-channel/pair",
        json={
            "pairingCode": pairing_code,
            "displayName": "采购电脑 A",
            "platform": "macos",
            "architecture": "arm64",
            "clientVersion": "0.12.5",
            "protocolVersion": 1,
            "capabilities": capabilities
            or ["config.read.v1", "config.write.v1"],
        },
        headers={"X-Xynigo-Source": "local_executor_device"},
    )
    assert response.status_code == 201, response.text
    return response.json()


def device_headers(credential: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {credential}",
        "X-Xynigo-Source": "local_executor_device",
        "X-Xynigo-Client-Version": "0.12.5",
    }


def heartbeat(
    device_client: TestClient,
    credential: str,
    *,
    revision: str | None = None,
    capabilities: list[str] | None = None,
    config_summary: dict[str, object] | None = None,
    hub_status: str = "ready",
    client_version: str = "0.13.18",
) -> dict[str, object]:
    body = {
        "waitSeconds": 0,
        "configRevision": revision,
        "hubStatus": hub_status,
        "clientVersion": client_version,
        "protocolVersion": 1,
        "capabilities": capabilities
        or ["config.read.v1", "config.write.v1"],
    }
    if config_summary is not None:
        body["configSummary"] = config_summary
    response = device_client.post(
        "/v1/executor-channel/poll",
        json=body,
        headers=device_headers(credential),
    )
    assert response.status_code == 200, response.text
    return response.json()


def safe_config_summary() -> dict[str, object]:
    return {
        "schemaVersion": 2,
        "configRevision": "c" * 64,
        "capturedAt": datetime.now(UTC).isoformat(),
        "runtimeConfig": {
            "hubPort": 6873,
            "concurrency": 2,
            "envCreateWorkers": 5,
            "verifySampleCount": 1,
            "safeParallelTasks": True,
        },
        "configured": {
            "hubApiKey": True,
            "larkAppCredentials": True,
            "larkLegacyTarget": False,
            "purchaseAssistantDataSources": True,
            "teamDefaultDataSource": False,
        },
        "dataSources": {
            "dataSourceCount": 2,
            "buyerProfileCount": 1,
            "environmentBindingCount": 3,
            "pendingOwnerConfirmationCount": 1,
            "mappingConflictCount": 0,
        },
        "compliance": {"status": "ready", "issueCodes": []},
    }


def test_environment_plan_is_parsed_in_cloud_and_site_errors_are_structured(
    tmp_path,
) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/v1/environment-plans/parse",
            json={
                "idempotencyKey": "cloud-environment-plan-0001",
                "filename": "buyers.xlsx",
                "contentBase64": environment_workbook(1, site="MX"),
                "site": "MX",
                "environmentGroup": "MX采购",
            },
            headers=CSRF,
        )
        assert created.status_code == 201, created.text
        result = created.json()
        assert result["runtime"] == "cloud"
        assert result["reused"] is False
        assert result["count"] == 1
        assert result["preview"][0]["emailMasked"] != "buyer1@example.test"

        latest = client.get(
            "/v1/environment-plans/latest",
            params={"site": "MX", "environmentGroup": "MX采购"},
        )
        assert latest.status_code == 200
        assert latest.json()["plan"]["cloudPlanId"] == result["cloudPlanId"]

        mismatch = client.post(
            "/v1/environment-plans/parse",
            json={
                "idempotencyKey": "cloud-environment-plan-0002",
                "filename": "buyers-us.xlsx",
                "contentBase64": environment_workbook(1, site="US"),
                "site": "MX",
                "environmentGroup": "MX采购",
            },
            headers=CSRF,
        )
        assert mismatch.status_code == 422
        assert mismatch.json()["detail"]["code"] == "environment_plan_site_mismatch"
        assert "第2行" in mismatch.json()["detail"]["message"]

    with database.session_factory() as session:
        record = session.scalar(select(EnvironmentAccountPlan))
        assert record is not None and record.encrypted_payload
        assert b"synthetic-password" not in record.encrypted_payload
        assert session.scalar(select(ExecutorTask)) is None


def test_pairing_is_single_use_and_credentials_are_only_stored_hashed(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        pairing = create_pairing_code_payload(web_client)
        code = str(pairing["pairingCode"])
        pairing_id = str(pairing["pairingRequestId"])
        pending = web_client.get(f"/v1/executors/pairing-codes/{pairing_id}")
        assert pending.status_code == 200
        assert pending.json()["status"] == "pending"
        assert pending.json()["executorId"] is None
        paired = pair(device_client, code)
        credential = str(paired["deviceCredential"])
        assert len(credential) >= 32

        reused = device_client.post(
            "/v1/executor-channel/pair",
            json={
                "pairingCode": code,
                "displayName": "另一台电脑",
                "platform": "windows",
                "architecture": "x86_64",
                "clientVersion": "0.12.5",
                "capabilities": ["config.read.v1"],
            },
        )
        assert reused.status_code == 409
        assert reused.json()["detail"]["code"] == "pairing_code_consumed"

        with database.session_factory() as session:
            executor = session.scalar(select(LocalExecutor))
            assert executor is not None
            assert executor.credential_digest == hash_token(credential)
            assert credential not in executor.credential_digest
            stored_code = session.scalar(select(ExecutorPairingCode))
            assert stored_code is not None
            assert stored_code.code_digest == hash_token(code.replace("-", ""))

        listed = web_client.get("/v1/executors")
        assert listed.status_code == 200
        item = listed.json()["items"][0]
        assert item["displayName"] == "采购电脑 A"
        assert item["connectivity"] == "offline"
        assert "deviceCredential" not in str(listed.json())

        consumed = web_client.get(f"/v1/executors/pairing-codes/{pairing_id}")
        assert consumed.status_code == 200
        assert consumed.json()["status"] == "consumed"
        assert consumed.json()["executorId"] == paired["executorId"]

        heartbeat(device_client, credential)
        online = web_client.get("/v1/executors").json()["items"][0]
        assert online["connectivity"] == "online"

        missing = web_client.get(f"/v1/executors/pairing-codes/{uuid.uuid4()}")
        assert missing.status_code == 404


def test_paired_executor_can_issue_its_owners_local_user_session(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(device_client, create_pairing_code(web_client))
        credential = str(paired["deviceCredential"])
        issued = device_client.post(
            "/v1/executor-channel/session",
            json={},
            headers=device_headers(credential),
        )
        assert issued.status_code == 200, issued.text
        session_token = str(issued.json()["sessionToken"])
        assert len(session_token) >= 32
        identity = device_client.get(
            "/v1/auth/me",
            headers={"Authorization": f"Bearer {session_token}"},
        )
        assert identity.status_code == 200, identity.text
        assert identity.json()["user"]["id"] == web_client.get(
            "/v1/auth/me"
        ).json()["user"]["id"]


def test_expired_pairing_code_and_device_cookie_are_rejected(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        pairing = create_pairing_code_payload(web_client)
        code = str(pairing["pairingCode"])
        pairing_id = str(pairing["pairingRequestId"])
        with database.session_factory() as session:
            record = session.scalar(select(ExecutorPairingCode))
            assert record is not None
            record.expires_at = datetime.now(UTC) - timedelta(seconds=1)
            session.commit()

        expired_status = web_client.get(
            f"/v1/executors/pairing-codes/{pairing_id}"
        )
        assert expired_status.status_code == 200
        assert expired_status.json()["status"] == "expired"
        assert "pairingCode" not in expired_status.text
        assert "credential" not in expired_status.text.lower()

        expired = device_client.post(
            "/v1/executor-channel/pair",
            json={
                "pairingCode": code,
                "displayName": "采购电脑",
                "platform": "macos",
                "architecture": "arm64",
                "clientVersion": "0.12.5",
                "capabilities": ["config.read.v1"],
            },
        )
        assert expired.status_code == 410
        assert expired.json()["detail"]["code"] == "pairing_code_expired"

        new_code = create_pairing_code(web_client)
        paired = pair(device_client, new_code)
        credential = str(paired["deviceCredential"])
        device_client.cookies.set("xynigo_session", "browser-cookie")
        rejected = device_client.post(
            "/v1/executor-channel/poll",
            json={
                "waitSeconds": 0,
                "hubStatus": "unknown",
                "clientVersion": "0.12.5",
                "capabilities": ["config.read.v1"],
            },
            headers=device_headers(credential),
        )
        assert rejected.status_code == 401
        assert rejected.json()["detail"]["code"] == "executor_cookie_not_allowed"


def test_config_read_write_lease_and_idempotent_finish(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(device_client, create_pairing_code(web_client))
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        assert heartbeat(device_client, credential, revision=REVISION_A)["task"] is None

        read_created = web_client.post(
            f"/v1/executors/{executor_id}/config/read",
            json={},
            headers=CSRF,
        )
        assert read_created.status_code == 202, read_created.text
        read_task_id = read_created.json()["task"]["id"]

        lease = heartbeat(device_client, credential, revision=REVISION_A)["task"]
        assert lease["id"] == read_task_id
        assert lease["type"] == "config.read.v1"
        lease_token = lease["leaseToken"]
        assert "leaseToken" not in str(read_created.json())

        started = device_client.post(
            f"/v1/executor-channel/tasks/{read_task_id}/start",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        )
        assert started.status_code == 200
        assert started.json()["task"]["status"] == "running"

        with database.session_factory() as session:
            running_task = session.get(ExecutorTask, uuid.UUID(read_task_id))
            assert running_task is not None
            running_task.lease_until = datetime.now(UTC) + timedelta(seconds=2)
            previous_lease_until = running_task.lease_until
            session.commit()
        progressed = device_client.post(
            f"/v1/executor-channel/tasks/{read_task_id}/progress",
            json={
                "leaseToken": lease_token,
                "phase": "config.executing",
                "current": 0,
                "total": 1,
            },
            headers=device_headers(credential),
        )
        assert progressed.status_code == 200, progressed.text
        with database.session_factory() as session:
            running_task = session.get(ExecutorTask, uuid.UUID(read_task_id))
            assert running_task is not None
            refreshed_lease_until = running_task.lease_until
            assert refreshed_lease_until is not None
            if refreshed_lease_until.tzinfo is None:
                refreshed_lease_until = refreshed_lease_until.replace(tzinfo=UTC)
            assert refreshed_lease_until > previous_lease_until + timedelta(
                seconds=30
            )

        finish_body = {
            "leaseToken": lease_token,
            "outcome": "succeeded",
            "resultCode": "config_read_succeeded",
            "resultSummary": {
                "configRevision": REVISION_A,
                "config": {
                    "hubPort": 6873,
                    "serverPort": 8765,
                    "concurrency": 2,
                    "importBuyerPlan": "1:新刚",
                    "verifySampleCount": 3,
                    "hiddenQueryColumns": ["envName", "ip"],
                    "purchaseSite": "MX",
                    "purchaseTags": {"MX": "MX采购", "US": "US采购"},
                    "envCreateWorkers": 5,
                    "safeParallelTasks": True,
                },
            },
        }
        finished = device_client.post(
            f"/v1/executor-channel/tasks/{read_task_id}/finish",
            json=finish_body,
            headers=device_headers(credential),
        )
        assert finished.status_code == 200, finished.text
        assert finished.json()["task"]["status"] == "succeeded"
        duplicate = device_client.post(
            f"/v1/executor-channel/tasks/{read_task_id}/finish",
            json=finish_body,
            headers=device_headers(credential),
        )
        assert duplicate.status_code == 200

        write_body = {
            "expectedRevision": REVISION_A,
            "idempotencyKey": "config-write-test-0001",
            "config": {
                "hubPort": 6873,
                "concurrency": 3,
                "verifySampleCount": 3,
                "envCreateWorkers": 5,
                "safeParallelTasks": True,
            },
        }
        write_created = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json=write_body,
            headers=CSRF,
        )
        assert write_created.status_code == 202, write_created.text
        duplicate_write = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json=write_body,
            headers=CSRF,
        )
        assert duplicate_write.status_code == 200 or duplicate_write.status_code == 202
        assert duplicate_write.json()["task"]["id"] == write_created.json()["task"]["id"]

        with database.session_factory() as session:
            tasks = list(session.scalars(select(ExecutorTask)))
            assert len(tasks) == 2
            assert all("proxyLink" not in str(task.payload_envelope) for task in tasks)


def test_cloud_operation_runs_dispatch_formal_tasks_and_restore_progress(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    capabilities = [
        "config.read.v1",
        "config.write.v1",
        "workspace.rpc.v1",
        "workspace.snapshot.v1",
        "environment.parse.v1",
        "environment.cloud-plan.v1",
        "environment.cloud-inventory.v1",
        "environment.create-bound.v1",
        "environment.create-backup.v1",
        "environment.retry-row.v1",
        "environment.retry-failed.v1",
        "logistics.query.v1",
    ]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        assert heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"] is None
        snapshot_refresh = web_client.post(
            f"/v1/executors/{executor_id}/workspace-snapshot",
            json={"idempotencyKey": "workspace-snapshot-0001"},
            headers=CSRF,
        )
        assert snapshot_refresh.status_code == 202, snapshot_refresh.text
        snapshot_task_id = snapshot_refresh.json()["task"]["id"]
        snapshot_lease = heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"]
        assert snapshot_lease["type"] == "workspace.snapshot.v1"
        snapshot_credential = snapshot_lease["leaseToken"]
        assert device_client.post(
            f"/v1/executor-channel/tasks/{snapshot_task_id}/start",
            json={"leaseToken": snapshot_credential},
            headers=device_headers(credential),
        ).status_code == 200
        snapshot_result = {
            "schemaVersion": 1,
            "snapshotRevision": "c" * 64,
            "capturedAt": "2026-09-01T18:00:00+08:00",
            "preferences": {
                "purchaseSite": "MX",
                "purchaseTags": {"MX": "MX采购", "US": "美国采购"},
                "importBuyerPlan": "2:新刚",
                "verifySampleCount": 1,
                "buyers": [{"name": "新刚", "code": "XG"}],
                "buyerDefaultSplit": ["新刚"],
                "backupMaxCount": 25,
            },
            "runtimeConfig": {
                "configRevision": REVISION_A,
                "hubPort": 6873,
                "concurrency": 2,
                "envCreateWorkers": 5,
                "verifySampleCount": 1,
                "safeParallelTasks": True,
            },
            "groups": ["MX采购", "美国采购"],
            "preflight": {
                "MX": {
                    "ready": True,
                    "hubConnected": True,
                    "groupFound": True,
                    "proxyConfigured": True,
                    "purchaseTag": "MX采购",
                    "configuredWorkers": 5,
                    "effectiveWorkers": 5,
                    "message": "预检通过",
                },
                "US": {
                    "ready": True,
                    "hubConnected": True,
                    "groupFound": True,
                    "proxyConfigured": True,
                    "purchaseTag": "美国采购",
                    "configuredWorkers": 5,
                    "effectiveWorkers": 5,
                    "message": "预检通过",
                },
            },
        }
        unsafe_snapshot = deepcopy(snapshot_result)
        unsafe_snapshot["password"] = "must-not-persist"
        assert device_client.post(
            f"/v1/executor-channel/tasks/{snapshot_task_id}/finish",
            json={
                "leaseToken": snapshot_credential,
                "outcome": "succeeded",
                "resultCode": "workspace_snapshot_completed",
                "resultSummary": unsafe_snapshot,
            },
            headers=device_headers(credential),
        ).status_code == 422
        snapshot_finished = device_client.post(
            f"/v1/executor-channel/tasks/{snapshot_task_id}/finish",
            json={
                "leaseToken": snapshot_credential,
                "outcome": "succeeded",
                "resultCode": "workspace_snapshot_completed",
                "resultSummary": snapshot_result,
            },
            headers=device_headers(credential),
        )
        assert snapshot_finished.status_code == 200, snapshot_finished.text
        snapshot_visible = web_client.get(
            f"/v1/executors/{executor_id}/workspace-snapshot"
        )
        assert snapshot_visible.status_code == 200
        assert snapshot_visible.json()["snapshot"]["groups"] == [
            "MX采购", "美国采购"
        ]
        assert snapshot_visible.json()["snapshot"]["runtimeConfig"][
            "configRevision"
        ] == REVISION_A
        assert snapshot_visible.json()["snapshotRevision"] == "c" * 64
        preference_write = web_client.put(
            "/v1/environment-preferences",
            json={
                "purchaseSite": "US",
                "purchaseTags": {"US": "美国采购"},
            },
            headers=CSRF,
        )
        assert preference_write.status_code == 200, preference_write.text
        assert heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"] is None
        updated_snapshot = web_client.get(
            f"/v1/executors/{executor_id}/workspace-snapshot"
        ).json()["snapshot"]
        assert updated_snapshot["preferences"]["purchaseSite"] == "US"
        assert updated_snapshot["groups"] == ["MX采购", "美国采购"]
        parse_created = web_client.post(
            "/v1/environment-plans/parse",
            json={
                "idempotencyKey": "environment-parse-0001",
                "filename": "synthetic-buyers.xlsx",
                "contentBase64": environment_workbook(2),
                "site": "MX",
                "environmentGroup": "MX采购测试",
            },
            headers=CSRF,
        )
        assert parse_created.status_code == 201, parse_created.text
        parse_result = parse_created.json()
        assert parse_result["runtime"] == "cloud"
        assert parse_result["count"] == 2
        assert parse_result["environmentGroup"] == "MX采购测试"
        cloud_plan_id = parse_result["cloudPlanId"]
        assert heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"] is None
        create_body = {
            "idempotencyKey": "environment-run-create-0001",
            "executorId": executor_id,
            "mode": "bound",
            "site": "MX",
            "purchaseDate": "20260901",
            "environmentGroup": "MX采购测试",
            "cloudPlanId": cloud_plan_id,
            "totalCount": 2,
            "verifySampleCount": 2,
            "assignments": [{"purchaserLabel": "合成采购员", "count": 2}],
        }
        created = web_client.post(
            "/v1/operation-runs/environment-creation",
            json=create_body,
            headers=CSRF,
        )
        assert created.status_code == 202, created.text
        snapshot = created.json()["data"]
        assert snapshot["status"] == "queued"
        assert snapshot["phase"] == "queued"
        assert snapshot["progressTotal"] == 2
        assert snapshot["terminal"] is False
        assert snapshot["executorTaskId"]

        overlapping_plan = web_client.post(
            "/v1/environment-plans/parse",
            json={
                "idempotencyKey": "environment-parse-overlap-0002",
                "filename": "synthetic-buyers.xlsx",
                "contentBase64": environment_workbook(2),
                "site": "MX",
                "environmentGroup": "MX采购测试",
            },
            headers=CSRF,
        )
        assert overlapping_plan.status_code == 201, overlapping_plan.text
        overlapping_body = {
            **create_body,
            "idempotencyKey": "environment-run-overlap-0002",
            "cloudPlanId": overlapping_plan.json()["cloudPlanId"],
        }
        overlapping = web_client.post(
            "/v1/operation-runs/environment-creation",
            json=overlapping_body,
            headers=CSRF,
        )
        assert overlapping.status_code == 409, overlapping.text
        assert (
            overlapping.json()["detail"]["code"]
            == "environment_cleanup_in_progress"
        )

        repeated = web_client.post(
            "/v1/operation-runs/environment-creation",
            json=create_body,
            headers=CSRF,
        )
        assert repeated.status_code == 202
        assert repeated.json()["data"]["unchanged"] is True
        assert repeated.json()["data"]["runId"] == snapshot["runId"]

        changed = deepcopy(create_body)
        changed["environmentGroup"] = "MX采购备用"
        conflict = web_client.post(
            "/v1/operation-runs/environment-creation",
            json=changed,
            headers=CSRF,
        )
        assert conflict.status_code == 409

        with database.session_factory() as session:
            run = session.scalar(select(EnvironmentCreationRun))
            task = session.scalar(
                select(ExecutorTask).where(
                    ExecutorTask.task_type == "environment.create-bound.v1"
                )
            )
            assert run is not None and task is not None
            assert run.executor_task_id == task.id
            assert task.task_type == "environment.create-bound.v1"
            guards = list(session.scalars(select(EnvironmentAccountRunGuard)))
            assert len(guards) == 2
            assert {guard.state for guard in guards} == {"active"}
            inventory = list(session.scalars(select(HubEnvironmentInventory)))
            assert len(inventory) == 2
            assert {row.state for row in inventory} == {"reserved"}
            assert len({row.environment_name for row in inventory}) == 2
            sequence = session.scalar(select(EnvironmentNameSequence))
            assert sequence is not None and sequence.last_value == 2
            serialized = json.dumps(task.payload_envelope)
            assert cloud_plan_id not in serialized
            assert "MX采购测试" not in serialized

        leased = heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"]
        assert leased["type"] == "environment.create-bound.v1"
        assert leased["payload"]["runId"] == snapshot["runId"]
        assert leased["payload"]["cloudPlanId"] == cloud_plan_id
        assert "planRef" not in leased["payload"]
        assert len(leased["payload"]["planAccounts"]) == 2
        assert leased["payload"]["cleanupBlockedAccountRefs"] == []
        assert len(leased["payload"]["plannedEnvironmentNames"]) == 2
        lease_token = leased["leaseToken"]
        leased_snapshot = web_client.get(
            f"/v1/operation-runs/environment-creation/{snapshot['runId']}"
        ).json()["data"]
        assert leased_snapshot["status"] == "leased"
        assert leased_snapshot["attempt"] == 1

        task_id = snapshot["executorTaskId"]
        started = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/start",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        )
        assert started.status_code == 200
        unsafe_progress = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/progress",
            json={
                "leaseToken": lease_token,
                "phase": "environment.creating",
                "current": 0,
                "total": 2,
                "snapshot": {
                    "rows": [
                        {
                            "accountRef": "sha256-progress-account-unsafe",
                            "accountLabel": "plain@example.test",
                            "purchaserLabel": "合成采购员",
                            "environmentName": "SYN-MX-UNSAFE",
                            "status": "running",
                            "currentStep": "env_created",
                            "completedSteps": [],
                            "password": "must-not-persist",
                        }
                    ]
                },
            },
            headers=device_headers(credential),
        )
        assert unsafe_progress.status_code == 422
        progress = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/progress",
            json={
                "leaseToken": lease_token,
                "phase": "environment.creating",
                "current": 1,
                "total": 2,
                "snapshot": {
                    "rows": [
                        {
                            "accountRef": "sha256-progress-account-0001",
                            "accountLabel": "pr***01@example.test",
                            "purchaserLabel": "合成采购员",
                            "environmentName": "SYN-MX-0901-001",
                            "status": "running",
                            "currentStep": "env_created",
                            "completedSteps": [],
                            "createdInRun": True,
                            "cleanupStatus": "pending",
                            "ipVerified": False,
                            "ipErrorCode": "hub_ip_missing",
                            "ipErrorSummary": "HubStudio 未返回出口 IP",
                        },
                        {
                            "accountRef": "sha256-progress-account-0002",
                            "accountLabel": "pr***02@example.test",
                            "purchaserLabel": "合成采购员",
                            "environmentName": "SYN-MX-0901-002",
                            "status": "queued",
                            "completedSteps": [],
                        },
                    ]
                },
            },
            headers=device_headers(credential),
        )
        assert progress.status_code == 200, progress.text
        running = web_client.get(
            f"/v1/operation-runs/environment-creation/{snapshot['runId']}"
        ).json()["data"]
        assert running["status"] == "running"
        assert running["phase"] == "environment.creating"
        assert running["progressCompleted"] == 1
        assert len(running["rows"]) == 2
        running_rows = {row["accountRef"]: row for row in running["rows"]}
        assert running_rows["sha256-progress-account-0001"]["status"] == "running"
        assert running_rows["sha256-progress-account-0001"]["createdInRun"] is True
        assert running_rows["sha256-progress-account-0001"]["cleanupStatus"] == "pending"
        assert running_rows["sha256-progress-account-0001"]["ipErrorCode"] == "hub_ip_missing"
        assert running_rows["sha256-progress-account-0002"]["status"] == "queued"

        unsafe_finish = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/finish",
            json={
                "leaseToken": lease_token,
                "outcome": "succeeded",
                "resultCode": "environment_run_completed",
                "resultSummary": {
                    "runStatus": "completed",
                    "phase": "environment.completed",
                    "password": "must-not-persist",
                },
            },
            headers=device_headers(credential),
        )
        assert unsafe_finish.status_code == 422

        finished = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/finish",
            json={
                "leaseToken": lease_token,
                "outcome": "succeeded",
                "resultCode": "environment_run_completed",
                "resultSummary": {
                    "runStatus": "partial_failure",
                    "phase": "completed",
                    "progressCompleted": 2,
                    "progressTotal": 2,
                    "successCount": 1,
                    "failedCount": 1,
                    "cleanupTotal": 1,
                    "cleanupDone": 0,
                    "cleanupFailed": 0,
                    "ipOkCount": 1,
                    "ipTotalCount": 1,
                },
            },
            headers=device_headers(credential),
        )
        assert finished.status_code == 200, finished.text
        with database.session_factory() as session:
            terminal_task = session.get(ExecutorTask, uuid.UUID(task_id))
            stored_plan = session.get(
                EnvironmentAccountPlan, uuid.UUID(cloud_plan_id)
            )
            assert terminal_task is not None
            assert "encryptedPayload" not in terminal_task.payload_envelope
            assert terminal_task.payload_envelope.get("purgedAt")
            assert stored_plan is not None
            assert stored_plan.status == "submitted"
            assert stored_plan.encrypted_payload is None
            assert session.scalar(
                select(func.count()).select_from(EnvironmentAccountRunGuard)
            ) == 0
        latest = web_client.get(
            "/v1/operation-runs/environment-creation/latest"
        )
        assert latest.status_code == 200
        final = latest.json()["data"]
        assert final["runId"] == snapshot["runId"]
        assert final["status"] == "partial_failure"
        assert final["terminal"] is True
        assert final["progressCompleted"] == 2
        assert final["successCount"] == 1
        assert final["failedCount"] == 1
        assert final["ipOkCount"] == 1

        ingested = web_client.put(
            "/v1/operations/environment-creation-runs",
            json={
                "source": "local_executor",
                "runKey": create_body["idempotencyKey"],
                "site": "MX",
                "purchaseDate": "20260901",
                "environmentGroup": "MX采购测试",
                "startedAt": "2026-09-01T08:00:00+08:00",
                "completedAt": "2026-09-01T08:02:00+08:00",
                "results": [
                    {
                        "accountRef": "sha256-progress-account-0001",
                        "accountLabel": "pr***01@example.test",
                        "purchaserLabel": "合成采购员",
                        "environmentName": "SYN-MX-0901-001",
                        "environmentRef": "hub-synthetic-mx-0001",
                        "environmentSerial": "9001",
                        "status": "success",
                        "bindingAt": "2026-09-01T08:01:00+08:00",
                    },
                    {
                        "accountRef": "sha256-progress-account-0002",
                        "accountLabel": "pr***02@example.test",
                        "purchaserLabel": "合成采购员",
                        "environmentName": "SYN-MX-0901-002",
                        "status": "failed",
                        "errorStep": "account_bound",
                        "errorSummary": "合成绑定失败",
                    },
                ],
                "ipChecks": [
                    {
                        "environmentName": "SYN-MX-0901-001",
                        "ipAddress": "192.0.2.20",
                        "country": "Mexico",
                        "city": "Example City",
                        "isp": "Synthetic ISP",
                        "ok": True,
                    }
                ],
            },
            headers=CSRF,
        )
        assert ingested.status_code == 200, ingested.text
        assert ingested.json()["data"]["runId"] == snapshot["runId"]
        assert ingested.json()["data"]["unchanged"] is False
        restored = web_client.get(
            f"/v1/operation-runs/environment-creation/{snapshot['runId']}"
        ).json()["data"]
        assert len(restored["rows"]) == 2
        restored_rows = {row["accountRef"]: row for row in restored["rows"]}
        assert restored_rows["sha256-progress-account-0001"]["status"] == "success"
        assert restored_rows["sha256-progress-account-0001"]["ipVerified"] is True

        retry_created = web_client.post(
            f"/v1/operation-runs/environment-creation/{snapshot['runId']}/retry",
            json={
                "idempotencyKey": "environment-retry-row-0001",
                "retryMode": "single",
                "accountRefs": ["sha256-progress-account-0002"],
            },
            headers=CSRF,
        )
        assert retry_created.status_code == 202, retry_created.text
        retry_snapshot = retry_created.json()["data"]
        assert retry_snapshot["mode"] == "retry_row"
        assert retry_snapshot["parentRunId"] == snapshot["runId"]
        assert retry_snapshot["progressTotal"] == 1
        retry_lease = heartbeat(
            device_client, credential, capabilities=capabilities
        )["task"]
        assert retry_lease["type"] == "environment.retry-row.v1"
        assert retry_lease["payload"]["accountRefs"] == [
            "sha256-progress-account-0002"
        ]
        retry_task_id = retry_snapshot["executorTaskId"]
        retry_credential = retry_lease["leaseToken"]
        assert device_client.post(
            f"/v1/executor-channel/tasks/{retry_task_id}/start",
            json={"leaseToken": retry_credential},
            headers=device_headers(credential),
        ).status_code == 200
        retry_progress = device_client.post(
            f"/v1/executor-channel/tasks/{retry_task_id}/progress",
            json={
                "leaseToken": retry_credential,
                "phase": "environment.creating",
                "current": 1,
                "total": 1,
                "snapshot": {"rows": [{
                    "accountRef": "sha256-progress-account-0002",
                    "accountLabel": "pr***02@example.test",
                    "purchaserLabel": "合成采购员",
                    "environmentName": "SYN-MX-0901-002",
                    "environmentRef": "hub-synthetic-mx-0002",
                    "environmentSerial": "9002",
                    "status": "success",
                    "currentStep": "done",
                    "completedSteps": ["done"],
                }]},
            },
            headers=device_headers(credential),
        )
        assert retry_progress.status_code == 200, retry_progress.text
        retry_finished = device_client.post(
            f"/v1/executor-channel/tasks/{retry_task_id}/finish",
            json={
                "leaseToken": retry_credential,
                "outcome": "succeeded",
                "resultCode": "environment_completed",
                "resultSummary": {
                    "runStatus": "completed",
                    "phase": "environment.completed",
                    "progressCompleted": 1,
                    "progressTotal": 1,
                    "totalCount": 1,
                    "successCount": 1,
                    "failedCount": 0,
                    "stoppedCount": 0,
                    "ipOkCount": 0,
                    "ipTotalCount": 0,
                },
            },
            headers=device_headers(credential),
        )
        assert retry_finished.status_code == 200, retry_finished.text
        retry_restored = web_client.get(
            f"/v1/operation-runs/environment-creation/{retry_snapshot['runId']}"
        ).json()["data"]
        assert retry_restored["status"] == "completed"
        assert retry_restored["rows"][0]["status"] == "success"
        stale_retry = web_client.post(
            f"/v1/operation-runs/environment-creation/{retry_snapshot['runId']}/retry",
            json={
                "idempotencyKey": "environment-retry-stale-0001",
                "retryMode": "failed",
                "accountRefs": ["sha256-progress-account-0002"],
            },
            headers=CSRF,
        )
        assert stale_retry.status_code == 409

        logistics = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                "idempotencyKey": "logistics-run-create-0001",
                "executorId": executor_id,
                "queryMode": "initial",
                "browserMode": "visible",
                "site": "MX",
                "environmentSerials": ["9001", "9002"],
            },
            headers=CSRF,
        )
        assert logistics.status_code == 202, logistics.text
        logistics_snapshot = logistics.json()["data"]
        assert logistics_snapshot["status"] == "queued"
        assert logistics_snapshot["progressTotal"] == 2
        cancelled_logistics = web_client.post(
            f"/v1/operation-runs/logistics-query/{logistics_snapshot['runId']}/cancel",
            json={},
            headers=CSRF,
        )
        assert cancelled_logistics.status_code == 200, cancelled_logistics.text
        assert cancelled_logistics.json()["data"]["status"] == "cancelled"
        assert cancelled_logistics.json()["data"]["stopRequested"] is True
        with database.session_factory() as session:
            logistics_run = session.scalar(select(LogisticsQueryRun))
            assert logistics_run is not None
            assert logistics_run.executor_task_id is not None


def test_offline_revision_conflict_revoke_and_sensitive_config_are_blocked(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(device_client, create_pairing_code(web_client))
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(device_client, credential, revision=REVISION_A)

        conflict = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json={
                "expectedRevision": REVISION_B,
                "config": {"concurrency": 2},
            },
            headers=CSRF,
        )
        assert conflict.status_code == 409
        assert conflict.json()["detail"]["code"] == "config_revision_conflict"

        sensitive = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json={
                "expectedRevision": REVISION_A,
                "config": {"proxyLink": "https://secret.invalid"},
            },
            headers=CSRF,
        )
        assert sensitive.status_code == 422
        nested_sensitive = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json={
                "expectedRevision": REVISION_A,
                "config": {"purchaseTags": {"password": "must-not-queue"}},
            },
            headers=CSRF,
        )
        assert nested_sensitive.status_code == 422
        for legacy_config in (
            {"purchaseSite": "MX"},
            {"purchaseTags": {"MX": "MX采购"}},
            {"importBuyerPlan": "1:新刚"},
            {"serverPort": 8765},
            {"hiddenQueryColumns": ["envName"]},
        ):
            legacy_write = web_client.put(
                f"/v1/executors/{executor_id}/config",
                json={
                    "expectedRevision": REVISION_A,
                    "config": legacy_config,
                },
                headers=CSRF,
            )
            assert legacy_write.status_code == 422

        with database.session_factory() as session:
            executor = session.scalar(select(LocalExecutor))
            assert executor is not None
            executor.last_seen_at = datetime.now(UTC) - timedelta(minutes=5)
            session.commit()
        offline = web_client.post(
            f"/v1/executors/{executor_id}/config/read",
            json={},
            headers=CSRF,
        )
        assert offline.status_code == 409
        assert offline.json()["detail"]["code"] == "executor_offline"

        heartbeat(device_client, credential, revision=REVISION_A)
        revoked = web_client.post(
            f"/v1/executors/{executor_id}/revoke",
            json={},
            headers=CSRF,
        )
        assert revoked.status_code == 200
        assert revoked.json()["executor"]["connectivity"] == "revoked"
        after_revoke = heartbeat_response = device_client.post(
            "/v1/executor-channel/poll",
            json={
                "waitSeconds": 0,
                "configRevision": REVISION_A,
                "hubStatus": "ready",
                "clientVersion": "0.12.5",
                "capabilities": ["config.read.v1", "config.write.v1"],
            },
            headers=device_headers(credential),
        )
        assert after_revoke.status_code == 401
        assert heartbeat_response.json()["detail"]["code"] == "executor_revoked"


def test_new_executor_reports_summary_and_rejects_cloud_config_writes(
    tmp_path,
) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = [
        "config.read.v1",
        "config.write.v1",
        "config.summary.v2",
        "local.config.desktop.v1",
        "workspace.rpc.v1",
    ]
    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
            config_summary=safe_config_summary(),
        )

        runtime = web_client.get(
            f"/v1/executors/{executor_id}/runtime-summary"
        )
        assert runtime.status_code == 200, runtime.text
        payload = runtime.json()
        assert payload["configSummary"]["schemaVersion"] == 2
        assert payload["configSummary"]["dataSources"] == {
            "dataSourceCount": 2,
            "buyerProfileCount": 1,
            "environmentBindingCount": 3,
            "pendingOwnerConfirmationCount": 1,
            "mappingConflictCount": 0,
        }
        assert payload["configSummaryStale"] is False
        workspace = web_client.get(
            f"/v1/executors/{executor_id}/workspace-snapshot"
        )
        assert workspace.status_code == 200
        assert workspace.json()["snapshot"] is None
        rendered = json.dumps(payload)
        for forbidden in (
            "spreadsheetToken", "sheetId", "appSecret", "apiKey",
            "containerCode", "environmentName",
        ):
            assert forbidden not in rendered

        read_blocked = web_client.post(
            f"/v1/executors/{executor_id}/config/read",
            json={}, headers=CSRF,
        )
        write_blocked = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json={
                "expectedRevision": REVISION_A,
                "config": {"concurrency": 3},
            },
            headers=CSRF,
        )
        rpc_blocked = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "POST", "path": "/api/lark/config", "body": {}},
            headers=CSRF,
        )
        for response in (read_blocked, write_blocked, rpc_blocked):
            assert response.status_code == 410, response.text
            assert response.json()["detail"]["code"] == (
                "local_config_desktop_only"
            )

        business_rpc = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/groups"},
            headers=CSRF,
        )
        assert business_rpc.status_code == 202, business_rpc.text


def test_config_summary_rejects_unknown_sensitive_fields(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = ["config.summary.v2", "local.config.desktop.v1"]
    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        summary = safe_config_summary()
        summary["spreadsheetToken"] = "must-never-be-accepted"
        response = device_client.post(
            "/v1/executor-channel/poll",
            json={
                "waitSeconds": 0,
                "configRevision": REVISION_A,
                "hubStatus": "ready",
                "clientVersion": "0.12.5",
                "protocolVersion": 1,
                "capabilities": capabilities,
                "configSummary": summary,
            },
            headers=device_headers(str(paired["deviceCredential"])),
        )
        assert response.status_code == 422


def test_web_device_access_is_tenant_isolated(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client:
        login(web_client)
        with database.session_factory() as session:
            other_tenant = Tenant(
                feishu_tenant_key="tenant_other",
                name="Other tenant",
                status="active",
            )
            session.add(other_tenant)
            session.flush()
            other_user = User(
                tenant_id=other_tenant.id,
                feishu_open_id="ou_other",
                display_name="Other user",
                status="active",
            )
            session.add(other_user)
            session.flush()
            other_executor = LocalExecutor(
                tenant_id=other_tenant.id,
                owner_user_id=other_user.id,
                display_name="Other device",
                platform="windows",
                architecture="x86_64",
                client_version="0.12.5",
                protocol_version=1,
                capabilities=["config.read.v1"],
                credential_digest=hash_token("other-device-credential-" + "x" * 40),
                status="active",
                hub_status="unknown",
            )
            session.add(other_executor)
            session.commit()
            other_id = str(other_executor.id)

        response = web_client.get(
            f"/v1/executors/{other_id}/runtime-summary"
        )
        assert response.status_code == 404
        assert response.json()["detail"]["code"] == "executor_not_found"
        listed = web_client.get("/v1/executors")
        assert listed.status_code == 200
        assert listed.json()["items"] == []


def test_web_device_and_task_access_is_owner_isolated_within_tenant(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client:
        login(web_client)
        with database.session_factory() as session:
            tenant = session.scalar(
                select(Tenant).where(Tenant.feishu_tenant_key == "tenant_allowed")
            )
            assert tenant is not None
            other_user = User(
                tenant_id=tenant.id,
                feishu_open_id="ou_same_tenant_other",
                display_name="Same tenant other user",
                status="active",
            )
            session.add(other_user)
            session.flush()
            other_executor = LocalExecutor(
                tenant_id=tenant.id,
                owner_user_id=other_user.id,
                display_name="Other member device",
                platform="windows",
                architecture="x86_64",
                client_version="0.12.7",
                protocol_version=1,
                capabilities=["config.read.v1", "config.write.v1"],
                credential_digest=hash_token(
                    "same-tenant-other-device-credential-" + "x" * 40
                ),
                status="active",
                hub_status="ready",
                last_seen_at=datetime.now(UTC),
            )
            session.add(other_executor)
            session.flush()
            other_task = ExecutorTask(
                tenant_id=tenant.id,
                executor_id=other_executor.id,
                task_type="config.read.v1",
                idempotency_key="same-tenant-other-task",
                payload_envelope={},
                created_by_user_id=other_user.id,
            )
            session.add(other_task)
            session.commit()
            other_executor_id = str(other_executor.id)
            other_task_id = str(other_task.id)

        listed = web_client.get("/v1/executors")
        assert listed.status_code == 200
        assert listed.json()["items"] == []

        protected_requests = (
            web_client.get(
                f"/v1/executors/{other_executor_id}/runtime-summary"
            ),
            web_client.post(
                f"/v1/executors/{other_executor_id}/config/read",
                json={},
                headers=CSRF,
            ),
            web_client.put(
                f"/v1/executors/{other_executor_id}/config",
                json={
                    "expectedRevision": REVISION_A,
                    "config": {"concurrency": 2},
                },
                headers=CSRF,
            ),
            web_client.post(
                f"/v1/executors/{other_executor_id}/revoke",
                json={},
                headers=CSRF,
            ),
        )
        for response in protected_requests:
            assert response.status_code == 404
            assert response.json()["detail"]["code"] == "executor_not_found"

        task_status = web_client.get(f"/v1/executor-tasks/{other_task_id}")
        assert task_status.status_code == 404
        assert task_status.json()["detail"]["code"] == "executor_task_not_found"
        task_cancel = web_client.post(
            f"/v1/executor-tasks/{other_task_id}/cancel",
            json={},
            headers=CSRF,
        )
        assert task_cancel.status_code == 404
        assert task_cancel.json()["detail"]["code"] == "executor_task_not_found"


def test_workspace_rpc_payload_and_result_are_encrypted_at_rest(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    capabilities = ["config.read.v1", "config.write.v1", "workspace.rpc.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )

        created = web_client.post(
            f"/v1/executors/{paired['executorId']}/workspace-rpc",
            json={
                "method": "POST",
                "path": "/api/envbatch/parse",
                "body": {
                    "filename": "buyers.xlsx",
                    "contentBase64": "secret-cookie-payload",
                },
            },
            headers=CSRF,
        )
        assert created.status_code == 202, created.text
        task_id = str(created.json()["task"]["id"])

        with database.session_factory() as session:
            task = session.get(ExecutorTask, uuid.UUID(task_id))
            assert task is not None
            stored = json.dumps(task.payload_envelope)
            assert "secret-cookie-payload" not in stored
            assert "encryptedPayload" in task.payload_envelope

        leased = heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )["task"]
        assert leased["id"] == task_id
        assert leased["payload"]["body"]["contentBase64"] == (
            "secret-cookie-payload"
        )
        lease_token = str(leased["leaseToken"])
        started = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/start",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        )
        assert started.status_code == 200, started.text
        with database.session_factory() as session:
            executor = session.scalar(select(LocalExecutor))
            assert executor is not None
            executor.last_seen_at = datetime.now(UTC) - timedelta(minutes=5)
            session.commit()
        renewed = device_client.put(
            f"/v1/executor-channel/tasks/{task_id}/lease",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        )
        assert renewed.status_code == 200, renewed.text
        with database.session_factory() as session:
            executor = session.scalar(select(LocalExecutor))
            assert executor is not None
            last_seen = executor.last_seen_at.replace(tzinfo=UTC)
            assert datetime.now(UTC) - last_seen < timedelta(seconds=5)
        finished = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/finish",
            json={
                "leaseToken": lease_token,
                "outcome": "succeeded",
                "resultCode": "workspace_rpc_completed",
                "resultSummary": {
                    "httpStatus": 200,
                    "responseType": "json",
                    "contentType": "application/json",
                    "body": {"password": "sensitive-result"},
                },
            },
            headers=device_headers(credential),
        )
        assert finished.status_code == 200, finished.text

        visible = web_client.get(f"/v1/executor-tasks/{task_id}")
        assert visible.status_code == 200, visible.text
        assert visible.json()["task"]["resultSummary"]["body"]["password"] == (
            "sensitive-result"
        )
        with database.session_factory() as session:
            task = session.get(ExecutorTask, uuid.UUID(task_id))
            assert task is not None
            stored = json.dumps(task.result_summary)
            assert "sensitive-result" not in stored
            assert "encryptedResult" in task.result_summary


def test_workspace_rpc_short_reads_queue_while_config_tasks_remain_exclusive(
    tmp_path,
) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = ["config.read.v1", "config.write.v1", "workspace.rpc.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])

        first = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/groups"},
            headers=CSRF,
        )
        second = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/tasks"},
            headers=CSRF,
        )
        assert first.status_code == 202, first.text
        assert second.status_code == 202, second.text
        assert first.json()["task"]["id"] != second.json()["task"]["id"]
        duplicate = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/groups"},
            headers=CSRF,
        )
        assert duplicate.status_code == 202, duplicate.text
        assert duplicate.json()["task"]["id"] == first.json()["task"]["id"]

        config_read = web_client.post(
            f"/v1/executors/{executor_id}/config/read",
            json={},
            headers=CSRF,
        )
        assert config_read.status_code == 409
        assert config_read.json()["detail"]["code"] == "executor_task_busy"


def test_formal_logistics_run_requires_online_executor_and_ready_hubstudio(
    tmp_path,
) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = ["logistics.query.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        base_payload = {
            "executorId": executor_id,
            "queryMode": "initial",
            "site": "MX",
            "environmentSerials": ["5564"],
        }

        offline = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                **base_payload,
                "idempotencyKey": "logistics-offline-blocked-0001",
            },
            headers=CSRF,
        )
        assert offline.status_code == 409, offline.text
        assert offline.json()["detail"]["code"] == "executor_offline"

        heartbeat(
            device_client,
            credential,
            capabilities=capabilities,
            hub_status="limited",
        )
        unavailable = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                **base_payload,
                "idempotencyKey": "logistics-hub-blocked-0001",
            },
            headers=CSRF,
        )
        assert unavailable.status_code == 409, unavailable.text
        assert unavailable.json()["detail"]["code"] == \
            "executor_hub_unavailable"

        heartbeat(
            device_client,
            credential,
            capabilities=capabilities,
            hub_status="ready",
            client_version="0.13.17",
        )
        upgrade_required = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                **base_payload,
                "idempotencyKey": "logistics-upgrade-blocked-0001",
            },
            headers=CSRF,
        )
        assert upgrade_required.status_code == 409, upgrade_required.text
        assert upgrade_required.json()["detail"]["code"] == \
            "executor_upgrade_required"


def test_formal_logistics_run_queues_ahead_of_background_workspace_read(
    tmp_path,
) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    capabilities = ["workspace.rpc.v1", "logistics.query.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )

        background = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/groups"},
            headers=CSRF,
        )
        assert background.status_code == 202, background.text

        formal = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                "idempotencyKey": "logistics-priority-over-read-0001",
                "executorId": executor_id,
                "queryMode": "initial",
                "browserMode": "visible",
                "site": "MX",
                "environmentSerials": ["5564"],
            },
            headers=CSRF,
        )
        assert formal.status_code == 202, formal.text
        leased = heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )["task"]
        assert leased["type"] == "logistics.query.v1"
        assert leased["payload"]["browserMode"] == "visible"

        with database.session_factory() as session:
            tasks = list(
                session.scalars(
                    select(ExecutorTask).order_by(ExecutorTask.priority.asc())
                )
            )
            assert [task.task_type for task in tasks] == [
                "logistics.query.v1",
                "workspace.rpc.v1",
            ]
            assert [task.priority for task in tasks] == [10, 100]


def test_formal_logistics_progress_preserves_complete_headless_row(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = ["workspace.rpc.v1", "logistics.query.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )
        created = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                "idempotencyKey": "logistics-progress-complete-row-0001",
                "executorId": executor_id,
                "queryMode": "initial",
                "site": "MX",
                "environmentSerials": ["4973"],
            },
            headers=CSRF,
        )
        assert created.status_code == 202, created.text
        run = created.json()["data"]
        leased = heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )["task"]
        task_id = str(leased["id"])
        lease_token = str(leased["leaseToken"])
        assert task_id == run["executorTaskId"]
        assert device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/start",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        ).status_code == 200

        progress = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/progress",
            json={
                "leaseToken": lease_token,
                "phase": "logistics.running",
                "current": 1,
                "total": 1,
                "snapshot": {"rows": [{
                    "environmentSerial": "4973",
                    "environmentName": "XG-MX-0829-105",
                    "status": "ok",
                    "currentStep": "ok",
                    "completedSteps": ["query_completed"],
                    "platformOrderNo": "GSH1RD06000M253",
                    "orderTime": "2026-09-01 10:20:30",
                    "amount": "$MXN123.45",
                    "platformStatus": "Enviado",
                    "statusLabel": "已发货",
                    "fulfillmentStage": "En tránsito.",
                    "trackingNumbers": ["49426326790694"],
                    "packageNumbers": ["PKG-001"],
                    "carrier": "IMILE",
                    "cancelled": False,
                    "riskOrder": False,
                    "riskSummary": "",
                    "ipAddress": "203.0.113.10",
                    "timeZone": "America/Mexico_City",
                    "utcOffsetMinutes": -360,
                    "queriedAt": "2026-09-01T10:21:31-06:00",
                    "errorSummary": "",
                    "screenshotStatus": "ok",
                }]},
            },
            headers=device_headers(credential),
        )
        assert progress.status_code == 200, progress.text
        restored = web_client.get(
            f"/v1/operation-runs/logistics-query/{run['runId']}"
        ).json()["data"]
        row = restored["rows"][0]
        assert row["orderTime"] == "2026-09-01 10:20:30"
        assert row["amount"] == "$MXN123.45"
        assert row["fulfillmentStage"] == "En tránsito."
        assert row["ipAddress"] == "203.0.113.10"
        assert row["queriedAt"] == "2026-09-01T10:21:31"
        assert row["screenshotStatus"] == "ok"

        finished = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/finish",
            json={
                "leaseToken": lease_token,
                "outcome": "succeeded",
                "resultCode": "logistics_completed",
                "resultSummary": {
                    "runStatus": "completed",
                    "phase": "logistics.completed",
                    "progressCompleted": 1,
                    "progressTotal": 1,
                    "totalCount": 1,
                    "successCount": 1,
                    "failedCount": 0,
                    "stoppedCount": 0,
                    "errorCode": "",
                    "errorSummary": "",
                },
            },
            headers=device_headers(credential),
        )
        assert finished.status_code == 200, finished.text
        terminal = web_client.get(
            f"/v1/operation-runs/logistics-query/{run['runId']}"
        ).json()["data"]
        assert terminal["status"] == "completed"


def test_busy_executor_config_read_returns_last_safe_snapshot(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)
    capabilities = ["config.read.v1", "workspace.rpc.v1"]

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client,
            credential,
            revision=REVISION_A,
            capabilities=capabilities,
        )
        captured_at = datetime.now(UTC)
        with database.session_factory() as session:
            executor = session.get(LocalExecutor, uuid.UUID(executor_id))
            assert executor is not None
            executor.workspace_snapshot = {
                "runtimeConfig": {
                    "configRevision": REVISION_A,
                    "hubPort": 6873,
                    "concurrency": 2,
                    "envCreateWorkers": 5,
                    "verifySampleCount": 3,
                    "safeParallelTasks": True,
                }
            }
            executor.workspace_snapshot_at = captured_at
            session.commit()

        active = web_client.post(
            f"/v1/executors/{executor_id}/workspace-rpc",
            json={"method": "GET", "path": "/api/groups"},
            headers=CSRF,
        )
        assert active.status_code == 202, active.text

        config_read = web_client.post(
            f"/v1/executors/{executor_id}/config/read",
            json={},
            headers=CSRF,
        )
        assert config_read.status_code == 202, config_read.text
        payload = config_read.json()
        assert payload["task"] is None
        assert payload["cached"] is True
        assert payload["cachedResult"]["configRevision"] == REVISION_A
        assert payload["cachedResult"]["config"] == {
            "hubPort": 6873,
            "concurrency": 2,
            "envCreateWorkers": 5,
            "verifySampleCount": 3,
            "safeParallelTasks": True,
            "queryBrowserMode": "headless",
        }
        assert payload["cachedResult"]["capturedAt"]


def test_started_config_write_becomes_uncertain_after_lease_expiry(tmp_path) -> None:
    app, database, _oauth = build_test_app(tmp_path)

    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(device_client, create_pairing_code(web_client))
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(device_client, credential, revision=REVISION_A)
        created = web_client.put(
            f"/v1/executors/{executor_id}/config",
            json={
                "expectedRevision": REVISION_A,
                "config": {"concurrency": 3},
            },
            headers=CSRF,
        )
        assert created.status_code == 202
        task_id = created.json()["task"]["id"]
        lease = heartbeat(device_client, credential, revision=REVISION_A)["task"]
        started = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/start",
            json={"leaseToken": lease["leaseToken"]},
            headers=device_headers(credential),
        )
        assert started.status_code == 200

        with database.session_factory() as session:
            task = session.get(ExecutorTask, uuid.UUID(task_id))
            assert task is not None
            task.lease_until = datetime.now(UTC) - timedelta(seconds=1)
            session.commit()

        assert heartbeat(device_client, credential, revision=REVISION_A)["task"] is None
        status_response = web_client.get(f"/v1/executor-tasks/{task_id}")
        assert status_response.status_code == 200
        task_payload = status_response.json()["task"]
        assert task_payload["status"] == "uncertain"
        assert task_payload["resultCode"] == "lease_expired_after_start"


def test_logistics_retry_keeps_parent_order_and_serves_uploaded_screenshot(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    capabilities = ["workspace.rpc.v1", "logistics.query.v1"]
    with TestClient(app) as web_client, TestClient(app) as device_client:
        login(web_client)
        paired = pair(
            device_client,
            create_pairing_code(web_client),
            capabilities=capabilities,
        )
        executor_id = str(paired["executorId"])
        credential = str(paired["deviceCredential"])
        heartbeat(
            device_client, credential, revision=REVISION_A,
            capabilities=capabilities, client_version="0.13.19",
        )
        created = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                "idempotencyKey": "logistics-parent-merge-0001",
                "executorId": executor_id,
                "queryMode": "initial",
                "site": "MX",
                "environmentSerials": ["20", "10"],
            },
            headers=CSRF,
        )
        assert created.status_code == 202, created.text
        root = created.json()["data"]
        task = heartbeat(
            device_client, credential, revision=REVISION_A,
            capabilities=capabilities, client_version="0.13.19",
        )["task"]
        lease_token = task["leaseToken"]
        task_id = task["id"]
        assert device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/start",
            json={"leaseToken": lease_token},
            headers=device_headers(credential),
        ).status_code == 200
        screenshot = b"\xff\xd8synthetic-jpeg-content\xff\xd9"
        rows = [
            {
                "environmentSerial": serial,
                "environmentName": f"ENV-{serial}",
                "status": "ok" if serial == "20" else "fail",
                "currentStep": "ok" if serial == "20" else "fail",
                "completedSteps": ["query_completed"],
                "platformOrderNo": "ORDER-20" if serial == "20" else "",
                "platformStatus": "Enviado" if serial == "20" else "",
                "statusLabel": "已发货" if serial == "20" else "",
                "screenshotStatus": "ok" if serial == "20" else "",
                "screenshotSizeKb": 1 if serial == "20" else 0,
                "errorSummary": "" if serial == "20" else "synthetic failure",
            }
            for serial in ["20", "10"]
        ]
        progress = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/progress",
            json={
                "leaseToken": lease_token,
                "phase": "logistics.running",
                "current": 2,
                "total": 2,
                "snapshot": {
                    "rows": rows,
                    "screenshots": [{
                        "environmentSerial": "20",
                        "contentType": "image/jpeg",
                        "contentBase64": base64.b64encode(screenshot).decode("ascii"),
                        "sha256": hashlib.sha256(screenshot).hexdigest(),
                        "size": len(screenshot),
                    }],
                },
            },
            headers=device_headers(credential),
        )
        assert progress.status_code == 200, progress.text
        finished = device_client.post(
            f"/v1/executor-channel/tasks/{task_id}/finish",
            json={
                "leaseToken": lease_token,
                "outcome": "succeeded",
                "resultCode": "logistics_partial_failure",
                "resultSummary": {
                    "runStatus": "partial_failure",
                    "phase": "logistics.completed",
                    "progressCompleted": 2,
                    "progressTotal": 2,
                    "totalCount": 2,
                    "successCount": 1,
                    "failedCount": 1,
                    "stoppedCount": 0,
                    "errorCode": "logistics.partial_failure",
                    "errorSummary": "synthetic partial failure",
                },
            },
            headers=device_headers(credential),
        )
        assert finished.status_code == 200, finished.text
        exported = web_client.get(
            f"/v1/operation-runs/logistics-query/{root['runId']}/export"
        )
        assert exported.status_code == 200, exported.text
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        assert [
            str(workbook.active.cell(row=index, column=1).value)
            for index in (2, 3)
        ] == ["20", "10"]
        workbook.close()

        retry = web_client.post(
            "/v1/operation-runs/logistics-query",
            json={
                "idempotencyKey": "logistics-child-merge-0001",
                "executorId": executor_id,
                "queryMode": "single_retry",
                "parentRunId": root["runId"],
                "force": False,
                "site": "MX",
                "environmentSerials": ["10"],
            },
            headers=CSRF,
        )
        assert retry.status_code == 202, retry.text
        child = retry.json()["data"]
        assert child["parentRunId"] == root["runId"]
        assert [row["environmentSerial"] for row in child["rows"]] == ["20", "10"]
        shot = web_client.get(
            f"/v1/operation-runs/logistics-query/{child['runId']}/screenshots/20"
        )
        assert shot.status_code == 200
        assert shot.content == screenshot
        assert "no-store" in shot.headers["cache-control"]


def test_workspace_view_preferences_round_trip(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    with TestClient(app) as client:
        login(client)
        missing = client.get(
            "/v1/workspace/view-preferences/fulfillment.logistics.results"
        )
        assert missing.status_code == 200
        assert missing.json()["settings"] is None
        saved = client.put(
            "/v1/workspace/view-preferences/fulfillment.logistics.results",
            json={
                "schemaVersion": 1,
                "visibleFields": ["serial", "status", "result", "action"],
                "fieldOrder": ["serial", "status", "result", "action"],
            },
            headers=CSRF,
        )
        assert saved.status_code == 200, saved.text
        restored = client.get(
            "/v1/workspace/view-preferences/fulfillment.logistics.results"
        ).json()
        assert restored["settings"]["visibleFields"] == [
            "serial", "status", "result", "action"
        ]
