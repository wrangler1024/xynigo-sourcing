from __future__ import annotations

from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from xynigo_auth.logistics_export import (
    HEADERS,
    LogisticsWorkbookExport,
    build_logistics_workbook,
    build_logistics_workbook_export,
)


JPEG = b"\xff\xd8synthetic-logistics-screenshot\xff\xd9"


def _rows() -> list[dict[str, object]]:
    return [
        {
            "environmentSerial": "20",
            "environmentName": "ENV-20",
            "platformOrderNo": "ORDER-20",
            "orderTime": "2026-09-03 10:00:00",
            "amount": "MXN 20",
            "status": "ok",
            "platformStatus": "Enviado",
            "statusLabel": "已发货",
            "trackingNumbers": ["TRACK-20-A", "TRACK-20-B"],
            "packageNumbers": ["PKG-20"],
            "carrier": "iMile",
            "screenshotStatus": "ok",
            "ipAddress": "192.0.2.20",
            "queriedAt": "2026-09-03T10:01:00+08:00",
        },
        {
            "environmentSerial": "10",
            "environmentName": "ENV-10",
            "status": "fail",
            "screenshotStatus": "ok",
            "errorSummary": "synthetic failure",
        },
        {
            "environmentSerial": "30",
            "environmentName": "ENV-30",
            "status": "pending",
            "screenshotStatus": "none",
        },
    ]


def _sheet_values(content: bytes) -> list[tuple[object, ...]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    values = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    return values


def test_quick_export_does_not_read_or_embed_screenshots() -> None:
    calls: list[str] = []

    def reader(serial: str) -> bytes:
        calls.append(serial)
        raise AssertionError("quick export must not read screenshots")

    exported = build_logistics_workbook_export(
        _rows(), reader, include_screenshots=False
    )

    assert isinstance(exported, LogisticsWorkbookExport)
    assert calls == []
    assert exported.included_screenshot_count == 0
    assert exported.missing_screenshot_count == 0
    values = _sheet_values(exported.content)
    assert list(values[0]) == HEADERS
    assert [row[0] for row in values[1:]] == ["20", "10", "30"]
    assert [row[9] for row in values[1:]] == [
        "已生成（未导出）",
        "已生成（未导出）",
        "未生成",
    ]
    with ZipFile(BytesIO(exported.content)) as archive:
        assert not any(name.startswith("xl/media/") for name in archive.namelist())


def test_complete_export_embeds_available_jpeg_and_keeps_missing_as_text() -> None:
    calls: list[str] = []

    def reader(serial: str) -> bytes | None:
        calls.append(serial)
        if serial == "20":
            return JPEG
        raise FileNotFoundError("screenshot expired")

    exported = build_logistics_workbook_export(
        _rows(), reader, include_screenshots=True
    )

    assert calls == ["20", "10"]
    assert exported.included_screenshot_count == 1
    assert exported.missing_screenshot_count == 1
    with ZipFile(BytesIO(exported.content)) as archive:
        assert archive.read("xl/media/image1.jpeg") == JPEG
        assert not any(name == "xl/media/image2.jpeg" for name in archive.namelist())
    values = _sheet_values(exported.content)
    assert [row[0] for row in values[1:]] == ["20", "10", "30"]
    assert values[2][9] == "截图已过期或缺失"
    assert values[3][9] == "未生成"


def test_workbook_fields_and_legacy_bytes_contract() -> None:
    content = build_logistics_workbook(
        _rows(), include_screenshots=False
    )

    assert isinstance(content, bytes)
    values = _sheet_values(content)
    first = values[1]
    assert first == (
        "20",
        "ENV-20",
        "ORDER-20",
        "2026-09-03 10:00:00",
        "MXN 20",
        "Enviado 已发货",
        "TRACK-20-A; TRACK-20-B",
        "PKG-20",
        "iMile",
        "已生成（未导出）",
        "192.0.2.20",
        "成功",
        None,
        "2026-09-03T10:01:00+08:00",
    )
