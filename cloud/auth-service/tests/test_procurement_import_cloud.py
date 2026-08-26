from __future__ import annotations

import base64
import json
import threading
import time
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from test_auth_flow import build_test_app, start_login

from xynigo_auth.models import AuditEvent, ProcurementImportJob, ProcurementImportPlan
from xynigo_auth.procurement_import_core import OUTPUT_HEADERS
from xynigo_auth.procurement_import_sheet import SheetTable


class FakeCloudSheetGateway:
    def __init__(self) -> None:
        self.headers = tuple(OUTPUT_HEADERS)
        self.rows: tuple[tuple[int, tuple[object, ...]], ...] = ()
        self.backgrounds: dict[int, str] = {}
        self.links: dict[int, str] = {}
        self.lock = threading.Lock()

    def inspect(self, url):
        return {
            "url": "https://tenant.feishu.cn/sheets/SheetToken123",
            "spreadsheetToken": "SheetToken123",
            "revision": 1,
            "sheets": [
                {
                    "sheetId": "sheetA",
                    "sheetName": "采购执行协作区",
                    "rowCount": 200,
                    "columnCount": len(self.headers),
                    "hidden": False,
                }
            ],
        }

    def read_table(self, url, sheet_id):
        with self.lock:
            return SheetTable(self.headers, self.rows, revision=1)

    def append_table_rows(self, url, sheet_name, columns, rows, dtypes=None, formats=None):
        with self.lock:
            next_row = max([number for number, _values in self.rows] or [1]) + 1
            current = list(self.rows)
            current.extend(
                (next_row + index, tuple(row)) for index, row in enumerate(rows)
            )
            self.rows = tuple(current)
        return {"updated_rows_count": len(rows)}

    def normalize_collaboration_headers(self, *args, **kwargs):
        return {"operations": 0, "skipped": True}

    def reorder_collaboration_headers(self, *args, **kwargs):
        return {"operations": 0, "skipped": True}

    def normalize_date_column(self, *args, **kwargs):
        return {"operations": 0, "skipped": True}

    def apply_header_presentation(self, *args, **kwargs):
        return {"operations": 1}

    def row_backgrounds(self, url, sheet_id, row_numbers):
        return {int(row): self.backgrounds.get(int(row), "") for row in row_numbers}

    def apply_row_presentation(
        self, url, sheet_name, background_bands, row_ranges, row_height=52, last_column="AQ"
    ):
        for item in background_bands:
            for row in range(int(item["start"]), int(item["end"]) + 1):
                self.backgrounds[row] = str(item["color"])
        return {"operations": 1}

    def hyperlink_presence(self, url, sheet_id, expected_links, column="M"):
        return {
            int(row): self.links.get(int(row)) == link
            for row, link in dict(expected_links).items()
        }

    def set_hyperlinks(self, url, sheet_id, links, column="M"):
        for row, link in links:
            self.links[int(row)] = str(link)
        return {"operations": 1}

    def image_presence(self, url, sheet_id, row_numbers, column="L"):
        # Synthetic cloud fixture has no embedded image bytes. Pretend the
        # target already contains them so this test stays entirely offline.
        return {int(row): True for row in row_numbers}

    def set_image(self, *args, **kwargs):
        raise AssertionError("existing synthetic images must be skipped")

    def verify_image(self, *args, **kwargs):
        return True


def source_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "order_"
    worksheet.append(
        [
            "店铺账号",
            "订单号",
            "包裹号",
            "下单时间",
            "订单金额",
            "币种缩写",
            "收货人姓名",
            "收货人国家",
            "收货人州/省",
            "收货人城市",
            "地址1",
            "地址2",
            "邮编",
            "收货人电话",
            "SKU",
            "产品规格",
            "产品售价",
            "单个产品数量",
            "产品图片网址",
            "客服备注",
            "产品图片",
        ]
    )
    xyp2 = {
        "d": "mx",
        "c": "MXN",
        "i": [
            [
                "SOURCE-01",
                "422790137",
                "I8mmn32aip2g7d",
                "27_447",
                "Multicolor",
                "M",
                110.09,
                0.65,
                38.53,
                1,
            ]
        ],
    }
    worksheet.append(
        [
            "测试店铺-测试运营（二组）$",
            "GSH-CLOUD-TEST-001",
            "PKG-CLOUD-TEST-001",
            "2026-08-26 12:00:00",
            150,
            "MXN",
            "Recipient Test",
            "MEXICO",
            "State",
            "City",
            "Address 1",
            "Address 2",
            "00123",
            "0012345678",
            "ERP-SKU-01",
            "SOURCE-01:Multicolor-M",
            150,
            1,
            "https://img.ltwebstatic.com/test/source-one.jpg",
            "[XYP2]" + json.dumps(xyp2, separators=(",", ":")) + "[/XYP2]",
            "",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def login(client: TestClient) -> None:
    state, _challenge = start_login(client)
    response = client.get(
        "/v1/auth/feishu/callback",
        params={"code": "authorization-code", "state": state},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_cloud_parse_validate_export_and_durable_worker(tmp_path) -> None:
    gateway = FakeCloudSheetGateway()
    app, database, _oauth = build_test_app(
        tmp_path,
        procurement_import_enabled=True,
        procurement_import_gateway=gateway,
    )
    headers = {"X-Xynigo-Web-CSRF": "same-origin"}
    with TestClient(app) as client:
        login(client)
        parsed = client.post(
            "/v1/assistant/procurement-import/parse",
            headers=headers,
            json={
                "filename": "order_cloud_test.xlsx",
                "contentBase64": base64.b64encode(source_workbook()).decode("ascii"),
            },
        )
        assert parsed.status_code == 201, parsed.text
        plan = parsed.json()
        assert plan["runtime"] == "cloud"
        assert plan["orderCount"] == 1
        assert plan["detailCount"] == 1
        assert plan["preview"][0]["orderNo"] == "GSH-CLOUD-TEST-001"

        with database.session_factory() as session:
            stored = session.scalar(select(ProcurementImportPlan))
            assert stored is not None
            assert stored.status == "parsed"
            assert b"GSH-CLOUD-TEST-001" not in bytes(stored.encrypted_payload or b"")

        inspected = client.post(
            "/v1/assistant/procurement-import/target/inspect",
            headers=headers,
            json={
                "planId": plan["planId"],
                "spreadsheetUrl": "https://tenant.feishu.cn/sheets/SheetToken123",
            },
        )
        assert inspected.status_code == 200, inspected.text
        assert inspected.json()["sheets"][0]["sheetId"] == "sheetA"

        validated = client.post(
            "/v1/assistant/procurement-import/target/validate",
            headers=headers,
            json={
                "planId": plan["planId"],
                "spreadsheetUrl": "https://tenant.feishu.cn/sheets/SheetToken123",
                "sheetId": "sheetA",
            },
        )
        assert validated.status_code == 200, validated.text
        assert validated.json()["valid"] is True

        exported = client.get(
            "/v1/assistant/procurement-import/export",
            params={"planId": plan["planId"]},
        )
        assert exported.status_code == 200, exported.text
        workbook = load_workbook(BytesIO(exported.content), data_only=False)
        assert workbook.active["C2"].value == "GSH-CLOUD-TEST-001"

        started = client.post(
            "/v1/assistant/procurement-import/sheet-sync",
            headers=headers,
            json={"planId": plan["planId"], "confirmWrite": True},
        )
        assert started.status_code == 202, started.text
        job_id = started.json()["jobId"]
        deadline = time.time() + 5
        status_payload = None
        while time.time() < deadline:
            status_response = client.get(
                "/v1/assistant/procurement-import/sheet-sync/status",
                params={"jobId": job_id},
            )
            assert status_response.status_code == 200, status_response.text
            status_payload = status_response.json()
            if status_payload["state"] in {"completed", "partial", "failed"}:
                break
            time.sleep(0.05)
        assert status_payload is not None
        assert status_payload["state"] == "completed", status_payload
        assert status_payload["rowsWritten"] == 1
        assert len(gateway.rows) == 1
        operator_index = gateway.headers.index("导入操作人")
        assert gateway.rows[0][1][operator_index] == "合成测试用户"

        with database.session_factory() as session:
            job = session.scalar(select(ProcurementImportJob))
            assert job is not None
            assert job.state == "completed"
            assert job.progress["errors"] == []
            audit_payload = json.dumps(
                [
                    {
                        "action": event.action,
                        "details": event.details,
                        "changeSummary": event.change_summary,
                    }
                    for event in session.scalars(select(AuditEvent))
                ],
                ensure_ascii=False,
            )
            assert "GSH-CLOUD-TEST-001" not in audit_payload
            assert "Recipient Test" not in audit_payload


def test_cloud_import_endpoint_is_fail_closed_when_feature_is_disabled(tmp_path) -> None:
    app, _database, _oauth = build_test_app(tmp_path)
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/v1/assistant/procurement-import/parse",
            headers={"X-Xynigo-Web-CSRF": "same-origin"},
            json={"filename": "x.xlsx", "contentBase64": "eA=="},
        )
        assert response.status_code == 503
        assert response.json()["detail"]["code"] == "procurement_import_cloud_disabled"


def test_cloud_parser_copy_matches_the_canonical_local_source() -> None:
    root = Path(__file__).resolve().parents[3]
    source = (root / "src/purchase_tool/procurement_import.py").read_text(
        encoding="utf-8"
    )
    expected = source.replace(
        "from .xlsx_cell_images import embed_cell_images",
        "from .procurement_import_xlsx import embed_cell_images",
    ).replace(
        "from .lark_sheet_sync import LarkCliSheetsGateway, LarkSheetSyncError",
        "from .procurement_import_sheet import FeishuSheetsGateway as LarkCliSheetsGateway, LarkSheetSyncError",
    )
    generated = (
        root / "cloud/auth-service/src/xynigo_auth/procurement_import_core.py"
    ).read_text(encoding="utf-8")
    assert generated.endswith(expected)
    assert (
        root / "cloud/auth-service/src/xynigo_auth/procurement_import_xlsx.py"
    ).read_bytes() == (root / "src/purchase_tool/xlsx_cell_images.py").read_bytes()
