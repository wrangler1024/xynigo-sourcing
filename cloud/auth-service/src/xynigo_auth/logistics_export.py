"""Build the cloud-owned logistics workbook from the durable merged run view."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .procurement_import_xlsx import embed_cell_images


HEADERS = [
    "环境序号", "环境名", "订单号", "下单时间", "金额", "状态",
    "物流单号", "包裹号", "承运商", "物流轨迹截图", "出口IP",
    "结果", "失败原因", "查询时间（站点）",
]


@dataclass(frozen=True, slots=True)
class LogisticsWorkbookExport:
    """Generated workbook bytes and screenshot embedding outcome."""

    content: bytes
    included_screenshot_count: int
    missing_screenshot_count: int


def _screenshot_text(status: object, *, include_screenshots: bool) -> str:
    normalized = str(status or "").strip().lower()
    if normalized == "ok":
        return "查看截图" if include_screenshots else "已生成（未导出）"
    if normalized == "expired":
        return "截图已过期"
    if normalized in {"failed", "fail", "error"}:
        return "生成失败"
    if normalized in {"pending", "running"}:
        return "生成中"
    return "未生成"


def build_logistics_workbook_export(
    rows: list[dict[str, object]],
    screenshot_reader: Callable[[str], bytes | None] | None = None,
    *,
    include_screenshots: bool = True,
) -> LogisticsWorkbookExport:
    """Build a workbook and expose screenshot inclusion metadata.

    In quick-export mode the screenshot reader is never called.  In complete
    mode a missing, expired, unreadable, or invalid JPEG is represented as
    text in the workbook instead of failing the entire export.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "物流单号查询"
    sheet.append(HEADERS)
    images: list[tuple[str, bytes]] = []
    missing_screenshot_count = 0
    side = Side(style="thin", color="D8E2EA")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row_index, row in enumerate(rows, start=2):
        state = str(row.get("status") or "")
        result = {
            "ok": "成功", "login": "登录失效", "inuse": "环境使用中，已跳过",
            "fail": "失败", "running": "查询中", "pending": "未查询",
            "stopped": "已停止",
        }.get(state, state)
        if row.get("riskOrder"):
            result = "风险订单（待验证）"
        elif row.get("cancelled"):
            result = "成功（砍单退款中）"
        serial = str(row.get("environmentSerial") or "")
        screenshot_status = row.get("screenshotStatus")
        screenshot_text = _screenshot_text(
            screenshot_status, include_screenshots=include_screenshots
        )
        sheet.append([
            serial,
            row.get("environmentName") or "",
            row.get("platformOrderNo") or "",
            row.get("orderTime") or "",
            row.get("amount") or "",
            " ".join(filter(None, [
                str(row.get("platformStatus") or ""),
                str(row.get("statusLabel") or ""),
            ])),
            "; ".join(row.get("trackingNumbers") or []),
            "; ".join(row.get("packageNumbers") or []),
            row.get("carrier") or "",
            screenshot_text,
            row.get("ipAddress") or "",
            result,
            row.get("errorSummary") or "",
            row.get("queriedAt") or "",
        ])
        if include_screenshots and str(screenshot_status or "").lower() == "ok":
            image: bytes | None = None
            if screenshot_reader is not None:
                try:
                    image = screenshot_reader(serial)
                except Exception:
                    image = None
            if (
                isinstance(image, bytes)
                and image.startswith(b"\xff\xd8")
                and image.endswith(b"\xff\xd9")
            ):
                images.append((f"J{row_index}", image))
                sheet.row_dimensions[row_index].height = 72
            else:
                missing_screenshot_count += 1
                sheet.cell(row=row_index, column=10).value = "截图已过期或缺失"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123B63")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    widths = [10, 24, 22, 20, 14, 18, 28, 22, 14, 18, 16, 20, 32, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(1, len(rows) + 1)}"
    sheet.sheet_view.showGridLines = False
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = embed_cell_images(output.getvalue(), images)
    return LogisticsWorkbookExport(
        content=content,
        included_screenshot_count=len(images),
        missing_screenshot_count=missing_screenshot_count,
    )


def build_logistics_workbook(
    rows: list[dict[str, object]],
    screenshot_reader: Callable[[str], bytes | None] | None = None,
    *,
    include_screenshots: bool = True,
) -> bytes:
    """Return workbook bytes while preserving the original call contract."""

    return build_logistics_workbook_export(
        rows,
        screenshot_reader,
        include_screenshots=include_screenshots,
    ).content
