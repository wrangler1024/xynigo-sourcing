# -*- coding: utf-8 -*-
"""模块三：号商 xlsx → HubStudio 批量采购环境。

默认 CLI 只做离线 dry-run。真实创建必须同时给出 ``--apply`` 和
``--confirm-env-write``。凭证只驻内存；断点文件只保存脱敏状态。
"""
import argparse
import csv
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
import hashlib
import json
import os
from pathlib import Path
import random
import re
import sys
import tempfile
import threading
import time
from urllib.parse import parse_qsl, unquote, urlsplit
import urllib.request

from .hub_api import DEFAULT_PORT, HubStudioApi
from .redaction import mask_email, scrub_text


SITE_ALIASES = {'MX': '希音墨西哥站', 'US': '希音美国站'}
DOMAINS = {
    'MX': 'https://www.shein.com.mx',
    'US': 'https://us.shein.com',
}
SUPPORTED_SITES = frozenset(SITE_ALIASES)

# (width, height, weight)。创建后不得再改指纹。
RES_POOL = (
    (2560, 1440, 0.20),
    (1920, 1080, 0.30),
    (1920, 1200, 0.12),
    (1680, 1050, 0.10),
    (1600, 900, 0.08),
    (1536, 864, 0.08),
    (1440, 900, 0.06),
    (1366, 768, 0.06),
)

STEPS = ('env_created', 'cookie_imported', 'account_bound', 'remarked',
         'done')
MAPPING_HEADERS = ('邮箱', '环境名', 'HUB序号', '采购员', '绑定时间', '状态')
EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
ORDER_RE = re.compile(r'(?:[?&])orderNo=([0-9a-f]+)(?:[&#]|$)', re.I)
REMARK_ORDER_RE = re.compile(r'(?:^|\|\s*)单号:([0-9a-f]+)', re.I)
EMAIL_KEY_PATH = '/api/boobar-graph'
MAIL_KEY_PATH = '/api'
ENV_NAME_RE = r'^{code}-{site}-{mmdd}-(\d{{3}})$'
TEST_ENV_NAME_RE = r'^{code}-{site}-测试-(\d{{2}})$'

# 采购员名单写死（Jeff 2026-08-19 确认）：(姓名, 英文代号)。
# HUB 环境名统一用英文代号；页面显示「姓名-代号」。变更需改代码随版本发布。
BUYER_ROSTER = (
    ('新刚', 'XG'),
    ('志恒', 'ZH'),
    ('康德', 'KD'),
    ('宇航', 'YH'),
)
BUYER_NAMES = tuple(name for name, _code in BUYER_ROSTER)
BUYER_CODES = dict(BUYER_ROSTER)
BUYER_NAME_BY_CODE = {code: name for name, code in BUYER_ROSTER}
# 默认均分范围（Jeff 确认）：宇航暂未正式加入采购团队，不参与默认均分。
DEFAULT_SPLIT_BUYERS = ('新刚', '志恒', '康德')

# 备用/测试环境（不绑号）：单批上限 Jeff 确认为 25；备注与存量批次一致。
BACKUP_MAX_COUNT = 25
BACKUP_TYPES = ('备用', '测试')
BACKUP_REMARK = {'备用': '备用环境', '测试': '测试环境'}
BACKUP_RESULT_HEADERS = ('环境名', 'HUB序号', 'containerCode', '站点', '状态')
VENDOR_TEMPLATE_HEADERS = ('邮箱账号', '密码', '接码Key链接', 'Cookie')
VENDOR_HEADER_ALIASES = (
    frozenset({'邮箱', '邮箱账号', '邮箱地址', '账号邮箱', '账号',
               'email', 'emailaddress', 'buyeremail',
               'buyeremailaddress', 'accountemail'}),
    frozenset({'密码', '账号密码', '登录密码', 'password',
               'accountpassword'}),
    frozenset({'接码key', '接码key链接', '接码链接', '接码url',
               '验证码链接', '邮箱接码key', 'verificationurl',
               'verificationcodeurl', 'keyurl'}),
    frozenset({'cookie', 'cookies', 'cookiejson', '登录cookie',
               '登录cookies'}),
)
VENDOR_HEADER_TOKENS = frozenset().union(*VENDOR_HEADER_ALIASES)


class EnvBatchError(ValueError):
    pass


def normalize_env_site(value):
    site = str(value or 'MX').strip().upper()
    if site not in SUPPORTED_SITES:
        raise EnvBatchError('建环境站点仅支持 MX（墨西哥）或 US（美国）')
    return site


def normalize_buyer(value):
    """采购员名单校验：姓名或英文代号均可，统一规范化为姓名。"""
    token = str(value or '').strip()
    if token in BUYER_CODES:
        return token
    code = token.upper()
    if code in BUYER_NAME_BY_CODE:
        return BUYER_NAME_BY_CODE[code]
    raise EnvBatchError(
        '采购员不在名单内：%s（可用：%s）' % (
            scrub_text(token)[:20],
            ' / '.join('%s-%s' % (name, code)
                       for name, code in BUYER_ROSTER)))


# Cookie 登录域是账号站点的可靠标记（2026-08-20 台账实测：墨表含 Cookie 记录
# 全部为 shein.com.mx，美表全部为 us.shein.com，零交叉；无 Cookie 交付无标记）。
COOKIE_SITE_MARKERS = {'MX': 'shein.com.mx', 'US': 'us.shein.com'}


def detect_cookie_site(cookie_text):
    """从 Cookie 域名识别账号站点；无标记返回 None，双标记返回 'CONFLICT'。"""
    text = str(cookie_text or '')
    hits = {site for site, marker in COOKIE_SITE_MARKERS.items()
            if re.search(re.escape(marker), text, re.I)}
    if len(hits) > 1:
        return 'CONFLICT'
    return next(iter(hits)) if hits else None


def _format_row_ranges(row_numbers):
    numbers = sorted(set(int(number) for number in row_numbers))
    if not numbers:
        return ''
    ranges = []
    start = previous = numbers[0]
    for number in numbers[1:]:
        if number == previous + 1:
            previous = number
            continue
        ranges.append((start, previous))
        start = previous = number
    ranges.append((start, previous))
    return '、'.join(
        '第%d行' % start if start == end else '第%d-%d行' % (start, end)
        for start, end in ranges)


def count_mixed_site_accounts(accounts):
    """统计同时携带 MX/US 登录域 Cookie 的账号，不返回任何凭证内容。"""
    return sum(
        detect_cookie_site(account.cookie_text) == 'CONFLICT'
        for account in accounts)


def validate_accounts_site(accounts, site, *, allow_mixed=False):
    """校验账号 Cookie 站点；可显式兼容同时包含所选站点的混合登录态。"""
    site = normalize_env_site(site)
    conflicts = []
    mismatches = []
    for account in accounts:
        detected = detect_cookie_site(account.cookie_text)
        if detected == 'CONFLICT':
            conflicts.append(account.row_number)
        elif detected is not None and detected != site:
            mismatches.append((account.row_number, detected))
    rejected_conflicts = [] if allow_mixed else conflicts
    if not rejected_conflicts and not mismatches:
        return
    issue_rows = rejected_conflicts + [
        row_number for row_number, _detected in mismatches]
    details = []
    if rejected_conflicts:
        details.append(
            '墨西哥与美国登录域混合 %d 行' % len(rejected_conflicts))
    if mismatches:
        detected_counts = {}
        for _row_number, detected in mismatches:
            detected_counts[detected] = detected_counts.get(detected, 0) + 1
        details.append('、'.join(
            'Cookie仅%s站、与所选%s站不一致 %d 行' %
            (detected, site, count)
            for detected, count in sorted(detected_counts.items())))
    raise EnvBatchError(
        'Cookie 站点校验失败：共 %d 行数据异常（%s），'
        '行号 %s；整批拒收' % (
            len(issue_rows), '；'.join(details), _format_row_ranges(issue_rows)))


def normalize_backup_type(value):
    kind = str(value or '').strip()
    if kind not in BACKUP_TYPES:
        raise EnvBatchError('备用模式类型仅支持 备用 或 测试')
    return kind


def validate_backup_count(value):
    try:
        count = int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise EnvBatchError('备用/测试环境数量必须是整数') from exc
    if not 1 <= count <= BACKUP_MAX_COUNT:
        raise EnvBatchError(
            '备用/测试环境单批数量必须是 1-%d 的整数' % BACKUP_MAX_COUNT)
    return count


def validate_purchase_tag(value):
    """Validate the exact HubStudio group name without guessing a fallback."""
    raw_tag = str(value or '')
    if any(char in raw_tag for char in ('\r', '\n', '\t')):
        raise EnvBatchError('采购分组不能包含换行或制表符')
    tag = raw_tag.strip()
    if not tag:
        raise EnvBatchError('请先在设置中填写采购分组')
    if len(tag) > 12:
        raise EnvBatchError('采购分组不能超过 12 个字符')
    return tag


def validate_purchase_group_site(value, site):
    """Reject a group that explicitly names the opposite purchase site."""
    tag = validate_purchase_tag(value)
    site = normalize_env_site(site)
    mx_named = bool(re.search(
        r'墨西哥|(?<![A-Za-z0-9])MX(?![A-Za-z0-9])', tag, re.I))
    us_named = bool(re.search(
        r'美国|(?<![A-Za-z0-9])US(?![A-Za-z0-9])', tag, re.I))
    if site == 'US' and mx_named:
        raise EnvBatchError('美国站不能使用墨西哥采购分组，请重新选择')
    if site == 'MX' and us_named:
        raise EnvBatchError('墨西哥站不能使用美国采购分组，请重新选择')
    return tag


# 内置默认动态代理提取链接（Jeff 2026-08-20 决策：前期写死，降低同事端配置成本；
# 运行时可在设置页用自定义链接覆盖，清除配置即回落到本默认值）。
DEFAULT_PROXY_LINK = (
    'http://global.rotgbapi.711proxy.com:8089/gen?zone=custom&ptype=1'
    '&region={region}&count=1&proto=socks5&stype=text&split=\\r\\n'
    '&sessType=sticky&sessTime=60&sessAuto=1')


def validate_proxy_link(value):
    """Validate a secret proxy extraction URL without echoing it in errors."""
    from urllib.parse import urlsplit

    raw_link = str(value or '')
    if any(char.isspace() for char in raw_link):
        raise EnvBatchError('动态代理提取链接格式无效')
    link = raw_link.strip()
    if not link:
        raise EnvBatchError('请先在设置中填写动态代理提取链接')
    remainder = link.replace('{region}', '')
    if '{' in remainder or '}' in remainder:
        raise EnvBatchError('动态代理提取链接只允许 {region} 占位符')
    try:
        parsed = urlsplit(link.replace('{region}', 'MX'))
    except ValueError as exc:
        raise EnvBatchError('动态代理提取链接格式无效') from exc
    if parsed.scheme.lower() not in ('http', 'https') or not parsed.netloc:
        raise EnvBatchError('动态代理提取链接必须是 http(s) URL')
    return link


def envbatch_preflight(hub, purchase_tag, proxy_link, site='MX'):
    """Return a non-sensitive readiness summary for module three."""
    site = normalize_env_site(site)
    result = {
        'ready': False,
        'site': site,
        'hubConnected': False,
        'purchaseTag': '',
        'proxyConfigured': bool(str(proxy_link or '').strip()),
        'groupFound': False,
        'message': '',
    }
    problems = []
    tag = ''
    try:
        tag = validate_purchase_group_site(purchase_tag, site)
        result['purchaseTag'] = tag
    except EnvBatchError as exc:
        problems.append(str(exc))
    try:
        validate_proxy_link(proxy_link)
    except EnvBatchError as exc:
        problems.append(str(exc))
    try:
        groups = list(hub.group_list() or [])
    except Exception:
        problems.append('HubStudio 未连接，请启动客户端并检查 Local API 端口')
    else:
        result['hubConnected'] = True
        result['groupFound'] = bool(tag) and tag in groups
        if tag and not result['groupFound']:
            problems.append('采购分组未在 HubStudio 中精确匹配，请核对设置')
    result['ready'] = not problems
    result['message'] = '；'.join(problems) if problems else '执行前预检通过'
    return result


def require_envbatch_ready(hub, purchase_tag, proxy_link, site='MX'):
    result = envbatch_preflight(hub, purchase_tag, proxy_link, site=site)
    if not result['ready']:
        raise EnvBatchError(result['message'])
    return result


def project_root():
    return Path(__file__).resolve().parents[2]


def _inside(path, root):
    try:
        return os.path.commonpath([str(path), str(root)]) == str(root)
    except ValueError:
        return False


def _sample_dir():
    return (project_root() / 'examples').resolve()


def validate_input_path(path, allow_repo_sample=False):
    resolved = Path(path).expanduser().resolve()
    root = project_root().resolve()
    if _inside(resolved, root):
        if not (allow_repo_sample and _inside(resolved, _sample_dir())):
            raise EnvBatchError('真实凭证 xlsx 不得放在项目目录内')
    if not resolved.is_file():
        raise EnvBatchError('找不到输入 xlsx')
    return resolved


def validate_output_path(path):
    resolved = Path(path).expanduser().resolve()
    if _inside(resolved, project_root().resolve()):
        raise EnvBatchError('运行产物不得写入项目目录，请选择仓库外路径')
    return resolved


@dataclass
class BuyerAccount:
    row_number: int
    email: str
    password: str
    key_url: str
    cookie_text: str
    order_no: str
    buyer: str = ''

    @property
    def safe_email(self):
        return mask_email(self.email)

    @property
    def account_id(self):
        return hashlib.sha256(self.email.strip().casefold().encode('utf-8')).hexdigest()


@dataclass
class BatchPlanItem:
    account: BuyerAccount
    env_name: str
    container_code: str = ''
    serial_number: object = None
    completed_steps: set = field(default_factory=set)
    state: str = 'pending'
    error_step: str = ''
    error: str = ''
    binding_time: str = ''
    recovered_existing: bool = False

    def public_dict(self):
        return {
            'accountId': self.account.account_id,
            'emailMasked': self.account.safe_email,
            'buyer': self.account.buyer,
            'envName': self.env_name,
            'containerCode': self.container_code,
            'serialNumber': self.serial_number,
            'completedSteps': [x for x in STEPS if x in self.completed_steps],
            'state': self.state,
            'errorStep': self.error_step,
            'error': scrub_text(self.error)[:300],
            'bindingTime': self.binding_time,
            'recoveredExisting': self.recovered_existing,
        }


def _cell_text(value):
    return value if isinstance(value, str) else str(value)


def _header_token(value):
    return re.sub(r'[\s_-]+', '', str(value or '').strip().casefold())


def extract_vendor_order_no(key_url, account_email):
    """Extract a vendor order number or derive a stable opaque reference.

    Legacy code links expose a hexadecimal ``orderNo``.  Some vendors instead
    deliver links whose identity is either ``id+email``, an ``email``/``mail``
    query parameter, or the account email in the final URL path segment.
    Those links do not contain a purchase order number, so use a SHA-256
    reference while keeping the existing hexadecimal remark and cross-group
    deduplication contract.  Email-only formats deliberately share the same
    reference even if a vendor moves the email between query and path.
    """
    try:
        parsed = urlsplit(key_url)
        hostname = parsed.hostname
    except ValueError as exc:
        raise EnvBatchError('接码Key链接参数格式无效') from exc
    if (parsed.scheme.lower() not in ('http', 'https') or
            not hostname or parsed.username is not None or
            parsed.password is not None or parsed.fragment):
        raise EnvBatchError('接码Key链接不含可识别的号商单号')

    # Only inspect the parsed query.  This preserves legacy links without
    # allowing a look-alike orderNo in URL user-info, path, or fragment to
    # bypass the URL safety checks above.
    legacy = ORDER_RE.search('?' + parsed.query)
    if legacy:
        return legacy.group(1)

    values = None
    link_email = ''
    normalized_path = parsed.path.rstrip('/')
    if normalized_path in (EMAIL_KEY_PATH, MAIL_KEY_PATH):
        try:
            pairs = parse_qsl(
                parsed.query, keep_blank_values=True, strict_parsing=True)
        except ValueError as exc:
            raise EnvBatchError('接码Key链接参数格式无效') from exc
        values = {}
        for name, value in pairs:
            if name in values:
                raise EnvBatchError('接码Key链接参数重复')
            values[name] = value
        if normalized_path == EMAIL_KEY_PATH:
            if set(values) not in ({'email'}, {'id', 'email'}):
                raise EnvBatchError('接码Key链接参数不符合新号商格式')
            link_email = values['email'].strip()
        else:
            if (set(values) != {'type', 'mail'} or
                    values['type'].strip() != 'html'):
                raise EnvBatchError('接码Key链接参数不符合新号商格式')
            link_email = values['mail'].strip()
    elif not parsed.query:
        raw_segments = parsed.path.rstrip('/').split('/')
        raw_segments = [segment for segment in raw_segments if segment]
        if len(raw_segments) < 2:
            raise EnvBatchError('接码Key链接不含可识别的号商单号')
        raw_email = raw_segments[-1]
        if re.search(r'%(?![0-9A-Fa-f]{2})', raw_email):
            raise EnvBatchError('接码Key链接路径编码无效')
        try:
            link_email = unquote(
                raw_email, encoding='utf-8', errors='strict').strip()
        except (UnicodeDecodeError, ValueError) as exc:
            raise EnvBatchError('接码Key链接路径编码无效') from exc
    else:
        raise EnvBatchError('接码Key链接不含可识别的号商单号')

    if (not EMAIL_RE.fullmatch(link_email) or
            link_email.casefold() != account_email.strip().casefold()):
        raise EnvBatchError('接码Key链接邮箱与账号邮箱不一致')

    if values is not None and 'id' in values:
        vendor_id = values['id'].strip()
        if not re.fullmatch(r'[A-Za-z0-9_-]{6,128}', vendor_id):
            raise EnvBatchError('接码Key链接 id 格式无效')
        material = 'id\0%s\0%s' % (
            hostname.casefold(), vendor_id)
    else:
        material = 'email\0%s' % link_email.casefold()
    return hashlib.sha256(
        ('xynigo-vendor-ref-v1\0' + material).encode('utf-8')).hexdigest()


def _is_vendor_header(values, row_number):
    tokens = tuple(_header_token(value) for value in values[:4])
    if all(token in aliases for token, aliases in
           zip(tokens, VENDOR_HEADER_ALIASES)):
        return True
    recognized_count = sum(
        token in VENDOR_HEADER_TOKENS for token in tokens)
    if (tokens[0] in VENDOR_HEADER_TOKENS or recognized_count >= 2):
        raise EnvBatchError(
            '第 %d 行疑似表头但名称或列序不匹配，必须为：%s' %
            (row_number, ' / '.join(VENDOR_TEMPLATE_HEADERS)))
    return False


def parse_vendor_workbook(source):
    """解析第一张工作表固定四列，兼容有表头或无表头。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise EnvBatchError('缺少 openpyxl，无法读取 xlsx') from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise EnvBatchError('xlsx 无法打开：%s' % scrub_text(exc)) from exc
    try:
        sheet = workbook.worksheets[0]
        accounts = []
        seen_emails, seen_orders = set(), set()
        first_content_row = True
        for row_number, raw_row in enumerate(
                sheet.iter_rows(values_only=True), start=1):
            values = list(raw_row)
            if not any(value not in (None, '') for value in values):
                continue
            extras = values[4:]
            if any(value not in (None, '') for value in extras):
                raise EnvBatchError('第 %d 行第 5 列后存在非空数据，严格模式拒收' % row_number)
            values += [None] * max(0, 4 - len(values))
            email_raw, password_raw, key_raw, cookie_raw = values[:4]
            if first_content_row:
                first_content_row = False
                if _is_vendor_header(values[:4], row_number):
                    continue
            if any(value in (None, '') for value in
                   (email_raw, password_raw, key_raw, cookie_raw)):
                raise EnvBatchError('第 %d 行缺少邮箱/密码/接码Key/Cookie 字段' % row_number)

            email = _cell_text(email_raw).strip()
            if not EMAIL_RE.fullmatch(email):
                raise EnvBatchError('第 %d 行邮箱格式错误：%s' %
                                    (row_number, mask_email(email)))
            email_key = email.casefold()
            if email_key in seen_emails:
                raise EnvBatchError('第 %d 行邮箱重复：%s' %
                                    (row_number, mask_email(email)))
            seen_emails.add(email_key)

            key_url = _cell_text(key_raw).strip()
            if not re.match(r'^https?://', key_url, re.I):
                raise EnvBatchError('第 %d 行接码Key不是 http(s) URL：%s' %
                                    (row_number, mask_email(email)))
            try:
                order_no = extract_vendor_order_no(key_url, email)
            except EnvBatchError as exc:
                raise EnvBatchError('第 %d 行%s：%s' % (
                    row_number, str(exc), mask_email(email))) from exc
            if order_no.casefold() in seen_orders:
                raise EnvBatchError('第 %d 行号商单号重复：%s' %
                                    (row_number, mask_email(email)))
            seen_orders.add(order_no.casefold())

            cookie_text = _cell_text(cookie_raw)
            try:
                parsed_cookie = json.loads(cookie_text)
            except Exception as exc:
                raise EnvBatchError('第 %d 行 Cookie 不是合法 JSON：%s' %
                                    (row_number, mask_email(email))) from exc
            if not isinstance(parsed_cookie, (list, dict)):
                raise EnvBatchError('第 %d 行 Cookie JSON 必须是数组或对象：%s' %
                                    (row_number, mask_email(email)))

            accounts.append(BuyerAccount(
                row_number=row_number,
                email=email,
                password=_cell_text(password_raw),
                key_url=key_url,
                cookie_text=cookie_text,
                order_no=order_no,
            ))
        if not accounts:
            raise EnvBatchError('Sheet1 没有有效账号行')
        return accounts
    finally:
        workbook.close()


def load_vendor_xlsx(path, allow_repo_sample=False):
    resolved = validate_input_path(path, allow_repo_sample=allow_repo_sample)
    accounts = parse_vendor_workbook(str(resolved))
    return resolved, accounts


def parse_assignment(spec, total):
    if not isinstance(spec, str) or not spec.strip():
        raise EnvBatchError('采购员分配不能为空')
    result, seen = [], set()
    for chunk in re.split(r'[,，]', spec):
        chunk = chunk.strip()
        match = re.fullmatch(r'(\d+)\s*[:：]\s*([^\s:/\\]+)', chunk)
        if not match:
            raise EnvBatchError('采购员分配格式错误：%s' % scrub_text(chunk))
        count = int(match.group(1))
        buyer = normalize_buyer(match.group(2))
        if count < 1:
            raise EnvBatchError('采购员分配数量必须大于 0')
        if buyer in seen:
            raise EnvBatchError('采购员重复：%s' % buyer)
        seen.add(buyer)
        result.append((count, buyer))
    if sum(count for count, _buyer in result) != total:
        raise EnvBatchError('分配数量合计 %d 与账号数 %d 不一致' %
                            (sum(count for count, _buyer in result), total))
    return result


def assign_buyers(accounts, assignment_spec):
    assignments = parse_assignment(assignment_spec, len(accounts))
    offset = 0
    for count, buyer in assignments:
        for account in accounts[offset:offset + count]:
            account.buyer = buyer
        offset += count
    return assignments


def validate_assignment_template(spec):
    """设置页分配模板：校验格式与采购员名单，不校验合计。"""
    text = str(spec or '').strip()
    if not text:
        raise EnvBatchError('采购员分配模板不能为空')
    for chunk in re.split(r'[,，]', text):
        chunk = chunk.strip()
        match = re.fullmatch(r'(\d+)\s*[:：]\s*([^\s:/\\]+)', chunk)
        if not match or int(match.group(1)) < 1:
            raise EnvBatchError(
                '采购员分配模板格式错误：%s' % scrub_text(chunk))
        normalize_buyer(match.group(2))
    return text


def choose_resolution(rng=None):
    rng = rng or random
    choices = [(width, height) for width, height, _weight in RES_POOL]
    weights = [weight for _width, _height, weight in RES_POOL]
    return rng.choices(choices, weights=weights, k=1)[0]


def build_env_create_body(name, site='MX', rng=None, proxy_link=None,
                          purchase_tag=None):
    site = normalize_env_site(site)
    tag = validate_purchase_tag(purchase_tag)
    link_code = validate_proxy_link(proxy_link)
    width, height = choose_resolution(rng)
    return {
        'containerName': name,
        'tagName': tag,
        'asDynamicType': 0,
        'proxyTypeName': 'Socks5_通用api',
        'linkCode': link_code.replace('{region}', site),
        'ipGetRuleType': 1,
        'coreVersion': 148,
        'advancedBo': {
            'width': width,
            'height': height,
            'languageType': 0,
            'webgl': 0,
            'canvas': 0,
            'audioContext': 0,
        },
    }


def format_remark(account, purchase_date):
    return ('邮箱接码:%s | 单号:%s | 采购员:%s | 购买:%s' %
            (account.key_url, account.order_no, account.buyer,
             purchase_date))


def batch_fingerprint(source_bytes, assignment_spec, site, purchase_date):
    digest = hashlib.sha256()
    digest.update(source_bytes)
    digest.update(json.dumps({
        'assignment': assignment_spec,
        'site': site,
        'purchaseDate': purchase_date,
    }, sort_keys=True, ensure_ascii=False).encode('utf-8'))
    return digest.hexdigest()


def _existing_by_order(existing_envs, site, target_orders):
    result = {}
    for env in existing_envs:
        match = REMARK_ORDER_RE.search(str(env.get('remark') or ''))
        if not match:
            continue
        order_key = match.group(1).casefold()
        if order_key not in target_orders:
            continue
        env_name = str(env.get('containerName') or '')
        env_site = re.search(r'-(MX|US)-', env_name, re.I)
        if env_site and env_site.group(1).upper() != site:
            raise EnvBatchError(
                'HubStudio 中同一号商单号已绑定到另一站点环境，需人工处理')
        if order_key in result:
            raise EnvBatchError('HubStudio 中同一号商单号对应多个环境，需人工处理')
        result[order_key] = env
    return result


def _env_identity(env):
    """Return a stable, non-sensitive identity for cross-list membership."""
    code = str(env.get('containerCode') or '').strip()
    if code:
        return ('containerCode', code)
    serial = str(env.get('serialNumber') or '').strip()
    if serial:
        return ('serialNumber', serial)
    name = str(env.get('containerName') or '').strip()
    if name:
        return ('containerName', name)
    return None


def validate_global_order_dedup(accounts, selected_envs, all_envs):
    """Reject target order numbers that already exist outside target group.

    ``selected_envs`` remains the authority for same-group idempotent recovery;
    ``all_envs`` is an unfiltered, fully paginated HubStudio snapshot used only
    as a global duplicate guard.  Error text intentionally omits order, group,
    environment and account values.
    """
    target_orders = {account.order_no.casefold() for account in accounts}
    selected_ids = {
        identity for identity in (_env_identity(env) for env in selected_envs)
        if identity is not None
    }
    matches = {}
    # Include selected rows in case the unfiltered API snapshot is briefly
    # incomplete; identities collapse the normal duplicate representation.
    for env in list(all_envs or []) + list(selected_envs or []):
        match = REMARK_ORDER_RE.search(str(env.get('remark') or ''))
        if not match:
            continue
        order_key = match.group(1).casefold()
        if order_key not in target_orders:
            continue
        identity = _env_identity(env)
        if identity is None:
            raise EnvBatchError(
                'HubStudio 已有号商单号记录缺少环境标识，无法安全查重')
        matches.setdefault(order_key, {})[identity] = env

    for identities in matches.values():
        if any(identity not in selected_ids for identity in identities):
            raise EnvBatchError(
                'HubStudio 中同一号商单号已存在于其他分组，'
                '已阻止重复建环境，请人工核对')
        if len(identities) > 1:
            raise EnvBatchError(
                'HubStudio 中同一号商单号对应多个环境，需人工处理')


def build_batch_plan(accounts, assignment_spec, existing_envs=None,
                     site='MX', purchase_date=None, resume_state=None,
                     all_existing_envs=None):
    site = normalize_env_site(site)
    purchase_date = purchase_date or date.today().strftime('%Y%m%d')
    if not re.fullmatch(r'20\d{6}', purchase_date):
        raise EnvBatchError('购买日期必须是 YYYYMMDD')
    assignments = assign_buyers(accounts, assignment_spec)
    # 采购现阶段只能取得同时含 MX/US 登录域的 Cookie。环境创建按
    # 用户所选站点继续执行并原样写入 Cookie；纯错站数据仍整批拒收。
    validate_accounts_site(accounts, site, allow_mixed=True)
    existing_envs = list(existing_envs or [])
    if all_existing_envs is not None:
        validate_global_order_dedup(
            accounts, existing_envs, list(all_existing_envs))
    existing_orders = _existing_by_order(
        existing_envs, site,
        {account.order_no.casefold() for account in accounts})
    resume_rows = {
        row.get('accountId'): row
        for row in ((resume_state or {}).get('rows') or [])
        if row.get('accountId')
    }
    mmdd = purchase_date[-4:]
    max_suffix = {buyer: 0 for _count, buyer in assignments}
    for buyer in max_suffix:
        pattern = re.compile(ENV_NAME_RE.format(
            code=re.escape(BUYER_CODES[buyer]), site=re.escape(site),
            mmdd=mmdd))
        for env in existing_envs:
            match = pattern.fullmatch(str(env.get('containerName') or ''))
            if match:
                max_suffix[buyer] = max(max_suffix[buyer], int(match.group(1)))
        for row in resume_rows.values():
            match = pattern.fullmatch(str(row.get('envName') or ''))
            if match:
                max_suffix[buyer] = max(max_suffix[buyer], int(match.group(1)))

    plan = []
    for account in accounts:
        recovered = existing_orders.get(account.order_no.casefold())
        saved = resume_rows.get(account.account_id) or {}
        if recovered:
            env_name = str(recovered.get('containerName') or '')
            if not env_name:
                raise EnvBatchError('已存在环境缺少名称，无法安全恢复')
            item = BatchPlanItem(
                account=account,
                env_name=env_name,
                container_code=str(recovered.get('containerCode') or ''),
                serial_number=recovered.get('serialNumber'),
                completed_steps=set(STEPS),
                state='done',
                binding_time=str(saved.get('bindingTime') or ''),
                recovered_existing=True,
            )
        elif saved:
            item = BatchPlanItem(
                account=account,
                env_name=str(saved.get('envName') or ''),
                container_code=str(saved.get('containerCode') or ''),
                serial_number=saved.get('serialNumber'),
                completed_steps=set(saved.get('completedSteps') or []),
                state=str(saved.get('state') or 'pending'),
                error_step=str(saved.get('errorStep') or ''),
                error=str(saved.get('error') or ''),
                binding_time=str(saved.get('bindingTime') or ''),
            )
            if not item.env_name:
                raise EnvBatchError('续跑状态缺少环境名，拒绝恢复')
        else:
            max_suffix[account.buyer] += 1
            env_name = '%s-%s-%s-%03d' % (
                BUYER_CODES[account.buyer], site, mmdd,
                max_suffix[account.buyer])
            item = BatchPlanItem(account=account, env_name=env_name)
        plan.append(item)
    return plan


def default_resume_dir():
    if os.name == 'nt':
        base = os.environ.get('LOCALAPPDATA') or tempfile.gettempdir()
        return Path(base) / 'PurchaseTool' / 'resume'
    if sys.platform == 'darwin':
        return Path.home() / 'Library' / 'Application Support' / '采购工具' / 'resume'
    return Path.home() / '.local' / 'state' / 'purchase-tool' / 'resume'


def _write_private_json(path, payload):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix='.envbatch-', suffix='.tmp',
                                    dir=str(path.parent))
    try:
        try:
            os.fchmod(fd, 0o600)
        except (AttributeError, OSError):
            pass
        with os.fdopen(fd, 'w', encoding='utf-8') as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_name, path)
        try:
            os.chmod(path, 0o600)
        except OSError:
            pass
    except Exception:
        try:
            os.close(fd)
        except OSError:
            pass
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


class ResumeStateStore(object):
    def __init__(self, batch_id, state_dir=None):
        self.batch_id = batch_id
        self.state_dir = Path(state_dir or default_resume_dir()).expanduser().resolve()
        if _inside(self.state_dir, project_root().resolve()):
            raise EnvBatchError('续跑目录必须位于项目目录外')
        self.path = self.state_dir / ('envbatch-%s.json' % batch_id[:20])

    def load(self):
        if not self.path.exists():
            return None
        try:
            with self.path.open(encoding='utf-8') as handle:
                payload = json.load(handle)
        except Exception as exc:
            raise EnvBatchError('续跑状态文件损坏，拒绝猜测恢复') from exc
        if payload.get('batchId') != self.batch_id:
            raise EnvBatchError('续跑状态与当前 xlsx/分配不匹配')
        return payload

    def save(self, rows, site, purchase_date):
        payload = {
            'version': 1,
            'batchId': self.batch_id,
            'site': site,
            'purchaseDate': purchase_date,
            'updatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'rows': [row.public_dict() for row in rows],
        }
        raw = json.dumps(payload, ensure_ascii=False).casefold()
        for forbidden in ('password', 'cookietext', 'keyurl'):
            if forbidden in raw:
                raise EnvBatchError('续跑状态意外包含凭证字段，已拒绝写入')
        _write_private_json(self.path, payload)

    def remove(self):
        try:
            self.path.unlink()
        except FileNotFoundError:
            pass


class BatchEnvOrchestrator(object):
    """HubStudio 写链路串行执行；每步落脱敏状态后再进入下一步。"""

    def __init__(self, hub, purchase_tag, proxy_link, site='MX',
                 purchase_date=None, state_store=None,
                 write_interval=0.3, sleep_fn=time.sleep, rng=None,
                 on_progress=None, max_workers=5, stop_event=None):
        self.hub = hub
        self.site = normalize_env_site(site)
        self.purchase_tag = validate_purchase_tag(purchase_tag)
        self.proxy_link = validate_proxy_link(proxy_link)
        self.purchase_date = purchase_date or date.today().strftime('%Y%m%d')
        self.state_store = state_store
        self.write_interval = write_interval
        self.sleep = sleep_fn
        self.rng = rng or random
        self.on_progress = on_progress
        self.stop_event = stop_event or threading.Event()
        # 并行建环境：纯 Local API 路径（不开浏览器窗口），默认 5、
        # 1-10 可调；模块一开窗口场景的并发结论不适用于此处。
        self.max_workers = max(1, min(10, int(max_workers)))
        self._persist_lock = threading.Lock()
        self.rows = []

    def prepare(self, accounts, assignment_spec):
        existing = self.hub.env_list(self.purchase_tag)
        all_existing = self.hub.env_list()
        saved = self.state_store.load() if self.state_store else None
        self.rows = build_batch_plan(
            accounts, assignment_spec, existing_envs=existing,
            site=self.site, purchase_date=self.purchase_date,
            resume_state=saved, all_existing_envs=all_existing)
        self._persist()
        return self.rows

    def _persist(self):
        with self._persist_lock:
            if self.state_store:
                self.state_store.save(self.rows, self.site,
                                      self.purchase_date)
            if self.on_progress:
                self.on_progress([row.public_dict() for row in self.rows])

    def _mark(self, row, step):
        row.completed_steps.add(step)
        row.error = ''
        row.error_step = ''
        row.state = 'done' if step == 'done' else step
        self._persist()

    def _find_env(self, env_name, attempts=25):
        for attempt in range(attempts):
            for env in self.hub.env_list(self.purchase_tag):
                if env.get('containerName') == env_name:
                    return env
            if attempt + 1 < attempts:
                self.sleep(1)
        raise EnvBatchError('新建环境后回读超时')

    def _run_one(self, row):
        if 'done' in row.completed_steps:
            row.state = 'done'
            return
        # 安全停止只在一整行开始前生效。已经进入五步写链路的行必须完整
        # 收尾，避免留下“环境已建但未绑号/未写备注”的半成品。
        if self.stop_event.is_set():
            row.state = 'stopped'
            row.error_step = ''
            row.error = '安全停止：未开始执行'
            self._persist()
            return
        current_step = 'env_created'
        try:
            row.state = 'running'
            self._persist()
            if 'env_created' not in row.completed_steps:
                existing = None
                for env in self.hub.env_list(self.purchase_tag):
                    if env.get('containerName') == row.env_name:
                        existing = env
                        break
                if existing is None:
                    self.hub.env_create(build_env_create_body(
                        row.env_name, self.site, self.rng,
                        proxy_link=self.proxy_link,
                        purchase_tag=self.purchase_tag))
                    self.sleep(self.write_interval)
                    existing = self._find_env(row.env_name)
                elif str(existing.get('remark') or '').strip():
                    # 收养防护：同名但备注非空 = 其他批次/机器的成品
                    # （备用/测试/已绑号），拒绝收养防错绑与备注覆盖；
                    # 合法收养只有本批次断点续跑（备注必为空）。
                    raise EnvBatchError(
                        '同名环境 %s 已存在且备注非空，可能被其他批次或'
                        '机器占用，拒绝收养，请人工核对' % row.env_name)
                row.container_code = str(existing.get('containerCode') or '')
                row.serial_number = existing.get('serialNumber')
                if not row.container_code or row.serial_number is None:
                    raise EnvBatchError('环境回读缺少 containerCode 或 HUB 序号')
                self._mark(row, 'env_created')

            current_step = 'cookie_imported'
            if 'cookie_imported' not in row.completed_steps:
                self.hub.env_import_cookie(
                    row.container_code, row.account.cookie_text)
                self.sleep(self.write_interval)
                self._mark(row, 'cookie_imported')

            current_step = 'account_bound'
            if 'account_bound' not in row.completed_steps:
                self.hub.container_add_account(
                    row.container_code, row.account.email,
                    row.account.password, self.site)
                self.sleep(self.write_interval)
                self._mark(row, 'account_bound')

            current_step = 'remarked'
            if 'remarked' not in row.completed_steps:
                self.hub.env_update(
                    row.container_code, row.env_name,
                    format_remark(row.account, self.purchase_date))
                self.sleep(self.write_interval)
                self._mark(row, 'remarked')

            current_step = 'done'
            row.binding_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self._mark(row, 'done')
        except Exception as exc:
            row.state = 'failed'
            row.error_step = current_step
            error_text = str(exc)
            for secret in (row.account.email, row.account.password,
                           row.account.key_url, row.account.cookie_text,
                           self.proxy_link,
                           self.proxy_link.replace('{region}', self.site)):
                if secret:
                    error_text = error_text.replace(secret, '<redacted>')
            row.error = scrub_text(error_text)[:300]
            self._persist()

    def run(self):
        if self.max_workers > 1 and len(self.rows) > 1:
            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                list(pool.map(self._run_one, self.rows))
        else:
            for row in self.rows:
                self._run_one(row)
        if self.state_store and all(row.state == 'done' for row in self.rows):
            self.state_store.remove()
        return self.rows

    def retry_one(self, account_id):
        row = next((item for item in self.rows
                    if item.account.account_id == account_id), None)
        if row is None:
            raise EnvBatchError('找不到待重试账号')
        if row.state == 'done':
            return row
        self._run_one(row)
        if self.state_store and all(item.state == 'done' for item in self.rows):
            self.state_store.remove()
        return row

    def retry_failed(self):
        """Retry every failed row from its persisted first incomplete step."""
        failed_rows = [row for row in self.rows if row.state == 'failed']
        if not failed_rows:
            raise EnvBatchError('当前批次没有失败行可重试')
        if self.max_workers > 1 and len(failed_rows) > 1:
            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                list(pool.map(self._run_one, failed_rows))
        else:
            for row in failed_rows:
                self._run_one(row)
        if self.state_store and all(item.state == 'done' for item in self.rows):
            self.state_store.remove()
        return failed_rows

    @staticmethod
    def _verification_sample(rows, count):
        done = [row for row in rows if row.state == 'done' and row.container_code]
        if count <= 0 or not done:
            return []
        chosen = []
        for buyer in dict.fromkeys(row.account.buyer for row in done):
            candidate = next(row for row in done if row.account.buyer == buyer)
            if candidate not in chosen:
                chosen.append(candidate)
            if len(chosen) >= count:
                return chosen
        for index in (0, len(done) // 2, len(done) - 1):
            candidate = done[index]
            if candidate not in chosen:
                chosen.append(candidate)
            if len(chosen) >= count:
                break
        return chosen[:count]

    def verify_ips(self, count=3, geo_lookup=None):
        geo_lookup = geo_lookup or lookup_ip_country
        results = []
        for row in self._verification_sample(self.rows, count):
            results.append(probe_env_ip(
                self.hub, self.site, row.env_name, row.container_code,
                geo_lookup))
        return results


def probe_env_ip(hub, site, env_name, container_code, geo_lookup):
    """无头启动环境读取出口 IP 并核对国家，随后立即关闭。"""
    started = False
    try:
        data = hub.browser_start(container_code, headless=True) or {}
        started = True
        ip = str(data.get('ip') or '')
        geo = geo_lookup(ip) if ip else {}
        country_code = str(geo.get('countryCode') or '').upper()
        return {
            'envName': env_name,
            'ip': ip,
            'country': geo.get('country') or '',
            'city': geo.get('city') or '',
            'isp': geo.get('isp') or '',
            'ok': country_code == site,
            'error': '' if country_code == site else '出口 IP 国家不匹配',
        }
    except Exception as exc:
        return {
            'envName': env_name, 'ip': '', 'country': '',
            'city': '', 'isp': '', 'ok': False,
            'error': scrub_text(exc)[:200],
        }
    finally:
        if started:
            try:
                hub.browser_stop(container_code)
            except Exception:
                pass


def backup_env_names(existing_envs, buyer, count, backup_type, site,
                     purchase_date):
    """备用/测试环境命名与续排（只读、按名幂等）。

    备用环境与绑号环境共享该采购员当日代号序号段并续排错开；
    测试环境走 `{code}-{site}-测试-{NN}` 独立命名空间，序号上限 99。
    """
    site = normalize_env_site(site)
    buyer = normalize_buyer(buyer)
    count = validate_backup_count(count)
    backup_type = normalize_backup_type(backup_type)
    purchase_date = str(purchase_date or '').strip()
    if not re.fullmatch(r'20\d{6}', purchase_date):
        raise EnvBatchError('购买日期必须是 YYYYMMDD')
    code = BUYER_CODES[buyer]
    existing = list(existing_envs or [])
    taken = {str(env.get('containerName') or '') for env in existing}
    if backup_type == '备用':
        pattern = re.compile(ENV_NAME_RE.format(
            code=re.escape(code), site=re.escape(site),
            mmdd=purchase_date[-4:]))

        def make(serial):
            return '%s-%s-%s-%03d' % (code, site, purchase_date[-4:], serial)
    else:
        pattern = re.compile(TEST_ENV_NAME_RE.format(
            code=re.escape(code), site=re.escape(site)))

        def make(serial):
            return '%s-%s-测试-%02d' % (code, site, serial)

    serial = 0
    for env in existing:
        match = pattern.fullmatch(str(env.get('containerName') or ''))
        if match:
            serial = max(serial, int(match.group(1)))
    names = []
    while len(names) < count:
        serial += 1
        if backup_type == '测试' and serial > 99:
            raise EnvBatchError('测试环境序号已达 99 上限，需人工处理')
        name = make(serial)
        if name in taken:
            continue
        names.append(name)
    return names


@dataclass
class BackupPlanItem:
    env_name: str
    container_code: str = ''
    serial_number: object = None
    state: str = 'pending'
    error: str = ''

    def public_dict(self):
        return {
            'envName': self.env_name,
            'containerCode': self.container_code,
            'serialNumber': self.serial_number,
            'state': self.state,
            'error': scrub_text(self.error)[:300],
        }


class BackupEnvOrchestrator(object):
    """备用/测试环境：只建环境+写固定备注。

    不导 Cookie、不绑号、不写飞书台账；无凭证流转。
    """

    def __init__(self, hub, purchase_tag, proxy_link, site='MX',
                 write_interval=0.3, sleep_fn=time.sleep, rng=None,
                 on_progress=None, max_workers=5, stop_event=None):
        self.hub = hub
        self.site = normalize_env_site(site)
        self.purchase_tag = validate_purchase_tag(purchase_tag)
        self.proxy_link = validate_proxy_link(proxy_link)
        self.write_interval = write_interval
        self.sleep = sleep_fn
        self.rng = rng or random
        self.on_progress = on_progress
        self.stop_event = stop_event or threading.Event()
        self.max_workers = max(1, min(10, int(max_workers)))
        self._persist_lock = threading.Lock()
        self.rows = []
        self.remark = ''

    def prepare(self, buyer, count, backup_type, purchase_date):
        backup_type = normalize_backup_type(backup_type)
        self.remark = BACKUP_REMARK[backup_type]
        existing = self.hub.env_list(self.purchase_tag)
        names = backup_env_names(
            existing, buyer, count, backup_type, self.site, purchase_date)
        self.rows = [BackupPlanItem(env_name=name) for name in names]
        self._persist()
        return self.rows

    def _persist(self):
        with self._persist_lock:
            if self.on_progress:
                self.on_progress([row.public_dict() for row in self.rows])

    def _find_env(self, env_name, attempts=25):
        for attempt in range(attempts):
            for env in self.hub.env_list(self.purchase_tag):
                if env.get('containerName') == env_name:
                    return env
            if attempt + 1 < attempts:
                self.sleep(1)
        raise EnvBatchError('新建环境后回读超时')

    def _run_one(self, row):
        if row.state == 'done':
            return
        # 与绑号环境一致：只阻止尚未开始的整行，不在建环境和写备注之间
        # 强切，确保不会制造缺备注的半成品环境。
        if self.stop_event.is_set():
            row.state = 'stopped'
            row.error = '安全停止：未开始执行'
            self._persist()
            return
        row.state = 'running'
        self._persist()
        try:
            existing = None
            for env in self.hub.env_list(self.purchase_tag):
                if env.get('containerName') == row.env_name:
                    existing = env
                    break
            if existing is None:
                self.hub.env_create(build_env_create_body(
                    row.env_name, self.site, self.rng,
                    proxy_link=self.proxy_link,
                    purchase_tag=self.purchase_tag))
                self.sleep(self.write_interval)
                existing = self._find_env(row.env_name)
            row.container_code = str(existing.get('containerCode') or '')
            row.serial_number = existing.get('serialNumber')
            if not row.container_code or row.serial_number is None:
                raise EnvBatchError('环境回读缺少 containerCode 或 HUB 序号')
            self.hub.env_update(
                row.container_code, row.env_name, self.remark)
            self.sleep(self.write_interval)
            row.state = 'done'
        except Exception as exc:
            row.state = 'failed'
            error_text = str(exc)
            for secret in (self.proxy_link,
                           self.proxy_link.replace('{region}', self.site)):
                if secret:
                    error_text = error_text.replace(secret, '<redacted>')
            row.error = scrub_text(error_text)[:300]
        self._persist()

    def run(self):
        if self.max_workers > 1 and len(self.rows) > 1:
            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                list(pool.map(self._run_one, self.rows))
        else:
            for row in self.rows:
                self._run_one(row)
        return self.rows

    def verify_ips(self, count=1, geo_lookup=None):
        geo_lookup = geo_lookup or lookup_ip_country
        done = [row for row in self.rows
                if row.state == 'done' and row.container_code]
        results = []
        for row in done[:max(0, int(count))]:
            results.append(probe_env_ip(
                self.hub, self.site, row.env_name, row.container_code,
                geo_lookup))
        return results


def backup_result_tsv_bytes(rows, site):
    """备用/测试结果清单：只含环境名/HUB 序号/containerCode，无凭证。"""
    output = StringIO(newline='')
    writer = csv.writer(output, dialect='excel-tab', lineterminator='\r\n')
    writer.writerow(list(BACKUP_RESULT_HEADERS))
    for row in rows:
        writer.writerow([
            row.env_name,
            row.serial_number if row.serial_number is not None else '',
            row.container_code,
            site,
            ('完成' if row.state == 'done' else
             ('已停止' if row.state == 'stopped' else '失败')),
        ])
    return output.getvalue().encode('utf-8-sig')


def lookup_ip_country(ip):
    if not re.fullmatch(r'[0-9a-fA-F:.]+', str(ip or '')):
        raise EnvBatchError('HubStudio 未返回有效出口 IP')
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    url = ('http://ip-api.com/json/%s?fields=status,message,country,'
           'countryCode,city,isp' % ip)
    with opener.open(url, timeout=10) as response:
        data = json.loads(response.read().decode('utf-8'))
    if data.get('status') != 'success':
        raise EnvBatchError('IP 归属地查询失败')
    return data


def mapping_workbook_bytes(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise EnvBatchError('缺少 openpyxl，无法导出映射清单') from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '绑定映射清单'
    sheet.append(list(MAPPING_HEADERS))
    for row in rows:
        sheet.append([
            row.account.email,
            row.env_name,
            row.serial_number,
            row.account.buyer,
            row.binding_time,
            ('完成' if row.state == 'done' else
             ('已停止' if row.state == 'stopped'
              else '失败:%s' % row.error_step)),
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    widths = (30, 30, 12, 12, 22, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = 'A1:F%d' % max(1, len(rows) + 1)
    sheet.sheet_view.showGridLines = False
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def write_private_bytes(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd = os.open(str(path), os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o600)
    with os.fdopen(fd, 'wb') as handle:
        handle.write(data)
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def build_parser():
    parser = argparse.ArgumentParser(
        prog='python -m purchase_tool env-batch',
        description='模块三：号商买家号批量创建 HubStudio 采购环境')
    parser.add_argument('--xlsx', '--input', dest='xlsx', required=True,
                        help='号商固定 4 列 xlsx，可有表头或无表头（真实文件必须在项目外）')
    parser.add_argument('--assign', required=True,
                        help='按文件顺序分配，如 3:Operator-A,2:Operator-B')
    parser.add_argument('--site', default='MX', choices=['MX', 'US'])
    parser.add_argument('--purchase-date',
                        default=date.today().strftime('%Y%m%d'))
    parser.add_argument('--hub-port', type=int, default=DEFAULT_PORT)
    parser.add_argument('--verify-sample-count', type=int, default=3)
    parser.add_argument('--mapping-output', help='仓库外的绑定映射清单路径')
    parser.add_argument('--resume-dir', help='仓库外的脱敏续跑状态目录')
    parser.add_argument('--apply', action='store_true',
                        help='真实创建环境；不加时仅离线 dry-run')
    parser.add_argument('--confirm-env-write', action='store_true',
                        help='确认执行 HubStudio 创建/导Cookie/绑号/备注写入')
    parser.add_argument('--no-verify-ip', action='store_true',
                        help='跳过批末出口 IP 抽查')
    return parser


def _print_dry_run(accounts, plan, assignments):
    print('计划校验通过：%d 个账号；Cookie 覆盖 %d/%d' %
          (len(accounts), len(accounts), len(accounts)))
    print('采购员分配：%s' % ', '.join(
        '%s×%d' % (buyer, count) for count, buyer in assignments))
    for row in plan[:5]:
        print('  - %s -> %s' % (row.account.safe_email, row.env_name))
    if len(plan) > 5:
        print('  ... 其余 %d 行已省略' % (len(plan) - 5))
    print('dry-run 完成；未连接 HubStudio，未写入任何数据。')


def main(argv=None):
    args = build_parser().parse_args(argv)
    try:
        input_path = validate_input_path(
            args.xlsx, allow_repo_sample=not args.apply)
        with input_path.open('rb') as handle:
            source_bytes = handle.read()
        accounts = parse_vendor_workbook(BytesIO(source_bytes))
        assignments = parse_assignment(args.assign, len(accounts))
        if not args.apply:
            plan = build_batch_plan(
                accounts, args.assign, existing_envs=[], site=args.site,
                purchase_date=args.purchase_date)
            _print_dry_run(accounts, plan, assignments)
            return 0
        if not args.confirm_env_write:
            print('拒绝执行：真实写入必须同时传 --confirm-env-write。')
            return 2

        output = (Path(args.mapping_output).expanduser() if args.mapping_output
                  else input_path.with_name(
                      '绑定映射清单_%s.xlsx' % args.purchase_date))
        output = validate_output_path(output)
        batch_id = batch_fingerprint(
            source_bytes, args.assign, args.site, args.purchase_date)
        state_store = ResumeStateStore(batch_id, args.resume_dir)
        hub = HubStudioApi(port=args.hub_port)
        purchase_tag = os.environ.get(
            'XYNIGO_PURCHASE_TAG_%s' % args.site,
            os.environ.get('XYNIGO_PURCHASE_TAG', '')
            if args.site == 'MX' else '')
        proxy_link = os.environ.get('XYNIGO_PROXY_LINK', '') or DEFAULT_PROXY_LINK
        require_envbatch_ready(
            hub, purchase_tag, proxy_link, site=args.site)
        runner = BatchEnvOrchestrator(
            hub, purchase_tag=purchase_tag, proxy_link=proxy_link,
            site=args.site, purchase_date=args.purchase_date,
            state_store=state_store)
        runner.prepare(accounts, args.assign)
        rows = runner.run()
        write_private_bytes(output, mapping_workbook_bytes(rows))
        done = sum(row.state == 'done' for row in rows)
        failed = len(rows) - done
        print('主链路结束：完成 %d，失败 %d；映射清单：%s' %
              (done, failed, output))
        if not args.no_verify_ip and done:
            checks = runner.verify_ips(max(0, args.verify_sample_count))
            ok = sum(item.get('ok') for item in checks)
            print('出口 IP 抽查：%d/%d 国家匹配' % (ok, len(checks)))
            if ok != len(checks):
                failed += 1
        return 2 if failed else 0
    except EnvBatchError as exc:
        print('执行失败：%s' % scrub_text(exc))
        return 2
    except Exception as exc:
        print('执行失败：%s' % scrub_text(exc))
        return 2


if __name__ == '__main__':
    sys.exit(main())
