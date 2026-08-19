# -*- coding: utf-8 -*-
"""模块三 Mac 补账命令：原始 xlsx + 绑定映射清单 → 飞书台账。

默认只做离线 dry-run；``--apply`` 才会调用 lark-cli 写入。所有含凭证
payload 均通过权限 0600 的临时文件传递，不进入 argv 或日志。
"""
import argparse
from dataclasses import dataclass
from io import BytesIO
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import time

from .env_batch import (EMAIL_RE, EnvBatchError, MAPPING_HEADERS,
                        load_vendor_xlsx, parse_vendor_workbook,
                        project_root, validate_input_path)
from .lark_ledger import BASE_TOKEN, TABLE_MX
from .redaction import mask_email, scrub_text


DEFAULT_OPERATOR_OPEN_ID = os.environ.get(
    'XYNIGO_LARK_OPERATOR_OPEN_ID', '')
BATCH_SIZE = 200
RETRYABLE_CODES = {1254291}
CREATE_FIELDS = (
    '邮箱账号', '密码', '接码Key链接', '号商购买单号', '账号状态',
    '采购员', 'Cookie', '备注',
)
READ_FIELDS = (
    '邮箱账号', '账号状态', '绑定环境', '环境序号', '绑定时间',
    '采购员', '操作人',
)
REQUIRED_FIELD_TYPES = {
    '邮箱账号': 'text',
    '密码': 'text',
    '接码Key链接': 'text',
    '号商购买单号': 'text',
    '账号状态': 'select',
    '采购员': 'select',
    'Cookie': 'text',
    '备注': 'text',
    '绑定环境': 'text',
    '环境序号': 'number',
    '绑定时间': 'datetime',
    '操作人': 'user',
}


class LedgerBackfillError(Exception):
    def __init__(self, message, code=None):
        super().__init__(message)
        self.code = code


@dataclass
class MappingRow:
    row_number: int
    email: str
    env_name: str
    serial_number: object
    buyer: str
    binding_time: str
    status: str

    @property
    def complete(self):
        return self.status == '完成'


def parse_mapping_workbook(source):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise LedgerBackfillError('缺少 openpyxl，无法读取映射清单') from exc
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise LedgerBackfillError('映射清单无法打开') from exc
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise LedgerBackfillError('映射清单为空')
        actual_headers = tuple(str(x or '').strip() for x in rows[0][:6])
        if actual_headers != MAPPING_HEADERS:
            raise LedgerBackfillError(
                '映射清单表头不匹配，必须为：%s' % '/'.join(MAPPING_HEADERS))
        result, seen = [], set()
        for row_number, raw in enumerate(rows[1:], start=2):
            values = list(raw) + [None] * max(0, 6 - len(raw))
            if not any(value not in (None, '') for value in values):
                continue
            if any(value not in (None, '') for value in values[6:]):
                raise LedgerBackfillError('映射清单第 %d 行存在额外非空列' % row_number)
            email = str(values[0] or '').strip()
            if not EMAIL_RE.fullmatch(email):
                raise LedgerBackfillError('映射清单第 %d 行邮箱错误：%s' %
                                          (row_number, mask_email(email)))
            email_key = email.casefold()
            if email_key in seen:
                raise LedgerBackfillError('映射清单邮箱重复：%s' %
                                          mask_email(email))
            seen.add(email_key)
            status = str(values[5] or '').strip()
            complete = status == '完成'
            env_name = str(values[1] or '').strip()
            buyer = str(values[3] or '').strip()
            binding_time = str(values[4] or '').strip()
            serial = values[2]
            if complete:
                if not env_name or not buyer or serial in (None, '') or not binding_time:
                    raise LedgerBackfillError(
                        '映射清单第 %d 行完成状态缺少绑定字段' % row_number)
                try:
                    serial = int(serial)
                except (TypeError, ValueError) as exc:
                    raise LedgerBackfillError(
                        '映射清单第 %d 行 HUB 序号不是整数' % row_number) from exc
            result.append(MappingRow(
                row_number=row_number, email=email, env_name=env_name,
                serial_number=serial, buyer=buyer,
                binding_time=binding_time, status=status))
        if not result:
            raise LedgerBackfillError('映射清单没有数据行')
        return result
    finally:
        workbook.close()


def load_mapping_xlsx(path, allow_repo_sample=False):
    try:
        resolved = validate_input_path(path, allow_repo_sample=allow_repo_sample)
    except EnvBatchError as exc:
        raise LedgerBackfillError(str(exc)) from exc
    return resolved, parse_mapping_workbook(str(resolved))


def _safe_lark_env():
    return dict(os.environ,
                LARKSUITE_CLI_NO_UPDATE_NOTIFIER='1',
                LARKSUITE_CLI_NO_SKILLS_NOTIFIER='1')


class LarkBaseClient(object):
    def __init__(self, lark_bin='lark-cli', base_token=BASE_TOKEN,
                 table_id=TABLE_MX, profile=None, runner=subprocess.run,
                 sleep_fn=time.sleep):
        self.lark_bin = lark_bin
        self.base_token = base_token
        self.table_id = table_id
        self.profile = profile
        self.runner = runner
        self.sleep = sleep_fn

    def _argv(self, command):
        if not self.base_token or not self.table_id:
            raise LedgerBackfillError(
                '未配置飞书台账；请设置 XYNIGO_LARK_BASE_TOKEN 和 '
                'XYNIGO_LARK_TABLE_ID')
        argv = [self.lark_bin, 'base', command,
                '--base-token', self.base_token,
                '--table-id', self.table_id,
                '--as', 'user', '--format', 'json']
        if self.profile:
            argv += ['--profile', self.profile]
        return argv

    @staticmethod
    def _parse(proc, command):
        raw = (proc.stdout or proc.stderr or '').strip()
        try:
            result = json.loads(raw)
        except Exception as exc:
            raise LedgerBackfillError('%s 返回了非 JSON 结果' % command) from exc
        if proc.returncode != 0 or result.get('ok') is not True:
            error = result.get('error') or {}
            message = error.get('message') or error.get('subtype') or '未知错误'
            code = (result.get('code') or error.get('code')
                    or (result.get('data') or {}).get('code'))
            if isinstance(code, str) and code.isdigit():
                code = int(code)
            raise LedgerBackfillError('%s 失败：%s' %
                                      (command, scrub_text(message)[:200]),
                                      code=code)
        return result

    def _run(self, command, extra=None, payload=None, high_risk=False):
        argv = self._argv(command) + list(extra or [])
        if high_risk:
            argv.append('--yes')
        if payload is None:
            proc = self.runner(argv, capture_output=True, text=True,
                               env=_safe_lark_env())
            return self._parse(proc, command)
        with tempfile.TemporaryDirectory(prefix='purchase-backfill-') as tmp:
            path = Path(tmp) / 'payload.json'
            fd = os.open(str(path), os.O_WRONLY | os.O_CREAT | os.O_TRUNC,
                         0o600)
            with os.fdopen(fd, 'w', encoding='utf-8') as handle:
                json.dump(payload, handle, ensure_ascii=False)
            proc = self.runner(
                argv + ['--json', '@payload.json'], cwd=tmp,
                capture_output=True, text=True, env=_safe_lark_env())
            return self._parse(proc, command)

    def list_fields(self):
        fields, offset = [], 0
        while True:
            result = self._run('+field-list', [
                '--limit', str(BATCH_SIZE), '--offset', str(offset)])
            data = result.get('data') or {}
            page = data.get('fields') or []
            fields.extend(page)
            if not data.get('has_more'):
                return fields
            if not page:
                raise LedgerBackfillError('field-list 分页停滞')
            offset += len(page)

    def list_records(self, fields=READ_FIELDS):
        records, offset = [], 0
        while True:
            extra = ['--limit', str(BATCH_SIZE), '--offset', str(offset)]
            for field_name in fields:
                extra += ['--field-id', field_name]
            result = self._run('+record-list', extra)
            data = result.get('data') or {}
            names = data.get('fields') or []
            values = data.get('data') or []
            record_ids = data.get('record_id_list') or []
            if len(record_ids) != len(values):
                raise LedgerBackfillError('record-list 的记录 ID 与数据行数量不一致')
            for record_id, row in zip(record_ids, values):
                record = dict(zip(names, row))
                record['_record_id'] = record_id
                records.append(record)
            if not data.get('has_more'):
                return records
            if not values:
                raise LedgerBackfillError('record-list 分页停滞')
            offset += len(values)

    def batch_create(self, create_records):
        record_ids = []
        for offset in range(0, len(create_records), BATCH_SIZE):
            batch = create_records[offset:offset + BATCH_SIZE]
            payload = {'create_records': batch}
            for attempt in range(5):
                try:
                    result = self._run('+record-batch-create', payload=payload)
                    break
                except LedgerBackfillError as exc:
                    code = getattr(exc, 'code', None)
                    if code not in RETRYABLE_CODES or attempt >= 4:
                        raise
                    self.sleep(2 ** attempt)
            ids = (result.get('data') or {}).get('record_id_list') or []
            if len(ids) != len(batch):
                raise LedgerBackfillError('batch-create 返回 record_id 数量不符')
            record_ids.extend(ids)
        return record_ids

    def batch_update(self, update_records):
        items = list(update_records.items())
        for offset in range(0, len(items), BATCH_SIZE):
            payload = {'update_records': dict(items[offset:offset + BATCH_SIZE])}
            for attempt in range(5):
                try:
                    self._run('+record-batch-update', payload=payload)
                    break
                except LedgerBackfillError as exc:
                    if exc.code not in RETRYABLE_CODES or attempt >= 4:
                        raise
                    self.sleep(2 ** attempt)

    def update_select_options(self, field, missing_options):
        existing = [dict(option) for option in (field.get('options') or [])]
        colors = ('Blue', 'Green', 'Orange', 'Purple', 'Turquoise')
        for index, name in enumerate(missing_options):
            existing.append({
                'name': name,
                'hue': colors[(len(existing) + index) % len(colors)],
                'lightness': 'Lighter',
            })
        payload = {
            'name': field.get('name') or '采购员',
            'type': 'select',
            'multiple': bool(field.get('multiple', False)),
            'options': existing,
        }
        if field.get('description'):
            payload['description'] = field['description']
        self._run(
            '+field-update',
            ['--field-id', str(field.get('id') or field.get('field_id')
                               or field.get('name') or '采购员')],
            payload=payload, high_risk=True)


def _field_index(fields):
    return {str(field.get('name') or ''): field for field in fields}


def validate_schema(fields):
    actual = _field_index(fields)
    bad = [name for name, expected in REQUIRED_FIELD_TYPES.items()
           if (actual.get(name) or {}).get('type') != expected]
    if bad:
        raise LedgerBackfillError('飞书台账字段缺失或类型不匹配：%s' %
                                  ', '.join(bad))
    return actual


def ensure_buyer_options(client, fields, buyers, allow_create=False):
    actual = _field_index(fields)
    field = actual['采购员']
    options = {str(item.get('name') or '')
               for item in (field.get('options') or [])}
    missing = sorted(set(buyers) - options)
    if not missing:
        return []
    if not allow_create:
        raise LedgerBackfillError(
            '采购员选项不存在：%s；需先补建或显式传 '
            '--create-missing-buyer-options' % ', '.join(missing))
    client.update_select_options(field, missing)
    return missing


def _normalize_user(value):
    if not isinstance(value, list):
        return []
    return sorted(str(item.get('id') or '') for item in value
                  if isinstance(item, dict) and item.get('id'))


def _same_binding(record, target):
    try:
        serial = int(record.get('环境序号'))
    except (TypeError, ValueError):
        serial = None
    return (
        str(record.get('账号状态') or '') == target['账号状态']
        and str(record.get('绑定环境') or '') == target['绑定环境']
        and serial == target['环境序号']
        and str(record.get('绑定时间') or '') == target['绑定时间']
        and str(record.get('采购员') or '') == target['采购员']
        and _normalize_user(record.get('操作人')) ==
        _normalize_user(target['操作人'])
    )


def _record_index(records):
    result = {}
    for record in records:
        email = str(record.get('邮箱账号') or '').strip()
        if not email:
            continue
        key = email.casefold()
        if key in result:
            raise LedgerBackfillError('台账存在重复邮箱记录：%s' %
                                      mask_email(email))
        result[key] = record
    return result


class LedgerBackfillService(object):
    def __init__(self, client, operator_open_id=DEFAULT_OPERATOR_OPEN_ID,
                 remark='模块三补账'):
        self.client = client
        self.operator_open_id = operator_open_id
        self.remark = remark

    @staticmethod
    def _validate_inputs(accounts, mapping_rows):
        account_keys = {account.email.casefold() for account in accounts}
        mapping_keys = {row.email.casefold() for row in mapping_rows}
        if account_keys != mapping_keys:
            missing_mapping = account_keys - mapping_keys
            extra_mapping = mapping_keys - account_keys
            detail = []
            if missing_mapping:
                detail.append('映射缺 %d 行' % len(missing_mapping))
            if extra_mapping:
                detail.append('映射多 %d 行' % len(extra_mapping))
            raise LedgerBackfillError('原始 xlsx 与映射清单邮箱集合不一致：%s' %
                                      '，'.join(detail))
        mapping = {row.email.casefold(): row for row in mapping_rows}
        for account in accounts:
            row = mapping[account.email.casefold()]
            if row.buyer:
                account.buyer = row.buyer
        return mapping

    def dry_run(self, accounts, mapping_rows):
        mapping = self._validate_inputs(accounts, mapping_rows)
        return {
            'total': len(accounts),
            'complete': sum(row.complete for row in mapping.values()),
            'failed': sum(not row.complete for row in mapping.values()),
            'buyers': sorted(set(row.buyer for row in mapping.values()
                                 if row.buyer)),
        }

    def apply(self, accounts, mapping_rows, create_missing_buyer_options=False):
        mapping = self._validate_inputs(accounts, mapping_rows)
        fields = self.client.list_fields()
        validate_schema(fields)
        ensure_buyer_options(
            self.client, fields,
            [row.buyer for row in mapping.values() if row.buyer],
            allow_create=create_missing_buyer_options)

        records = self.client.list_records()
        record_by_email = _record_index(records)
        missing_accounts = [account for account in accounts
                            if account.email.casefold() not in record_by_email]
        unique_password = len({account.password for account in accounts}) == 1
        remark = self.remark + ('，统一密码' if unique_password else '')
        create_records = []
        for account in missing_accounts:
            row = mapping[account.email.casefold()]
            create_records.append({
                '邮箱账号': account.email,
                '密码': account.password,
                '接码Key链接': account.key_url,
                '号商购买单号': account.order_no,
                '账号状态': '未绑定',
                '采购员': row.buyer,
                'Cookie': account.cookie_text,
                '备注': remark,
            })
        record_ids = self.client.batch_create(create_records) if create_records else []
        for account, record_id in zip(missing_accounts, record_ids):
            record_by_email[account.email.casefold()] = {
                '_record_id': record_id,
                '邮箱账号': account.email,
                '账号状态': '未绑定',
                '采购员': mapping[account.email.casefold()].buyer,
            }

        updates = {}
        skipped_complete = 0
        for account in accounts:
            row = mapping[account.email.casefold()]
            if not row.complete:
                continue
            record = record_by_email[account.email.casefold()]
            target = {
                '账号状态': '已绑定',
                '绑定环境': row.env_name,
                '环境序号': int(row.serial_number),
                '绑定时间': row.binding_time,
                '采购员': row.buyer,
                '操作人': [{'id': self.operator_open_id}],
            }
            if _same_binding(record, target):
                skipped_complete += 1
                continue
            updates[record['_record_id']] = target
        if updates:
            self.client.batch_update(updates)

        # 只读回邮箱和绑定字段，确认创建/更新真正落地。
        readback = _record_index(self.client.list_records())
        for account in accounts:
            row = mapping[account.email.casefold()]
            record = readback.get(account.email.casefold())
            if not record:
                raise LedgerBackfillError('补账读回缺少记录：%s' % account.safe_email)
            if row.complete:
                target = {
                    '账号状态': '已绑定', '绑定环境': row.env_name,
                    '环境序号': int(row.serial_number),
                    '绑定时间': row.binding_time, '采购员': row.buyer,
                    '操作人': [{'id': self.operator_open_id}],
                }
                if not _same_binding(record, target):
                    raise LedgerBackfillError('补账读回字段不一致：%s' %
                                              account.safe_email)
        return {
            'total': len(accounts),
            'created': len(create_records),
            'existing': len(accounts) - len(create_records),
            'updated': len(updates),
            'skippedComplete': skipped_complete,
            'failedRows': sum(not row.complete for row in mapping.values()),
        }


def build_parser():
    parser = argparse.ArgumentParser(
        prog='python -m purchase_tool backfill',
        description='模块三 Mac 补账：默认 dry-run，--apply 才写飞书')
    parser.add_argument('--xlsx', required=True, help='仓库外原始号商 xlsx')
    parser.add_argument('--mapping', required=True, help='仓库外绑定映射清单')
    parser.add_argument('--operator-open-id', default=DEFAULT_OPERATOR_OPEN_ID)
    parser.add_argument('--remark', default='模块三补账')
    parser.add_argument('--profile')
    parser.add_argument('--apply', action='store_true')
    parser.add_argument('--create-missing-buyer-options', action='store_true',
                        help='显式允许全量 PUT 补建缺失的采购员选项')
    return parser


def main(argv=None):
    args = build_parser().parse_args(argv)
    try:
        original_path = validate_input_path(
            args.xlsx, allow_repo_sample=not args.apply)
        mapping_path = validate_input_path(
            args.mapping, allow_repo_sample=not args.apply)
        with original_path.open('rb') as handle:
            accounts = parse_vendor_workbook(BytesIO(handle.read()))
        mapping_rows = parse_mapping_workbook(str(mapping_path))
        service = LedgerBackfillService(
            LarkBaseClient(profile=args.profile),
            operator_open_id=args.operator_open_id,
            remark=args.remark)
        if not args.apply:
            summary = service.dry_run(accounts, mapping_rows)
            print('补账 dry-run：总数 %(total)d，完成映射 %(complete)d，失败映射 %(failed)d' % summary)
            print('未连接飞书、未写入任何数据；采购员：%s' %
                  ','.join(summary['buyers']))
            return 0
        if not args.operator_open_id.startswith('ou_'):
            raise LedgerBackfillError('操作人必须是已确认的 ou_ open_id')
        summary = service.apply(
            accounts, mapping_rows,
            create_missing_buyer_options=args.create_missing_buyer_options)
        print('补账完成：新增 %(created)d，已有 %(existing)d，更新 %(updated)d，'
              '已闭环跳过 %(skippedComplete)d，失败映射 %(failedRows)d' % summary)
        return 2 if summary['failedRows'] else 0
    except (EnvBatchError, LedgerBackfillError) as exc:
        print('补账失败：%s' % scrub_text(exc))
        return 2
    except Exception as exc:
        print('补账失败：%s' % scrub_text(exc))
        return 2


if __name__ == '__main__':
    sys.exit(main())
