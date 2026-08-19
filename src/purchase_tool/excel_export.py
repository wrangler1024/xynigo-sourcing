# -*- coding: utf-8 -*-
"""物流查询 Excel/CSV 导出；JPEG 截图无需 Pillow 即可内嵌。"""
import time


EXPORT_HEAD = ['环境序号', '环境名', '订单号', '下单时间', '金额', '状态',
               '物流单号', '包裹号', '承运商', '物流轨迹截图', '出口IP',
               '结果', '失败原因', '查询时间（墨西哥）']

STATE_CN = {'ok': '成功', 'login': '登录失效', 'inuse': '环境使用中，已跳过',
            'fail': '失败', 'running': '查询中', 'pending': '未查询',
            'stopped': '已停止'}


def _screenshot_text(row):
    state = row.get('screenshotState')
    if state == 'ok':
        return '查看截图'
    if state == 'fail':
        return '截图失败：%s' % (row.get('screenshotError') or '未知原因')
    if row.get('tracks'):
        return '未生成'
    return '—（暂无物流）'


def export_bytes(rows, fmt, screenshot_reader=None):
    """生成导出文件，返回 (bytes, filename, mime)。"""
    stamp = time.strftime('%Y%m%d_%H%M')
    lines = []
    for r in rows:
        state = r['state']
        if state == 'ok':
            result = '成功（砍单退款中）' if r['kanDan'] else '成功'
        else:
            result = STATE_CN.get(state, state)
        lines.append([
            r['serial'], r['envName'], r['orderNo'], r['orderTime'],
            r['amount'],
            (r['status'] + ' ' + r['statusCn']).strip(),
            '; '.join(r['tracks']) or ('—（砍单退款，无物流）'
                                       if r['kanDan'] else ''),
            '; '.join(r['pkgs']), r['carrier'], _screenshot_text(r),
            r['ip'], result, r['error'], r['time']])
    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.drawing.spreadsheet_drawing import (
                AnchorMarker, TwoCellAnchor)
            from openpyxl.styles import (
                Alignment, Border, Font, PatternFill, Side)
            from io import BytesIO

            class EmbeddedJpeg(OpenpyxlImage):
                """无需 Pillow 的 JPEG 嵌入对象（截图尺寸由 CDP 提供）。"""

                def __init__(self, data, width, height):
                    self.ref = BytesIO(data)
                    self.width = width
                    self.height = height
                    self.format = 'jpeg'
                    self.anchor = 'A1'

                def _data(self):
                    return self.ref.getvalue()

            wb = Workbook()
            ws = wb.active
            ws.title = '物流单号查询'
            ws.append(EXPORT_HEAD)
            grid_side = Side(style='thin', color='B7C9E2')
            grid_border = Border(
                left=grid_side, right=grid_side,
                top=grid_side, bottom=grid_side)
            for row_index, (row, source) in enumerate(
                    zip(lines, rows), start=2):
                ws.append(row)
                if (source.get('screenshotState') == 'ok'
                        and screenshot_reader is not None):
                    image_data = screenshot_reader(source['serial'])
                    if image_data:
                        source_width = max(
                            1, int(source.get('screenshotWidth') or 1000))
                        source_height = max(
                            1, int(source.get('screenshotHeight') or 600))
                        # 主表直接放入缩略图。TwoCellAnchor 将图片约束在
                        # J 列当前行内，调整行列时会随单元格移动和缩放。
                        thumb_width = min(160, source_width)
                        thumb_height = max(
                            1, int(source_height * thumb_width /
                                   float(source_width)))
                        thumb = EmbeddedJpeg(
                            image_data, thumb_width, thumb_height)
                        thumb.anchor = TwoCellAnchor(
                            editAs='twoCell',
                            _from=AnchorMarker(col=9, row=row_index - 1),
                            to=AnchorMarker(col=10, row=row_index))
                        ws.add_image(thumb)
                        ws.row_dimensions[row_index].height = max(
                            24, thumb_height * 0.75)
                        cell = ws.cell(row_index, 10)
                        cell.value = ''
            header = ws[1]
            for cell in header:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='1F4E78')
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = grid_border
            widths = [10, 24, 22, 20, 14, 18, 28, 22, 14, 24,
                      16, 18, 32, 12]
            for index, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + index)].width = width
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = 'A1:N%s' % max(1, len(lines) + 1)
            ws.sheet_view.showGridLines = False
            for row in ws.iter_rows(min_row=2, max_row=len(lines) + 1):
                for cell in row:
                    cell.alignment = Alignment(
                        vertical='center', wrap_text=True)
                    cell.border = grid_border
                for index in (0, 2, 6, 7, 10):
                    row[index].number_format = '@'
                row[9].alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
            buf = BytesIO()
            wb.save(buf)
            return (buf.getvalue(),
                    '物流单号查询结果_%s.xlsx' % stamp,
                    'application/vnd.openxmlformats-officedocument'
                    '.spreadsheetml.sheet')
        except ImportError:
            fmt = 'csv'   # 无 openpyxl 时降级 CSV
    buf = []
    for row in lines:
        buf.append(','.join(
            '"%s"' % str(c).replace('"', '""') if ',' in str(c) or '"' in str(c)
            else str(c) for c in row))
    data = ('\ufeff' + '\r\n'.join(
        [','.join(EXPORT_HEAD)] + buf)).encode('utf-8')
    return data, '物流单号查询结果_%s.csv' % stamp, 'text/csv'
